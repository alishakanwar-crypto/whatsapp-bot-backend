"""
Law Minister Office Attendance Bot — auto-response handler.

Routes incoming WhatsApp messages for the Law Minister phone number
(+91 84489 43232, Phone Number ID: 1168433719678061) and sends
automated responses based on greeting/unrelated/office-topic classification.

Uses the same WABA token as the PPIS bot but a different phone number ID.
"""

import logging
import os
import re
import tempfile
from datetime import datetime, timezone, timedelta

import httpx

logger = logging.getLogger("law_minister_bot")

# Law Minister WhatsApp Phone Number ID
LAW_MINISTER_PHONE_ID = os.getenv(
    "LAW_MINISTER_PHONE_ID", "1168433719678061"
)

# WABA ID (shared with PPIS number)
_WABA_ID = os.getenv("LAW_MINISTER_WABA_ID", "2417647228700804")

# ---------- Admin Configuration ----------
# Admins: phone numbers (without country code prefix) who have elevated access.
# They can request message summaries via "summary" / "report" commands.
ADMINS = {
    "918796105084": "Ali",
}

# Token is shared across the WABA — reuse the same Cloud token
def _get_token() -> str:
    return os.getenv("WHATSAPP_CLOUD_TOKEN", "")


# ---------- Auto-Registration & Webhook Keep-Alive ----------

async def ensure_webhook_registration() -> dict:
    """Re-register the Law Minister phone number and subscribe the WABA.

    Meta periodically drops webhook subscriptions for secondary phone
    numbers.  This function is called on app startup and every 2 hours
    by the scheduler to keep the subscription alive.

    Returns a dict with the results of each API call.
    """
    token = _get_token()
    if not token:
        logger.error("LM REGISTRATION: WHATSAPP_CLOUD_TOKEN not set — skipping")
        return {"error": "no_token"}

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    results = {}

    async with httpx.AsyncClient(timeout=15.0) as client:
        # 1. Re-register the phone number (keeps Cloud API active)
        try:
            resp = await client.post(
                f"https://graph.facebook.com/v21.0/{LAW_MINISTER_PHONE_ID}/register",
                json={"messaging_product": "whatsapp", "pin": "123456"},
                headers=headers,
            )
            results["register"] = resp.json()
            logger.info(
                f"LM REGISTRATION: register phone → {resp.status_code}: {resp.text}"
            )
        except Exception as e:
            results["register"] = {"error": str(e)}
            logger.error(f"LM REGISTRATION: register failed: {e}")

        # 2. Subscribe WABA to app (ensures webhooks flow for all numbers)
        try:
            resp = await client.post(
                f"https://graph.facebook.com/v21.0/{_WABA_ID}/subscribed_apps",
                headers=headers,
            )
            results["subscribe_waba"] = resp.json()
            logger.info(
                f"LM REGISTRATION: subscribe WABA → {resp.status_code}: {resp.text}"
            )
        except Exception as e:
            results["subscribe_waba"] = {"error": str(e)}
            logger.error(f"LM REGISTRATION: WABA subscribe failed: {e}")

    return results


def ensure_webhook_registration_sync() -> None:
    """Synchronous wrapper for the scheduler (APScheduler runs sync jobs)."""
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as pool:
                pool.submit(lambda: asyncio.run(ensure_webhook_registration())).result(timeout=30)
        else:
            loop.run_until_complete(ensure_webhook_registration())
    except Exception as e:
        logger.error(f"LM REGISTRATION sync wrapper failed: {e}")


# ---------- Language Detection ----------

_DEVANAGARI_RE = re.compile(r"[\u0900-\u097F]")


def detect_language(text: str) -> str:
    """Detect Hindi, English, or Hinglish from text."""
    has_devanagari = bool(_DEVANAGARI_RE.search(text))
    has_latin = bool(re.search(r"[a-zA-Z]", text))
    if has_devanagari and has_latin:
        return "hinglish"
    if has_devanagari:
        return "hindi"
    return "english"


# ---------- Greeting Detection ----------

GREETING_KEYWORDS = {
    "hi", "hello", "hey", "namaste", "namashkar", "namaskar",
    "good morning", "good afternoon", "good evening",
    "shubh prabhat", "pranam", "jai hind",
    "नमस्ते", "नमस्कार", "प्रणाम", "शुभ प्रभात",
    "जय हिन्द", "जय हिंद",
}

ALLOWED_TOPICS = {
    "attendance", "registration", "face", "photo", "photograph",
    "present", "absent", "check-in", "checkin", "office", "timing",
    "time", "leave", "help", "register", "status", "camera",
    "name", "report", "summary",
    "उपस्थिति", "रजिस्ट्रेशन", "फोटो", "कार्यालय", "समय",
    "छुट्टी", "मदद", "स्थिति", "रिपोर्ट",
}


def _is_greeting(text: str) -> bool:
    normalised = text.strip().lower()
    if normalised in GREETING_KEYWORDS:
        return True
    for kw in GREETING_KEYWORDS:
        if normalised.startswith(kw):
            return True
    return False


# ---------- Response Templates (always bilingual) ----------

GREETING_RESPONSE = (
    "Hello / नमस्कार,\n\n"
    "Welcome to the Face Recognition Attendance System of "
    "Shri Arjun Ram Meghwal Ji.\n"
    "श्री अर्जुन राम मेघवाल जी की फेस रिकग्निशन अटेंडेंस सिस्टम "
    "में आपका स्वागत है।\n\n"
    "I am an automated bot designed only for office attendance "
    "and official work-related communication.\n"
    "मैं एक स्वचालित बॉट हूँ जो केवल कार्यालय उपस्थिति एवं "
    "आधिकारिक कार्य संबंधी संवाद हेतु बनाया गया है।\n\n"
    "Kindly share:\n"
    "• 2 clear front-face photographs for face registration "
    "along with your name.\n"
    "• फेस रजिस्ट्रेशन हेतु 2 स्पष्ट फ्रंट फेस फोटो के साथ "
    "अपना नाम साझा करें।\n\n"
    "Please ensure the photographs are clear, recent, "
    "and properly visible.\n"
    "कृपया सुनिश्चित करें कि फोटो स्पष्ट, हाल की एवं "
    "सही रूप से दिखाई देने वाली हों।\n\n"
    "Thankyou / धन्यवाद"
)

UNRELATED_RESPONSE = (
    "Kindly note that this automated system is restricted to official "
    "office attendance and work-related communication only. / "
    "कृपया ध्यान दें कि यह स्वचालित प्रणाली केवल कार्यालय उपस्थिति "
    "एवं आधिकारिक कार्य संबंधी संवाद हेतु सीमित है।"
)


def _get_response(text: str) -> str | None:
    """Return auto-response for the message, or None for allowed topics."""
    if _is_greeting(text):
        return GREETING_RESPONSE

    normalised = text.strip().lower()
    for topic in ALLOWED_TOPICS:
        if topic in normalised:
            return None  # Allowed topic — no auto-reply needed

    return UNRELATED_RESPONSE


# ---------- Send via Law Minister Phone Number ----------

async def _send_text(to: str, body: str) -> bool:
    """Send a text message from the Law Minister phone number."""
    token = _get_token()
    if not token:
        logger.error("WHATSAPP_CLOUD_TOKEN not set")
        return False

    recipient = to.split("@")[0] if "@" in to else to
    if len(recipient) == 10:
        recipient = "91" + recipient

    url = f"https://graph.facebook.com/v21.0/{LAW_MINISTER_PHONE_ID}/messages"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": recipient,
        "type": "text",
        "text": {"body": body},
    }

    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                url, json=payload, headers=headers, timeout=15.0
            )
            if resp.status_code == 200:
                logger.info(
                    f"Law Minister bot replied to {recipient}: {body[:50]}..."
                )
                return True
            logger.error(
                f"Law Minister bot send error {resp.status_code}: {resp.text}"
            )
            return False
    except Exception as e:
        logger.error(f"Law Minister bot send failed: {e}")
        return False


# ---------- Deduplication ----------

# In-memory set of recently processed message IDs (TTL-based cleanup)
_processed_ids: dict[str, float] = {}
_DEDUP_TTL = 300  # 5 minutes


def _is_duplicate(msg_id: str) -> bool:
    """Check if a message ID was already processed (prevents retry duplicates)."""
    import time
    now = time.time()

    # Clean old entries
    expired = [k for k, t in _processed_ids.items() if now - t > _DEDUP_TTL]
    for k in expired:
        del _processed_ids[k]

    if msg_id in _processed_ids:
        return True
    _processed_ids[msg_id] = now
    return False


# ---------- Message Log (DB) ----------

async def _init_lm_tables():
    """Create Law Minister message log table if not exists."""
    import aiosqlite
    from app.database import DB_PATH
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS lm_message_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                direction TEXT NOT NULL,
                sender TEXT NOT NULL,
                recipient TEXT NOT NULL,
                message_type TEXT NOT NULL DEFAULT 'text',
                content TEXT NOT NULL DEFAULT '',
                category TEXT DEFAULT '',
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.commit()


async def _log_message(direction: str, sender: str, recipient: str,
                       content: str, msg_type: str = "text",
                       category: str = ""):
    """Log a message to the lm_message_log table."""
    import aiosqlite
    from app.database import DB_PATH
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "INSERT INTO lm_message_log (direction, sender, recipient, message_type, content, category) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (direction, sender, recipient, msg_type, content[:500], category),
            )
            await db.commit()
    except Exception as e:
        logger.error(f"Failed to log message: {e}")


# ---------- Excel Summary Generation ----------

async def _generate_summary_excel(days: int = 7) -> str | None:
    """Generate Excel summary of all outbound messages for the last N days.

    Returns the file path of the generated .xlsx file, or None on failure.
    """
    import aiosqlite
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.database import DB_PATH

    try:
        ist = timezone(timedelta(hours=5, minutes=30))
        cutoff = (datetime.now(ist) - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")

        async with aiosqlite.connect(DB_PATH) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(
                "SELECT * FROM lm_message_log WHERE timestamp >= ? ORDER BY timestamp DESC",
                (cutoff,),
            )
            rows = await cursor.fetchall()

        if not rows:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Message Summary"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")

        headers = ["#", "Direction", "From", "To", "Type", "Content", "Category", "Timestamp (IST)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(rows, 2):
            ts_raw = row["timestamp"] or ""
            try:
                ts_dt = datetime.fromisoformat(ts_raw.replace("Z", "+00:00"))
                ts_ist = ts_dt.astimezone(ist).strftime("%d/%m/%Y %I:%M %p")
            except Exception:
                ts_ist = ts_raw

            ws.cell(row=i, column=1, value=i - 1)
            ws.cell(row=i, column=2, value=row["direction"])
            ws.cell(row=i, column=3, value=row["sender"])
            ws.cell(row=i, column=4, value=row["recipient"])
            ws.cell(row=i, column=5, value=row["message_type"])
            ws.cell(row=i, column=6, value=row["content"][:200])
            ws.cell(row=i, column=7, value=row["category"])
            ws.cell(row=i, column=8, value=ts_ist)

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        filepath = tempfile.mktemp(suffix=".xlsx", prefix="lm_summary_")
        wb.save(filepath)
        return filepath

    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        return None


async def _send_document(to: str, filepath: str, caption: str) -> bool:
    """Send a document (Excel file) via WhatsApp Cloud API."""
    token = _get_token()
    if not token:
        return False

    recipient = to.split("@")[0] if "@" in to else to
    if len(recipient) == 10:
        recipient = "91" + recipient

    headers = {"Authorization": f"Bearer {token}"}

    # Step 1: Upload the file to Meta
    upload_url = f"https://graph.facebook.com/v21.0/{LAW_MINISTER_PHONE_ID}/media"
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            with open(filepath, "rb") as f:
                resp = await client.post(
                    upload_url,
                    headers=headers,
                    data={"messaging_product": "whatsapp", "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                    files={"file": ("message_summary.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
            if resp.status_code != 200:
                logger.error(f"Media upload failed: {resp.status_code} {resp.text}")
                return False
            media_id = resp.json().get("id")

            # Step 2: Send document message
            msg_url = f"https://graph.facebook.com/v21.0/{LAW_MINISTER_PHONE_ID}/messages"
            payload = {
                "messaging_product": "whatsapp",
                "to": recipient,
                "type": "document",
                "document": {
                    "id": media_id,
                    "caption": caption,
                    "filename": "message_summary.xlsx",
                },
            }
            resp = await client.post(
                msg_url,
                json=payload,
                headers={**headers, "Content-Type": "application/json"},
                timeout=15.0,
            )
            if resp.status_code == 200:
                logger.info(f"Sent Excel summary to {recipient}")
                return True
            logger.error(f"Document send failed: {resp.status_code} {resp.text}")
            return False
    except Exception as e:
        logger.error(f"Send document error: {e}")
        return False


# ---------- Admin Command Handling ----------

async def _handle_admin_command(sender: str, text: str) -> str | None:
    """Handle admin-specific commands. Returns response text or None."""
    normalised = text.strip().lower()

    # Summary / Report command
    if normalised in ("summary", "report", "excel", "log", "messages"):
        admin_name = ADMINS.get(sender, "Admin")
        await _send_text(sender, f"Generating message summary for you, {admin_name}... Please wait.")

        filepath = await _generate_summary_excel(days=7)
        if filepath:
            ist = timezone(timedelta(hours=5, minutes=30))
            now_str = datetime.now(ist).strftime("%d/%m/%Y %I:%M %p")
            caption = f"Law Minister Bot — Message Summary\nGenerated: {now_str}\nPeriod: Last 7 days"
            sent = await _send_document(sender, filepath, caption)
            # Clean up temp file
            try:
                os.unlink(filepath)
            except Exception:
                pass
            if sent:
                return None  # Already sent the document
            return "Failed to send the Excel file. Please try again."
        else:
            return "No messages found in the last 7 days."

    # Help command for admin
    if normalised in ("admin help", "admin", "commands"):
        return (
            "Admin Commands:\n"
            "• *summary* / *report* — Get Excel summary of all bot messages (last 7 days)\n"
            "• *admin help* — Show this help menu"
        )

    return None  # Not an admin command


# ---------- Main Handler ----------

async def handle_webhook(body: dict) -> dict:
    """Process an incoming webhook payload for the Law Minister phone number.

    Returns a dict with status and actions taken.
    """
    # Ensure message log table exists
    await _init_lm_tables()

    actions = []

    for entry in body.get("entry", []):
        for change in entry.get("changes", []):
            value = change.get("value", {})
            for message in value.get("messages", []):
                if message.get("type") != "text":
                    continue

                sender = message.get("from", "")
                text = message.get("text", {}).get("body", "")
                msg_id = message.get("id", "")

                if msg_id and _is_duplicate(msg_id):
                    logger.info(f"Duplicate message {msg_id}, skipping.")
                    continue

                # Log incoming message
                await _log_message(
                    direction="incoming",
                    sender=sender,
                    recipient=LAW_MINISTER_PHONE_ID,
                    content=text,
                    category="incoming",
                )

                # Check if sender is admin and handle admin commands
                if sender in ADMINS:
                    admin_response = await _handle_admin_command(sender, text)
                    if admin_response is not None:
                        sent = await _send_text(sender, admin_response)
                        await _log_message(
                            direction="outgoing",
                            sender=LAW_MINISTER_PHONE_ID,
                            recipient=sender,
                            content=admin_response[:500],
                            category="admin_response",
                        )
                        actions.append({
                            "from": sender,
                            "text": text,
                            "category": "admin_command",
                            "response_sent": sent,
                        })
                        continue
                    elif text.strip().lower() in ("summary", "report", "excel", "log", "messages"):
                        # Admin command handled (excel sent directly)
                        actions.append({
                            "from": sender,
                            "text": text,
                            "category": "admin_command",
                            "response_sent": True,
                        })
                        continue

                auto_reply = _get_response(text)
                if auto_reply:
                    sent = await _send_text(sender, auto_reply)
                    lang = detect_language(text)
                    category = "greeting" if _is_greeting(text) else "unrelated"

                    # Log outgoing response
                    await _log_message(
                        direction="outgoing",
                        sender=LAW_MINISTER_PHONE_ID,
                        recipient=sender,
                        content=auto_reply[:500],
                        category=category,
                    )

                    actions.append({
                        "from": sender,
                        "text": text,
                        "category": category,
                        "language": lang,
                        "response_sent": sent,
                    })
                    logger.info(
                        f"Auto-replied to {sender} ({category}/{lang}): "
                        f"{text[:40]}"
                    )
                else:
                    actions.append({
                        "from": sender,
                        "text": text,
                        "category": "allowed",
                        "language": detect_language(text),
                        "response_sent": False,
                        "note": "Office topic — no auto-reply",
                    })

    return {"status": "ok", "bot": "law_minister", "actions": actions}
