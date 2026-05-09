"""Generate and maintain a daily teacher attendance Excel workbook.

The workbook lives at /data/teacher_attendance.xlsx and is updated after
each attendance window closes (9:30 AM IST).  Each calendar month gets its
own worksheet.  Rows = teachers, columns = days of the month.

Download via GET /api/dashboard/teacher-attendance-excel
"""

import asyncio
import logging
import os
from datetime import datetime, timezone, timedelta, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db

logger = logging.getLogger(__name__)

IST = timezone(timedelta(hours=5, minutes=30))

# Persistent path — survives deploys on Fly.io (volume-mounted /data)
EXCEL_DIR = Path("/data")
EXCEL_PATH = EXCEL_DIR / "teacher_attendance.xlsx"

# Fallback for local dev
if not EXCEL_DIR.exists():
    EXCEL_DIR = Path(os.path.dirname(__file__)).parent.parent
    EXCEL_PATH = EXCEL_DIR / "teacher_attendance.xlsx"

# Styling constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
PRESENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PRESENT_FONT = Font(name="Calibri", color="006100", size=10)
ABSENT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ABSENT_FONT = Font(name="Calibri", color="9C0006", size=10)
HOLIDAY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HOLIDAY_FONT = Font(name="Calibri", color="4472C4", italic=True, size=10)
NAME_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _month_sheet_name(dt: date) -> str:
    """e.g. 'May 2026'"""
    return dt.strftime("%B %Y")


async def generate_teacher_attendance_excel(target_date: date | None = None) -> str:
    """Rebuild / update the Excel workbook for the given date (default: today IST).

    Returns the path to the saved Excel file.
    """
    if target_date is None:
        target_date = datetime.now(IST).date()

    sheet_name = _month_sheet_name(target_date)
    days_in_month = (date(target_date.year, target_date.month % 12 + 1, 1)
                     - timedelta(days=1)).day if target_date.month < 12 else 31

    db = await get_db()
    try:
        # --- Fetch all registered teachers ---
        cursor = await db.execute(
            "SELECT DISTINCT person_id, name, phone FROM agent_registered_faces "
            "WHERE person_id LIKE 'TEACHER_%' ORDER BY name"
        )
        teachers = await cursor.fetchall()

        if not teachers:
            logger.warning("TEACHER EXCEL: No teachers found in face DB")
            return str(EXCEL_PATH)

        # De-duplicate by person_id (may have multiple face angles)
        teacher_map: dict[str, dict] = {}
        for row in teachers:
            pid = row[0]
            if pid not in teacher_map:
                teacher_map[pid] = {"name": row[1], "phone": row[2] or ""}

        teacher_list = sorted(teacher_map.items(), key=lambda x: x[1]["name"])

        # --- Fetch attendance records for this month ---
        month_start = date(target_date.year, target_date.month, 1).isoformat()
        if target_date.month < 12:
            month_end = date(target_date.year, target_date.month + 1, 1).isoformat()
        else:
            month_end = date(target_date.year + 1, 1, 1).isoformat()

        cursor = await db.execute(
            "SELECT person_id, student_name, logged_at, confidence, camera_label "
            "FROM attendance_records "
            "WHERE person_id LIKE 'TEACHER_%' "
            "AND logged_at >= ? AND logged_at < ? "
            "ORDER BY logged_at",
            (month_start, month_end),
        )
        records = await cursor.fetchall()

        # Build: person_id -> {day_num -> {time, confidence, camera}}
        attendance_data: dict[str, dict[int, dict]] = {}
        for rec in records:
            pid = rec[0]
            logged_at_str = rec[2]
            try:
                logged_dt = datetime.fromisoformat(logged_at_str)
                day_num = logged_dt.day
            except (ValueError, TypeError):
                continue
            if pid not in attendance_data:
                attendance_data[pid] = {}
            # Keep earliest detection for each day
            if day_num not in attendance_data[pid]:
                attendance_data[pid][day_num] = {
                    "time": logged_dt.strftime("%I:%M %p"),
                    "confidence": rec[3],
                    "camera": rec[4] or "",
                }

        # --- Fetch holidays for this month ---
        cursor = await db.execute(
            "SELECT date, reason FROM school_holidays "
            "WHERE date >= ? AND date < ?",
            (month_start, month_end),
        )
        holidays_rows = await cursor.fetchall()
        holidays: dict[int, str] = {}
        for h in holidays_rows:
            try:
                h_date = date.fromisoformat(h[0])
                holidays[h_date.day] = h[1] or "Holiday"
            except (ValueError, TypeError):
                continue

    finally:
        await db.close()

    # --- Build or load workbook ---
    if EXCEL_PATH.exists():
        try:
            wb = load_workbook(str(EXCEL_PATH))
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Create or replace month sheet (delete + recreate to avoid merged-cell issues)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    # Remove default sheet if this is the first real sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Write headers ---
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=3 + days_in_month)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"PP International School — Teacher Attendance — {sheet_name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    # Row 2: Generated timestamp
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=3 + days_in_month)
    ts_cell = ws.cell(row=2, column=1)
    ts_cell.value = f"Last updated: {datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}"
    ts_cell.font = Font(name="Calibri", italic=True, size=9, color="808080")
    ts_cell.alignment = Alignment(horizontal="center")

    # Row 3: Column headers
    header_row = 3
    headers = ["S.No.", "Teacher Name", "Phone"]
    for day in range(1, days_in_month + 1):
        dt = date(target_date.year, target_date.month, day)
        day_name = dt.strftime("%a")  # Mon, Tue, etc.
        headers.append(f"{day}\n{day_name}")
    headers.append("Total Present")
    headers.append("Total Absent")
    headers.append("Attendance %")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN_BORDER

    # --- Write teacher rows ---
    data_start_row = 4
    for idx, (pid, tinfo) in enumerate(teacher_list):
        row = data_start_row + idx

        # S.No.
        cell = ws.cell(row=row, column=1)
        cell.value = idx + 1
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Teacher name
        cell = ws.cell(row=row, column=2)
        cell.value = tinfo["name"]
        cell.font = NAME_FONT
        cell.border = THIN_BORDER

        # Phone
        cell = ws.cell(row=row, column=3)
        cell.value = tinfo["phone"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        total_present = 0
        working_days = 0

        for day in range(1, days_in_month + 1):
            col = 3 + day
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            dt = date(target_date.year, target_date.month, day)

            # Future dates: leave blank
            if dt > target_date:
                cell.value = ""
                continue

            # Sunday
            if dt.weekday() == 6:  # Sunday
                cell.value = "SUN"
                cell.fill = HOLIDAY_FILL
                cell.font = HOLIDAY_FONT
                continue

            # Holiday
            if day in holidays:
                cell.value = "H"
                cell.fill = HOLIDAY_FILL
                cell.font = HOLIDAY_FONT
                continue

            # Only 2nd Saturday of the month is off
            if dt.weekday() == 5:  # Saturday
                # Which Saturday of the month? (1st, 2nd, 3rd, 4th, 5th)
                saturday_number = (day - 1) // 7 + 1
                if saturday_number == 2:
                    cell.value = "2nd SAT"
                    cell.fill = HOLIDAY_FILL
                    cell.font = HOLIDAY_FONT
                    continue
                # All other Saturdays are working days — fall through

            # Working day
            working_days += 1

            att = attendance_data.get(pid, {}).get(day)
            if att:
                cell.value = f"P\n{att['time']}"
                cell.fill = PRESENT_FILL
                cell.font = PRESENT_FONT
                total_present += 1
            else:
                cell.value = "A"
                cell.fill = ABSENT_FILL
                cell.font = ABSENT_FONT

        # Summary columns
        total_absent = working_days - total_present
        pct = (total_present / working_days * 100) if working_days > 0 else 0

        summary_col = 3 + days_in_month + 1
        for scol, val in [(summary_col, total_present),
                          (summary_col + 1, total_absent),
                          (summary_col + 2, f"{pct:.1f}%")]:
            cell = ws.cell(row=row, column=scol)
            cell.value = val
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", bold=True, size=10)

    # --- Column widths ---
    ws.column_dimensions[get_column_letter(1)].width = 6   # S.No.
    ws.column_dimensions[get_column_letter(2)].width = 25  # Name
    ws.column_dimensions[get_column_letter(3)].width = 15  # Phone
    for day in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(3 + day)].width = 10
    for extra in range(1, 4):
        ws.column_dimensions[get_column_letter(3 + days_in_month + extra)].width = 12

    # Row height for header
    ws.row_dimensions[header_row].height = 35
    # Row heights for data rows with time
    for r in range(data_start_row, data_start_row + len(teacher_list)):
        ws.row_dimensions[r].height = 30

    # Freeze panes: freeze teacher name columns + header row
    ws.freeze_panes = "D4"

    # --- Save ---
    wb.save(str(EXCEL_PATH))
    logger.info(f"TEACHER EXCEL: Saved {len(teacher_list)} teachers × "
                f"{days_in_month} days to {EXCEL_PATH}")
    return str(EXCEL_PATH)


def generate_teacher_attendance_excel_sync() -> None:
    """Synchronous wrapper for the scheduler."""
    loop = asyncio.new_event_loop()
    try:
        loop.run_until_complete(generate_teacher_attendance_excel())
    except Exception as e:
        logger.error(f"TEACHER EXCEL: Scheduled generation failed: {e}", exc_info=True)
    finally:
        loop.close()
