"""Dashboard API routes for the PPIS School Command Center web app."""

import io
import logging
import os
from datetime import datetime, timedelta, date, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, Header, HTTPException, Query, Request
from fastapi.responses import FileResponse, StreamingResponse

from app.database import get_db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])

# ---------------------------------------------------------------------------
# Authentication: dashboard write operations require AGENT_SECRET
# ---------------------------------------------------------------------------
AGENT_SECRET = os.environ.get("AGENT_SECRET", "")


async def verify_dashboard_secret(x_agent_secret: str = Header("")) -> None:
    """Require AGENT_SECRET for write/delete operations on dashboard data."""
    if not AGENT_SECRET:
        return
    if x_agent_secret != AGENT_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized")


# ── Attendance ───────────────────────────────────────────────────────────────

@router.get("/attendance/today")
async def attendance_today():
    """Get today's attendance summary."""
    db = await get_db()
    try:
        # Total present today
        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        row = await cursor.fetchone()
        present_today = row[0] if row else 0

        # Total registered students
        cursor = await db.execute("SELECT COUNT(DISTINCT person_id) FROM agent_registered_faces")
        row = await cursor.fetchone()
        registered = row[0] if row else 0

        # Total students in PI sheet
        cursor = await db.execute("SELECT COUNT(*) FROM pi_sheet_students")
        row = await cursor.fetchone()
        total_students = row[0] if row else 0

        # Grade-wise breakdown
        cursor = await db.execute(
            "SELECT grade, COUNT(DISTINCT person_id) as count "
            "FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes') "
            "GROUP BY grade ORDER BY grade"
        )
        grades = [{"grade": r[0], "count": r[1]} for r in await cursor.fetchall()]

        # Recent attendance entries
        cursor = await db.execute(
            "SELECT person_id, student_name, grade, camera_label, confidence, "
            "notification_sent, logged_at "
            "FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes') "
            "ORDER BY logged_at DESC LIMIT 500"
        )
        recent = [
            {
                "person_id": r[0], "name": r[1], "grade": r[2],
                "camera": r[3], "confidence": round(r[4] * 100, 1),
                "notified": bool(r[5]), "time": r[6],
            }
            for r in await cursor.fetchall()
        ]

        return {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "present_today": present_today,
            "registered_faces": registered,
            "total_students": total_students,
            "grade_breakdown": grades,
            "recent_entries": recent,
        }
    finally:
        await db.close()


@router.get("/attendance/history")
async def attendance_history(
    days: int = Query(7, ge=1, le=90),
):
    """Get attendance counts per day for the last N days."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT date(logged_at) as day, COUNT(DISTINCT person_id) as count "
            "FROM attendance_records "
            "WHERE logged_at >= datetime('now', ?) "
            "GROUP BY day ORDER BY day",
            (f"-{days} days",),
        )
        rows = await cursor.fetchall()
        return {"days": [{"date": r[0], "count": r[1]} for r in rows]}
    finally:
        await db.close()


MINIMUM_CONFIDENCE = 0.30  # Backend rejects anything below 30%
ATTENDANCE_WINDOW_START = 7  # 7:00 AM IST
ATTENDANCE_WINDOW_END_HOUR = 9  # 9:00 AM IST
ATTENDANCE_WINDOW_END_MIN = 0  # Matches student phase end


def _attendance_notif_name(person_id: str, name: str) -> str:
    """Format notification greeting: title-case ALL-CAPS names."""
    display = name.title() if name == name.upper() else name
    if person_id.startswith("TEACHER_"):
        return display  # Just the name — template has "Dear {{1}}, you have been"
    return f"{display} has been"


def _attendance_template(person_id: str) -> tuple[str, str]:
    """Return (template_name, language_code) for attendance notifications."""
    if person_id.startswith("TEACHER_"):
        return "ppis_teacher_present_text", "en"
    return "ppis_attendance_alert", "en"


@router.post("/attendance/report")
async def report_attendance(request: Request):
    """Receive attendance records from the campus agent.

    Validates: off-days, holidays, attendance window, confidence floor.
    Logs all rejected records to audit_log for forensics.
    """
    from datetime import timezone, timedelta
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    day_name = now_ist.strftime("%A")

    # Block entirely on Sundays (no one works)
    if day_name == "Sunday":
        return {"status": "blocked", "reason": "Sunday — school is closed", "inserted": 0}
    # Block entirely on 2nd Saturday (no one works)
    is_saturday = day_name == "Saturday"
    if is_saturday:
        sat_number = (now_ist.day - 1) // 7 + 1
        if sat_number == 2:
            return {"status": "blocked", "reason": "2nd Saturday — school is closed", "inserted": 0}
    # Flag: students are blocked on ALL Saturdays (teachers may still work)
    _block_students_today = is_saturday
    # Block on holidays
    today_str = now_ist.strftime("%Y-%m-%d")
    try:
        _hdb = await get_db()
        try:
            _hcur = await _hdb.execute(
                "SELECT reason FROM school_holidays WHERE date = ?", (today_str,))
            _hrow = await _hcur.fetchone()
            if _hrow:
                return {"status": "blocked", "reason": f"Holiday: {_hrow[0]}", "inserted": 0}
        finally:
            await _hdb.close()
    except Exception:
        pass

    # Validate attendance time window (7:00 AM - 9:00 AM IST)
    window_start = now_ist.replace(hour=ATTENDANCE_WINDOW_START, minute=0, second=0, microsecond=0)
    window_end = now_ist.replace(hour=ATTENDANCE_WINDOW_END_HOUR, minute=ATTENDANCE_WINDOW_END_MIN, second=0, microsecond=0)
    if not (window_start <= now_ist <= window_end):
        return {
            "status": "blocked",
            "reason": f"Outside attendance window (7:00-9:00 AM IST, current: {now_ist.strftime('%I:%M %p')})",
            "inserted": 0,
        }

    body = await request.json()
    records = body.get("records", [])
    if not records:
        return {"status": "ok", "inserted": 0}

    from app.services.whatsapp_service import send_cloud_template_message

    db = await get_db()
    try:
        inserted = 0
        updated = 0
        rejected = 0
        backend_notified = 0
        for rec in records:
            person_id = rec.get("person_id", "")
            name = rec.get("name", "")
            grade = rec.get("grade", "")
            camera = rec.get("camera", "")
            confidence = rec.get("confidence", 0)
            notified = 1 if rec.get("notification_sent") else 0
            phones = rec.get("parent_phones", "")
            logged_at = rec.get("logged_at", datetime.now().isoformat())

            # Skip student records on ALL Saturdays (teachers/principals still allowed)
            if _block_students_today and not person_id.startswith(("TEACHER_", "PRINCIPAL_")):
                rejected += 1
                logger.info(f"Student attendance skipped on Saturday: {name} ({person_id})")
                continue

            # Confidence floor: reject low-confidence matches at backend level
            if confidence < MINIMUM_CONFIDENCE:
                rejected += 1
                logger.warning(
                    f"Attendance rejected: {name} ({person_id}) confidence "
                    f"{confidence:.1%} < {MINIMUM_CONFIDENCE:.0%} minimum"
                )
                await db.execute(
                    "INSERT INTO audit_log (action, table_name, record_id, details) "
                    "VALUES (?, ?, ?, ?)",
                    ("rejected_low_confidence", "attendance_records", person_id,
                     f"confidence={confidence:.4f}, camera={camera}, name={name}"),
                )
                continue

            # ALWAYS send notification from backend if not already notified.
            # The campus agent may not have the phone number locally, or may
            # have failed to send. Backend is the safety net — guaranteed delivery.
            if not notified:
                # Look up phone from backend face DB
                if not phones:
                    cursor = await db.execute(
                        "SELECT phone FROM agent_registered_faces "
                        "WHERE person_id = ? AND phone IS NOT NULL AND phone != '' "
                        "LIMIT 1",
                        (person_id,),
                    )
                    row = await cursor.fetchone()
                    if row and row[0]:
                        phones = row[0]

                if phones:
                    try:
                        _ts = datetime.fromisoformat(logged_at)
                        _ist = _ts.astimezone(
                            __import__("datetime").timezone(timedelta(hours=5, minutes=30))
                        )
                        time_str = _ist.strftime("%I:%M %p")
                    except Exception:
                        time_str = "this morning"

                    notif_name = _attendance_notif_name(person_id, name)
                    tpl_name, tpl_lang = _attendance_template(person_id)

                    phone_list = [p.strip() for p in phones.split(",") if p.strip()]
                    _record_notified = False
                    for ph in phone_list:
                        digits = "".join(c for c in ph if c.isdigit())
                        if len(digits) == 10:
                            digits = "91" + digits
                        if len(digits) >= 12:
                            ok = await send_cloud_template_message(
                                digits, tpl_name,
                                language_code=tpl_lang,
                                body_params=[notif_name, time_str],
                            )
                            if ok:
                                logger.info(
                                    f"Backend sent attendance notification for "
                                    f"{name} ({person_id}) to {digits}"
                                )
                                backend_notified += 1
                                _record_notified = True
                            else:
                                logger.warning(
                                    f"Backend notification failed for {name} to {digits}"
                                )
                    if _record_notified:
                        notified = 1

            # Check if already reported today for this person
            cursor = await db.execute(
                "SELECT id FROM attendance_records "
                "WHERE person_id = ? AND date(logged_at) = date(?)",
                (person_id, logged_at),
            )
            existing = await cursor.fetchone()
            if existing:
                await db.execute(
                    "UPDATE attendance_records SET notification_sent = ?, "
                    "logged_at = ?, confidence = ?, parent_phones = ? "
                    "WHERE id = ?",
                    (notified, logged_at, confidence, phones, existing[0]),
                )
                updated += 1
                continue

            await db.execute(
                "INSERT INTO attendance_records "
                "(person_id, student_name, grade, camera_label, confidence, "
                "notification_sent, parent_phones, logged_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (person_id, name, grade, camera, confidence, notified, phones, logged_at),
            )
            inserted += 1

        await db.commit()
        return {
            "status": "ok",
            "inserted": inserted,
            "updated": updated,
            "rejected": rejected,
            "backend_notified": backend_notified,
        }
    finally:
        await db.close()


@router.post("/attendance/resend-missed")
async def resend_missed_notifications():
    """Find today's attendance records with notification_sent=0, look up phones,
    and send attendance notifications for all of them."""
    from app.services.whatsapp_service import send_cloud_template_message

    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT ar.person_id, ar.student_name, ar.logged_at "
            "FROM attendance_records ar "
            "WHERE date(ar.logged_at) = date('now', '+5 hours', '+30 minutes') "
            "AND ar.notification_sent = 0"
        )
        rows = await cursor.fetchall()
        sent = 0
        failed = 0
        for r in rows:
            person_id, name, logged_at = r[0], r[1], r[2]
            # Look up phone from face DB
            pcur = await db.execute(
                "SELECT phone FROM agent_registered_faces "
                "WHERE person_id = ? AND phone IS NOT NULL AND phone != '' LIMIT 1",
                (person_id,),
            )
            prow = await pcur.fetchone()
            if not prow or not prow[0]:
                logger.warning(f"Resend: no phone for {person_id}")
                failed += 1
                continue
            phones = prow[0]
            try:
                _ts = datetime.fromisoformat(logged_at)
                time_str = _ts.strftime("%I:%M %p")
            except Exception:
                time_str = "this morning"

            notif_name = _attendance_notif_name(person_id, name)
            tpl_name, tpl_lang = _attendance_template(person_id)

            phone_list = [p.strip() for p in phones.split(",") if p.strip()]
            any_ok = False
            for ph in phone_list:
                digits = "".join(c for c in ph if c.isdigit())
                if len(digits) == 10:
                    digits = "91" + digits
                if len(digits) >= 12:
                    ok = await send_cloud_template_message(
                        digits, tpl_name,
                        language_code=tpl_lang,
                        body_params=[notif_name, time_str],
                    )
                    if ok:
                        any_ok = True
                        logger.info(f"Resend: sent notification for {name} to {digits}")
            if any_ok:
                await db.execute(
                    "UPDATE attendance_records SET notification_sent = 1, "
                    "parent_phones = ? WHERE person_id = ? "
                    "AND date(logged_at) = date('now', '+5 hours', '+30 minutes')",
                    (phones, person_id),
                )
                sent += 1
            else:
                failed += 1
        await db.commit()
        return {"status": "ok", "sent": sent, "failed": failed, "total_missed": len(rows)}
    finally:
        await db.close()


@router.post("/attendance/resend-all")
async def resend_all_notifications():
    """Send attendance notifications for ALL detected people today,
    regardless of whether they were already notified. Useful for
    ensuring no one was missed."""
    from app.services.whatsapp_service import send_cloud_template_message

    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT ar.person_id, ar.student_name, ar.logged_at "
            "FROM attendance_records ar "
            "WHERE date(ar.logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        rows = await cursor.fetchall()
        sent = 0
        failed = 0
        no_phone = 0
        for r in rows:
            person_id, name, logged_at = r[0], r[1], r[2]
            pcur = await db.execute(
                "SELECT phone FROM agent_registered_faces "
                "WHERE person_id = ? AND phone IS NOT NULL AND phone != '' LIMIT 1",
                (person_id,),
            )
            prow = await pcur.fetchone()
            if not prow or not prow[0]:
                logger.warning(f"Resend-all: no phone for {person_id}")
                no_phone += 1
                continue
            phones = prow[0]
            try:
                _ts = datetime.fromisoformat(logged_at)
                time_str = _ts.strftime("%I:%M %p")
            except Exception:
                time_str = "this morning"

            notif_name = _attendance_notif_name(person_id, name)
            tpl_name, tpl_lang = _attendance_template(person_id)

            phone_list = [p.strip() for p in phones.split(",") if p.strip()]
            any_ok = False
            for ph in phone_list:
                digits = "".join(c for c in ph if c.isdigit())
                if len(digits) == 10:
                    digits = "91" + digits
                if len(digits) >= 12:
                    ok = await send_cloud_template_message(
                        digits, tpl_name,
                        language_code=tpl_lang,
                        body_params=[notif_name, time_str],
                    )
                    if ok:
                        any_ok = True
            if any_ok:
                await db.execute(
                    "UPDATE attendance_records SET notification_sent = 1, "
                    "parent_phones = ? WHERE person_id = ? "
                    "AND date(logged_at) = date('now', '+5 hours', '+30 minutes')",
                    (phones, person_id),
                )
                sent += 1
            else:
                failed += 1
        await db.commit()
        return {
            "status": "ok",
            "sent": sent,
            "failed": failed,
            "no_phone": no_phone,
            "total_records": len(rows),
        }
    finally:
        await db.close()


@router.delete("/attendance/record/{person_id}",
               dependencies=[Depends(verify_dashboard_secret)])
async def delete_attendance_record(person_id: str):
    """Delete an attendance record by person_id. Requires AGENT_SECRET."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "DELETE FROM attendance_records WHERE LOWER(person_id) = LOWER(?)", (person_id,)
        )
        deleted_count = cursor.rowcount
        if deleted_count > 0:
            await db.execute(
                "INSERT INTO audit_log (action, table_name, record_id, details) "
                "VALUES (?, ?, ?, ?)",
                ("delete", "attendance_records", person_id,
                 f"Deleted {deleted_count} record(s)"),
            )
        await db.commit()
        return {"status": "ok", "deleted": deleted_count, "person_id": person_id}
    finally:
        await db.close()


# ── Chats ────────────────────────────────────────────────────────────────────

@router.get("/chats")
async def get_chats(
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    phone: str | None = None,
    direction: str | None = None,
    search: str | None = None,
):
    """Get chat messages with filters."""
    db = await get_db()
    try:
        conditions: list[str] = []
        params: list = []

        if phone:
            conditions.append("(sender LIKE ? OR receiver LIKE ?)")
            params.extend([f"%{phone}%", f"%{phone}%"])
        if direction:
            conditions.append("direction = ?")
            params.append(direction)
        if search:
            conditions.append("content LIKE ?")
            params.append(f"%{search}%")

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        cursor = await db.execute(
            f"SELECT id, sender, receiver, content, channel, direction, "
            f"datetime(timestamp, '+5 hours', '+30 minutes') "
            f"FROM messages {where} ORDER BY timestamp DESC LIMIT ? OFFSET ?",
            params + [limit, offset],
        )
        rows = await cursor.fetchall()

        # Total count
        cursor = await db.execute(
            f"SELECT COUNT(*) FROM messages {where}", params
        )
        total = (await cursor.fetchone())[0]

        return {
            "total": total,
            "messages": [
                {
                    "id": r[0], "sender": r[1], "receiver": r[2],
                    "content": r[3], "channel": r[4], "direction": r[5],
                    "timestamp": r[6],
                }
                for r in rows
            ],
        }
    finally:
        await db.close()


@router.get("/chats/conversations")
async def get_conversations(limit: int = Query(30, ge=1, le=100)):
    """Get unique conversations (grouped by phone number) with latest message."""
    db = await get_db()
    try:
        cursor = await db.execute(
            """
            SELECT
                CASE WHEN direction = 'incoming' THEN sender ELSE receiver END as phone,
                content as last_message,
                datetime(timestamp, '+5 hours', '+30 minutes') as last_time,
                direction
            FROM messages
            WHERE id IN (
                SELECT MAX(id) FROM messages
                GROUP BY CASE WHEN direction = 'incoming' THEN sender ELSE receiver END
            )
            ORDER BY timestamp DESC
            LIMIT ?
            """,
            (limit,),
        )
        rows = await cursor.fetchall()

        conversations = []
        for r in rows:
            phone = r[0]
            # Look up parent name from pi_sheet_students
            cur2 = await db.execute(
                "SELECT student_name, grade, father_name, mother_name "
                "FROM pi_sheet_students "
                "WHERE father_mobile LIKE ? OR mother_mobile LIKE ? LIMIT 1",
                (f"%{phone[-10:]}%", f"%{phone[-10:]}%"),
            )
            parent_row = cur2 and await cur2.fetchone()
            label = ""
            if parent_row:
                label = f"{parent_row[0]} ({parent_row[1]})"

            conversations.append({
                "phone": phone,
                "label": label,
                "last_message": r[1][:100] if r[1] else "",
                "last_time": r[2],
                "direction": r[3],
            })

        return {"conversations": conversations}
    finally:
        await db.close()


# ── Notifications ────────────────────────────────────────────────────────────

@router.get("/notifications/stats")
async def notification_stats():
    """Get notification delivery statistics."""
    db = await get_db()
    try:
        # Count messages by direction today
        cursor = await db.execute(
            "SELECT direction, COUNT(*) FROM messages "
            "WHERE date(timestamp) = date('now', '+5 hours', '+30 minutes') GROUP BY direction"
        )
        msg_counts = {r[0]: r[1] for r in await cursor.fetchall()}

        # Total messages
        cursor = await db.execute("SELECT COUNT(*) FROM messages")
        total_messages = (await cursor.fetchone())[0]

        # Today's attendance notifications
        cursor = await db.execute(
            "SELECT COUNT(*) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes') AND notification_sent = 1"
        )
        notified_today = (await cursor.fetchone())[0]

        return {
            "today_incoming": msg_counts.get("incoming", 0),
            "today_outgoing": msg_counts.get("outgoing", 0),
            "total_messages": total_messages,
            "attendance_notified_today": notified_today,
        }
    finally:
        await db.close()


# ── Students ─────────────────────────────────────────────────────────────────

@router.get("/students")
async def get_students(
    grade: str | None = None,
    search: str | None = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """Get student directory from PI sheet."""
    db = await get_db()
    try:
        conditions: list[str] = []
        params: list = []

        if grade:
            conditions.append("grade LIKE ?")
            params.append(f"%{grade}%")
        if search:
            conditions.append("(student_name LIKE ? OR father_name LIKE ? OR mother_name LIKE ?)")
            params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        cursor = await db.execute(
            f"SELECT id, student_name, grade, father_name, mother_name, "
            f"father_mobile, mother_mobile, address, transport "
            f"FROM pi_sheet_students {where} ORDER BY grade, student_name "
            f"LIMIT ? OFFSET ?",
            params + [limit, offset],
        )
        rows = await cursor.fetchall()

        cursor = await db.execute(
            f"SELECT COUNT(*) FROM pi_sheet_students {where}", params
        )
        total = (await cursor.fetchone())[0]

        return {
            "total": total,
            "students": [
                {
                    "id": r[0], "name": r[1], "grade": r[2],
                    "father_name": r[3], "mother_name": r[4],
                    "father_phone": r[5], "mother_phone": r[6],
                    "address": r[7], "transport": r[8],
                }
                for r in rows
            ],
        }
    finally:
        await db.close()


@router.post("/students/refresh")
async def refresh_students():
    """Re-import all students from the PI Sheet (all grade tabs).

    Fetches every grade tab, excludes withdrawn students, deduplicates,
    and replaces the pi_sheet_students table.
    """
    from app.services.sheet_refresh_service import fetch_all_pi_sheet_tabs

    result = await fetch_all_pi_sheet_tabs()
    if result:
        db = await get_db()
        try:
            cursor = await db.execute("SELECT COUNT(*) FROM pi_sheet_students")
            total = (await cursor.fetchone())[0]
            cursor = await db.execute(
                "SELECT COUNT(*) FROM pi_sheet_students "
                "WHERE (father_mobile != '' AND father_mobile IS NOT NULL) "
                "OR (mother_mobile != '' AND mother_mobile IS NOT NULL)"
            )
            with_phones = (await cursor.fetchone())[0]
            return {
                "status": "ok",
                "total_students": total,
                "with_phones": with_phones,
                "missing_phones": total - with_phones,
            }
        finally:
            await db.close()
    return {"status": "error", "message": "Failed to refresh PI Sheet data"}


@router.delete("/students/{student_id}")
async def delete_student(student_id: int):
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT student_name, grade FROM pi_sheet_students WHERE id = ?",
            (student_id,),
        )
        row = await cursor.fetchone()
        if not row:
            return {"status": "error", "message": "Student not found"}
        await db.execute("DELETE FROM pi_sheet_students WHERE id = ?", (student_id,))
        await db.commit()
        return {"status": "ok", "deleted": {"id": student_id, "name": row[0], "grade": row[1]}}
    finally:
        await db.close()


@router.get("/students/grades")
async def get_grades():
    """Get list of all grades with student counts."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT grade, COUNT(*) as count FROM pi_sheet_students "
            "GROUP BY grade ORDER BY grade"
        )
        return {"grades": [{"grade": r[0], "count": r[1]} for r in await cursor.fetchall()]}
    finally:
        await db.close()


# ── Face Registration ────────────────────────────────────────────────────────

@router.get("/faces")
async def get_registered_faces():
    """Get all registered face entries."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT person_id, name, role, phone, angle, registered_at "
            "FROM agent_registered_faces ORDER BY registered_at DESC"
        )
        rows = await cursor.fetchall()

        # Group by person
        persons: dict = {}
        for r in rows:
            pid = r[0]
            if pid not in persons:
                persons[pid] = {
                    "person_id": pid, "name": r[1], "role": r[2],
                    "phone": r[3], "registered_at": r[5], "face_count": 0,
                }
            persons[pid]["face_count"] += 1

        return {"registered": list(persons.values()), "total": len(persons)}
    finally:
        await db.close()


@router.delete("/faces/{person_id}",
               dependencies=[Depends(verify_dashboard_secret)])
async def dashboard_delete_face(person_id: str):
    """Delete a face entry. Requires AGENT_SECRET."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT person_id FROM agent_registered_faces "
            "WHERE person_id = ? COLLATE NOCASE LIMIT 1",
            (person_id,),
        )
        row = await cursor.fetchone()
        if not row:
            return {"status": "error", "message": "Person not found"}
        original_pid = row[0]
        cursor = await db.execute(
            "DELETE FROM agent_registered_faces WHERE person_id = ? COLLATE NOCASE",
            (person_id,),
        )
        deleted = cursor.rowcount
        await db.execute(
            "INSERT INTO audit_log (action, table_name, record_id, details) "
            "VALUES (?, ?, ?, ?)",
            ("delete", "agent_registered_faces", original_pid,
             f"Deleted {deleted} face(s)"),
        )
        await db.commit()
        logger.info(f"Dashboard: deleted {deleted} face(s) for {original_pid}")
        return {"status": "ok", "deleted": deleted, "person_id": original_pid}
    finally:
        await db.close()


@router.patch("/faces/{person_id}/phone",
              dependencies=[Depends(verify_dashboard_secret)])
async def dashboard_update_face_phone(person_id: str, request: Request):
    """Update the phone number for a face entry. Requires AGENT_SECRET."""
    body = await request.json()
    phone = body.get("phone", "")
    if not phone:
        return {"status": "error", "message": "Missing phone"}
    db = await get_db()
    try:
        cursor = await db.execute(
            "UPDATE agent_registered_faces SET phone = ? "
            "WHERE person_id = ? COLLATE NOCASE",
            (phone, person_id),
        )
        await db.commit()
        updated = cursor.rowcount
        logger.info(f"Dashboard: updated phone for {person_id}: {phone} ({updated} rows)")
        return {"status": "ok", "updated": updated, "person_id": person_id, "phone": phone}
    finally:
        await db.close()


# ── Cameras ──────────────────────────────────────────────────────────────────

@router.get("/cameras")
async def get_cameras():
    """Get camera mapping."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT location, dvr_index, channel, description, cam_type "
            "FROM agent_camera_mapping ORDER BY location"
        )
        rows = await cursor.fetchall()
        return {
            "total": len(rows),
            "cameras": [
                {
                    "location": r[0], "dvr_index": r[1], "channel": r[2],
                    "description": r[3],
                    "cam_type": r[4] if r[4] else f"DVR {r[1] + 1}",
                }
                for r in rows
            ],
        }
    finally:
        await db.close()


# ── Overview ─────────────────────────────────────────────────────────────────

@router.get("/overview")
async def dashboard_overview():
    """Get a full overview of the system for the dashboard home page."""
    db = await get_db()
    try:
        # Students
        cursor = await db.execute("SELECT COUNT(*) FROM pi_sheet_students")
        total_students = (await cursor.fetchone())[0]

        # Registered faces
        cursor = await db.execute("SELECT COUNT(DISTINCT person_id) FROM agent_registered_faces")
        registered_faces = (await cursor.fetchone())[0]

        # Today's attendance
        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        present_today = (await cursor.fetchone())[0]

        # Today's messages
        cursor = await db.execute(
            "SELECT COUNT(*) FROM messages WHERE date(timestamp) = date('now', '+5 hours', '+30 minutes')"
        )
        messages_today = (await cursor.fetchone())[0]

        # Total messages
        cursor = await db.execute("SELECT COUNT(*) FROM messages")
        total_messages = (await cursor.fetchone())[0]

        # Cameras
        cursor = await db.execute("SELECT COUNT(*) FROM agent_camera_mapping")
        total_cameras = (await cursor.fetchone())[0]

        # Recent activity (last 10 messages) — convert to IST
        cursor = await db.execute(
            "SELECT sender, content, direction, "
            "datetime(timestamp, '+5 hours', '+30 minutes') "
            "FROM messages ORDER BY timestamp DESC LIMIT 10"
        )
        recent_messages = [
            {"sender": r[0], "content": r[1][:80], "direction": r[2], "time": r[3]}
            for r in await cursor.fetchall()
        ]

        return {
            "total_students": total_students,
            "registered_faces": registered_faces,
            "present_today": present_today,
            "messages_today": messages_today,
            "total_messages": total_messages,
            "total_cameras": total_cameras,
            "recent_activity": recent_messages,
        }
    finally:
        await db.close()


# ── Teacher Attendance Excel ─────────────────────────────────────────────────

@router.get("/teacher-attendance-excel")
async def download_teacher_attendance_excel(month: str | None = None):
    """Download the teacher attendance Excel workbook.

    Optional query param `month` in YYYY-MM format regenerates for that month.
    Without it, regenerates for the current month and returns the file.
    """
    from app.services.teacher_attendance_excel import (
        generate_teacher_attendance_excel,
        EXCEL_PATH,
    )

    try:
        if month:
            parts = month.split("-")
            target = date(int(parts[0]), int(parts[1]), 1)
        else:
            from app.services.teacher_attendance_excel import IST
            target = datetime.now(IST).date()

        await generate_teacher_attendance_excel(target)
    except Exception as e:
        logger.error(f"Teacher attendance Excel generation failed: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

    if not Path(str(EXCEL_PATH)).exists():
        return {"status": "error", "message": "Excel file not found"}

    filename = f"PPIS_Teacher_Attendance_{target.strftime('%B_%Y')}.xlsx"
    return FileResponse(
        path=str(EXCEL_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


# ── Manual Review Queue ──────────────────────────────────────────────────────

@router.get("/review/pending")
async def get_pending_reviews(limit: int = Query(50, ge=1, le=200)):
    """Get low-confidence detections pending manual review."""
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT id, person_id, matched_name, grade, camera_label, "
            "confidence, snapshot_path, created_at "
            "FROM manual_review_queue WHERE status = 'pending' "
            "ORDER BY created_at DESC LIMIT ?",
            (limit,),
        )
        rows = await cursor.fetchall()
        return {
            "pending_count": len(rows),
            "reviews": [
                {
                    "id": r[0], "person_id": r[1], "matched_name": r[2],
                    "grade": r[3], "camera": r[4],
                    "confidence": round(r[5] * 100, 1),
                    "snapshot": r[6], "time": r[7],
                }
                for r in rows
            ],
        }
    finally:
        await db.close()


@router.post("/review/{review_id}/approve")
async def approve_review(review_id: int):
    """Approve a manual review — marks attendance and sends notification."""
    from app.services.whatsapp_service import send_cloud_template_message
    from datetime import timezone

    ist = timezone(timedelta(hours=5, minutes=30))
    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT person_id, matched_name, grade, camera_label, confidence, "
            "created_at FROM manual_review_queue WHERE id = ? AND status = 'pending'",
            (review_id,),
        )
        row = await cursor.fetchone()
        if not row:
            return {"status": "error", "message": "Review not found or already processed"}

        person_id, name, grade, camera, confidence, logged_at = row

        # Mark as approved
        await db.execute(
            "UPDATE manual_review_queue SET status = 'approved', "
            "reviewed_by = 'admin', reviewed_at = datetime('now') WHERE id = ?",
            (review_id,),
        )

        # Insert attendance record
        await db.execute(
            "INSERT INTO attendance_records "
            "(person_id, student_name, grade, camera_label, confidence, "
            "notification_sent, logged_at) VALUES (?, ?, ?, ?, ?, 0, ?)",
            (person_id, name, grade, camera, confidence, logged_at),
        )

        # Send notification
        pcur = await db.execute(
            "SELECT phone FROM agent_registered_faces "
            "WHERE person_id = ? AND phone IS NOT NULL AND phone != '' LIMIT 1",
            (person_id,),
        )
        prow = await pcur.fetchone()
        notified = False
        if prow and prow[0]:
            phones = prow[0]
            try:
                _ts = datetime.fromisoformat(logged_at)
                time_str = _ts.strftime("%I:%M %p")
            except Exception:
                time_str = "this morning"

            notif_name = _attendance_notif_name(person_id, name)
            tpl_name, tpl_lang = _attendance_template(person_id)

            phone_list = [p.strip() for p in phones.split(",") if p.strip()]
            for ph in phone_list:
                digits = "".join(c for c in ph if c.isdigit())
                if len(digits) == 10:
                    digits = "91" + digits
                if len(digits) >= 12:
                    ok = await send_cloud_template_message(
                        digits, tpl_name,
                        language_code=tpl_lang,
                        body_params=[notif_name, time_str],
                    )
                    if ok:
                        notified = True

            if notified:
                await db.execute(
                    "UPDATE attendance_records SET notification_sent = 1, "
                    "parent_phones = ? WHERE person_id = ? "
                    "AND date(logged_at) = date(?)",
                    (phones, person_id, logged_at),
                )

        await db.commit()
        return {"status": "ok", "person_id": person_id, "name": name, "notified": notified}
    finally:
        await db.close()


@router.post("/review/{review_id}/reject")
async def reject_review(review_id: int):
    """Reject a manual review — false positive, do not mark attendance."""
    db = await get_db()
    try:
        await db.execute(
            "UPDATE manual_review_queue SET status = 'rejected', "
            "reviewed_by = 'admin', reviewed_at = datetime('now') WHERE id = ?",
            (review_id,),
        )
        await db.commit()
        return {"status": "ok", "review_id": review_id}
    finally:
        await db.close()


# ── Camera Status ────────────────────────────────────────────────────────────

@router.get("/cameras/status")
async def get_camera_status():
    """Get current camera status summary."""
    db = await get_db()
    try:
        # Get latest status for each camera
        cursor = await db.execute(
            "SELECT camera_label, dvr_ip, channel, status, error_code, "
            "consecutive_failures, last_success_at, last_failure_at "
            "FROM camera_status_log "
            "WHERE id IN (SELECT MAX(id) FROM camera_status_log GROUP BY camera_label) "
            "ORDER BY camera_label"
        )
        rows = await cursor.fetchall()
        online = [r for r in rows if r[3] == "online"]
        offline = [r for r in rows if r[3] != "online"]
        return {
            "total_cameras": len(rows),
            "online": len(online),
            "offline": len(offline),
            "cameras": [
                {
                    "label": r[0], "dvr_ip": r[1], "channel": r[2],
                    "status": r[3], "error": r[4],
                    "failures": r[5], "last_success": r[6], "last_failure": r[7],
                }
                for r in rows
            ],
        }
    finally:
        await db.close()


@router.post("/cameras/status/report")
async def report_camera_status(request: Request):
    """Receive camera status report from the campus agent."""
    body = await request.json()
    cameras = body.get("cameras", [])
    if not cameras:
        return {"status": "ok", "updated": 0}

    db = await get_db()
    try:
        updated = 0
        for cam in cameras:
            label = cam.get("label", "")
            dvr_ip = cam.get("dvr_ip", "")
            channel = cam.get("channel", 0)
            status = cam.get("status", "online")
            error_code = cam.get("error_code", "")
            failures = cam.get("consecutive_failures", 0)

            now = datetime.now().isoformat()
            last_success = now if status == "online" else None
            last_failure = now if status != "online" else None

            await db.execute(
                "INSERT INTO camera_status_log "
                "(camera_label, dvr_ip, channel, status, error_code, "
                "consecutive_failures, last_success_at, last_failure_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (label, dvr_ip, channel, status, error_code,
                 failures, last_success, last_failure),
            )
            updated += 1

        await db.commit()
        return {"status": "ok", "updated": updated}
    finally:
        await db.close()


# ── Daily Summary ────────────────────────────────────────────────────────────

@router.get("/summary/today")
async def get_today_summary():
    """Get comprehensive live dashboard data for today."""
    from datetime import timezone
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    today_str = now_ist.strftime("%Y-%m-%d")

    db = await get_db()
    try:
        # Present count
        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        present = (await cursor.fetchone())[0]

        # Teachers present
        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes') "
            "AND person_id LIKE 'TEACHER_%'"
        )
        teachers_present = (await cursor.fetchone())[0]

        # Students present
        students_present = present - teachers_present

        # Total registered
        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM agent_registered_faces"
        )
        total_registered = (await cursor.fetchone())[0]

        # Notifications sent vs failed
        cursor = await db.execute(
            "SELECT "
            "SUM(CASE WHEN notification_sent = 1 THEN 1 ELSE 0 END), "
            "SUM(CASE WHEN notification_sent = 0 THEN 1 ELSE 0 END) "
            "FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        notif_row = await cursor.fetchone()
        notifications_sent = notif_row[0] or 0
        notifications_pending = notif_row[1] or 0

        # Manual reviews pending
        cursor = await db.execute(
            "SELECT COUNT(*) FROM manual_review_queue "
            "WHERE status = 'pending' AND date(created_at) = date('now', '+5 hours', '+30 minutes')"
        )
        reviews_pending = (await cursor.fetchone())[0]

        # Camera status
        cursor = await db.execute(
            "SELECT status, COUNT(*) FROM camera_status_log "
            "WHERE id IN (SELECT MAX(id) FROM camera_status_log GROUP BY camera_label) "
            "GROUP BY status"
        )
        cam_status = {r[0]: r[1] for r in await cursor.fetchall()}
        cameras_online = cam_status.get("online", 0)
        cameras_offline = cam_status.get("offline", 0) + cam_status.get("error", 0)

        # Grade breakdown for absent tracking — extract grade from person_id
        # person_id format: NAME_GRADE (e.g. SUHAAN_AHUJA_GRADE3C)
        cursor = await db.execute(
            "SELECT DISTINCT person_id FROM agent_registered_faces "
            "WHERE person_id NOT LIKE 'TEACHER_%'"
        )
        registered_by_grade: dict[str, int] = {}
        for r in await cursor.fetchall():
            pid = r[0]
            # Extract grade: last segment that starts with common grade patterns
            parts = pid.split("_")
            grade_part = ""
            for p in parts:
                if p.startswith("GRADE") or p.startswith("PREP"):
                    grade_part = p
            if grade_part:
                registered_by_grade[grade_part] = registered_by_grade.get(grade_part, 0) + 1

        cursor = await db.execute(
            "SELECT grade, COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes') "
            "AND person_id NOT LIKE 'TEACHER_%' GROUP BY grade"
        )
        present_by_grade = {r[0]: r[1] for r in await cursor.fetchall()}

        grade_summary = []
        for grade, total in sorted(registered_by_grade.items()):
            p = present_by_grade.get(grade, 0)
            grade_summary.append({
                "grade": grade, "registered": total,
                "present": p, "absent": total - p,
            })

        return {
            "date": today_str,
            "time": now_ist.strftime("%I:%M %p"),
            "students_present": students_present,
            "teachers_present": teachers_present,
            "total_present": present,
            "total_registered": total_registered,
            "notifications_sent": notifications_sent,
            "notifications_pending": notifications_pending,
            "manual_reviews_pending": reviews_pending,
            "cameras_online": cameras_online,
            "cameras_offline": cameras_offline,
            "grade_summary": grade_summary,
        }
    finally:
        await db.close()


@router.post("/summary/generate")
async def generate_daily_summary():
    """Generate and store the daily summary report (call at end of school day)."""
    from datetime import timezone
    ist = timezone(timedelta(hours=5, minutes=30))
    now_ist = datetime.now(ist)
    today_str = now_ist.strftime("%Y-%m-%d")

    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        total_present = (await cursor.fetchone())[0]

        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes') "
            "AND person_id LIKE 'TEACHER_%'"
        )
        teachers_present = (await cursor.fetchone())[0]

        cursor = await db.execute(
            "SELECT COUNT(DISTINCT person_id) FROM agent_registered_faces "
            "WHERE person_id NOT LIKE 'TEACHER_%'"
        )
        total_students = (await cursor.fetchone())[0]
        total_absent = total_students - (total_present - teachers_present)

        cursor = await db.execute(
            "SELECT "
            "SUM(CASE WHEN notification_sent = 1 THEN 1 ELSE 0 END), "
            "SUM(CASE WHEN notification_sent = 0 THEN 1 ELSE 0 END) "
            "FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        n_row = await cursor.fetchone()
        notif_sent = n_row[0] or 0
        notif_failed = n_row[1] or 0

        cursor = await db.execute(
            "SELECT COUNT(*) FROM manual_review_queue "
            "WHERE status = 'pending' AND date(created_at) = date('now', '+5 hours', '+30 minutes')"
        )
        reviews = (await cursor.fetchone())[0]

        await db.execute(
            "INSERT OR REPLACE INTO daily_summary "
            "(report_date, total_present, total_absent, total_teachers_present, "
            "total_notifications_sent, total_notifications_failed, "
            "cameras_online, cameras_offline, manual_reviews_pending) "
            "VALUES (?, ?, ?, ?, ?, ?, 0, 0, ?)",
            (today_str, total_present, total_absent, teachers_present,
             notif_sent, notif_failed, reviews),
        )
        await db.commit()
        return {
            "status": "ok",
            "date": today_str,
            "present": total_present,
            "absent": total_absent,
            "teachers": teachers_present,
            "notifications_sent": notif_sent,
            "notifications_failed": notif_failed,
            "reviews_pending": reviews,
        }
    finally:
        await db.close()


# ── Notification Retry ───────────────────────────────────────────────────────

@router.post("/notifications/retry-failed")
async def retry_failed_notifications():
    """Retry all failed notification deliveries from today."""
    from app.services.whatsapp_service import send_cloud_template_message

    db = await get_db()
    try:
        cursor = await db.execute(
            "SELECT nd.id, nd.person_id, nd.student_name, nd.phone, nd.attempts "
            "FROM notification_delivery nd "
            "WHERE nd.status = 'failed' AND nd.attempts < 3 "
            "AND date(nd.created_at) = date('now', '+5 hours', '+30 minutes')"
        )
        rows = await cursor.fetchall()
        retried = 0
        still_failed = 0
        for r in rows:
            nd_id, person_id, name, phone, attempts = r

            notif_name = _attendance_notif_name(person_id, name)
            tpl_name, tpl_lang = _attendance_template(person_id)

            ok = await send_cloud_template_message(
                phone, tpl_name,
                language_code=tpl_lang,
                body_params=[notif_name, "this morning"],
            )
            if ok:
                await db.execute(
                    "UPDATE notification_delivery SET status = 'delivered', "
                    "delivered_at = datetime('now'), attempts = ? WHERE id = ?",
                    (attempts + 1, nd_id),
                )
                retried += 1
            else:
                new_status = "failed" if attempts + 1 < 3 else "permanently_failed"
                await db.execute(
                    "UPDATE notification_delivery SET status = ?, "
                    "attempts = ?, last_attempt_at = datetime('now') WHERE id = ?",
                    (new_status, attempts + 1, nd_id),
                )
                still_failed += 1

        await db.commit()
        return {"status": "ok", "retried": retried, "still_failed": still_failed}
    finally:
        await db.close()


# ── Absent Students ──────────────────────────────────────────────────────────

@router.get("/attendance/absent")
async def get_absent_students(grade: str | None = None):
    """Get students who are registered but NOT marked present today."""
    db = await get_db()
    try:
        # Get all present person_ids today
        cursor = await db.execute(
            "SELECT DISTINCT person_id FROM attendance_records "
            "WHERE date(logged_at) = date('now', '+5 hours', '+30 minutes')"
        )
        present_ids = {r[0] for r in await cursor.fetchall()}

        # Get all registered faces (students only)
        conditions = ["person_id NOT LIKE 'TEACHER_%'"]
        params: list = []
        if grade:
            conditions.append("person_id LIKE ?")
            params.append(f"%{grade}%")

        where = " AND ".join(conditions)
        cursor = await db.execute(
            f"SELECT DISTINCT person_id, name FROM agent_registered_faces "
            f"WHERE {where}",
            params,
        )
        all_students = await cursor.fetchall()

        absent = []
        for r in all_students:
            if r[0] not in present_ids:
                # Extract grade from person_id
                parts = r[0].rsplit("_", 1)
                g = parts[-1] if len(parts) > 1 else ""
                absent.append({"person_id": r[0], "name": r[1], "grade": g})

        return {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "total_absent": len(absent),
            "students": sorted(absent, key=lambda x: (x["grade"], x["name"])),
        }
    finally:
        await db.close()


# ── Manual Review Report from Agent ──────────────────────────────────────────

@router.post("/review/report")
async def report_manual_review(request: Request):
    """Receive low-confidence detection from agent for manual review."""
    body = await request.json()
    records = body.get("records", [])
    if not records:
        return {"status": "ok", "queued": 0}

    db = await get_db()
    try:
        queued = 0
        for rec in records:
            await db.execute(
                "INSERT INTO manual_review_queue "
                "(person_id, matched_name, grade, camera_label, confidence, snapshot_path) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (
                    rec.get("person_id", ""),
                    rec.get("name", ""),
                    rec.get("grade", ""),
                    rec.get("camera", ""),
                    rec.get("confidence", 0),
                    rec.get("snapshot_path", ""),
                ),
            )
            queued += 1
        await db.commit()
        return {"status": "ok", "queued": queued}
    finally:
        await db.close()


# ── Teacher/Staff Attendance Report ──────────────────────────────────────────

REPORT_RECIPIENTS = os.environ.get(
    "ATTENDANCE_REPORT_EMAIL", "alisha.kanwar@ppischool.in,leave@ppischool.in"
)


def _generate_teacher_report_excel(
    teachers: list[dict], attendance: list[dict], report_date: str
) -> bytes:
    """Generate Excel bytes for the daily teacher/staff attendance report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Attendance"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    present_fill = PatternFill("solid", fgColor="C6EFCE")
    absent_fill = PatternFill("solid", fgColor="FFC7CE")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"PP International School — Teacher/Staff Attendance — {report_date}"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["S.No", "Teacher Name", "Attendance Status", "Entry Time (IST)",
               "Camera Source", "Confidence %", "Notification Status", "Remarks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Build attendance lookup
    att_map = {}
    for a in attendance:
        pid = a["person_id"]
        if pid not in att_map:
            att_map[pid] = a

    # Data rows
    row = 4
    present_count = 0
    absent_count = 0
    for idx, t in enumerate(teachers, 1):
        pid = t["person_id"]
        att = att_map.get(pid)
        status = "Present" if att else "Absent"
        time_str = ""
        camera = ""
        conf = ""
        notif_status = ""
        remarks = ""
        if att:
            present_count += 1
            time_str = att.get("time", "")
            camera = att.get("camera", "")
            conf_val = att.get("confidence", 0)
            conf = f"{conf_val}%"
            notif_status = "Sent" if att.get("notification_sent") else "Pending"
            remarks = "Face verified via recognition"
        else:
            absent_count += 1
            notif_status = "N/A"
            remarks = "Not detected during attendance window"

        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=t["name"]).border = border
        status_cell = ws.cell(row=row, column=3, value=status)
        status_cell.border = border
        status_cell.fill = present_fill if att else absent_fill
        status_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=time_str).border = border
        ws.cell(row=row, column=5, value=camera).border = border
        ws.cell(row=row, column=6, value=conf).border = border
        notif_cell = ws.cell(row=row, column=7, value=notif_status)
        notif_cell.border = border
        notif_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=8, value=remarks).border = border
        row += 1

    # Summary row
    row += 1
    ws.cell(row=row, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"Present: {present_count}").font = Font(
        bold=True, color="006100"
    )
    ws.cell(row=row, column=3, value=f"Absent: {absent_count}").font = Font(
        bold=True, color="9C0006"
    )
    ws.cell(row=row, column=4, value=f"Total: {len(teachers)}").font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 38

    # ── Sheet 2: Teacher Database ──
    ws2 = wb.create_sheet("Teacher Database")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"PP International School — Teacher/Staff Database — {report_date}"
    ws2["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws2["A1"].alignment = Alignment(horizontal="center")

    db_headers = ["S.No", "Name", "Phone Number", "Role", "Face Registered", "Today's Status"]
    for col, h in enumerate(db_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    reg_fill = PatternFill("solid", fgColor="C6EFCE")
    noreg_fill = PatternFill("solid", fgColor="FFC7CE")
    row2 = 4
    with_phone = 0
    without_phone = 0
    for idx, t in enumerate(teachers, 1):
        pid = t["person_id"]
        phone = t.get("phone", "")
        has_phone = "Yes" if phone else "No"
        if phone:
            with_phone += 1
        else:
            without_phone += 1
        att = att_map.get(pid)
        today_status = "Present" if att else "Absent"

        ws2.cell(row=row2, column=1, value=idx).border = border
        ws2.cell(row=row2, column=2, value=t["name"]).border = border
        phone_cell = ws2.cell(row=row2, column=3, value=phone if phone else "MISSING")
        phone_cell.border = border
        phone_cell.fill = reg_fill if phone else noreg_fill
        ws2.cell(row=row2, column=4, value=t.get("role", "Teacher")).border = border
        face_cell = ws2.cell(row=row2, column=5, value="Yes")
        face_cell.border = border
        face_cell.fill = reg_fill
        face_cell.alignment = Alignment(horizontal="center")
        status_cell2 = ws2.cell(row=row2, column=6, value=today_status)
        status_cell2.border = border
        status_cell2.fill = present_fill if att else absent_fill
        status_cell2.alignment = Alignment(horizontal="center")
        row2 += 1

    row2 += 1
    ws2.cell(row=row2, column=1, value="Summary:").font = Font(bold=True)
    ws2.cell(row=row2, column=2, value=f"With Phone: {with_phone}").font = Font(bold=True, color="006100")
    ws2.cell(row=row2, column=3, value=f"Missing Phone: {without_phone}").font = Font(bold=True, color="9C0006")
    ws2.cell(row=row2, column=4, value=f"Total: {len(teachers)}").font = Font(bold=True)

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/attendance/teacher-report")
async def teacher_attendance_report(report_date: str | None = None):
    """Generate and download today's teacher attendance Excel report."""
    ist = timezone(timedelta(hours=5, minutes=30))
    today = report_date or datetime.now(ist).strftime("%Y-%m-%d")

    db = await get_db()
    try:
        # All registered teachers (including principal)
        cur = await db.execute(
            "SELECT DISTINCT person_id, name, phone, role "
            "FROM agent_registered_faces "
            "WHERE person_id LIKE 'TEACHER_%' OR person_id LIKE 'PRINCIPAL_%' "
            "ORDER BY name"
        )
        teachers = [
            {"person_id": r[0], "name": r[1], "phone": r[2] or "", "role": r[3] or "Teacher"}
            for r in await cur.fetchall()
        ]

        # Today's teacher attendance
        cur = await db.execute(
            "SELECT person_id, student_name, camera_label, confidence, logged_at, "
            "notification_sent "
            "FROM attendance_records "
            "WHERE (person_id LIKE 'TEACHER_%' OR person_id LIKE 'PRINCIPAL_%') "
            "AND date(logged_at) = ? "
            "ORDER BY logged_at",
            (today,),
        )
        attendance = []
        for r in await cur.fetchall():
            logged = r[4] or ""
            time_str = ""
            if logged:
                try:
                    dt = datetime.fromisoformat(logged)
                    time_str = dt.strftime("%I:%M %p")
                except Exception:
                    time_str = logged
            attendance.append({
                "person_id": r[0],
                "name": r[1],
                "camera": r[2] or "",
                "confidence": round(r[3] * 100, 1) if r[3] else 0,
                "time": time_str,
                "notification_sent": bool(r[5]) if r[5] is not None else False,
            })
    finally:
        await db.close()

    xlsx_bytes = _generate_teacher_report_excel(teachers, attendance, today)
    filename = f"Teacher_Attendance_{today}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/attendance/teacher-report/email")
async def email_teacher_attendance_report(
    request: Request,
    _: None = Depends(verify_dashboard_secret),
):
    """Generate teacher attendance Excel and email it.

    Called by the agent after Phase 1 (teacher scanning) completes.
    """
    from app.services.email_service import send_email_async

    ist = timezone(timedelta(hours=5, minutes=30))
    today = datetime.now(ist).strftime("%Y-%m-%d")
    today_display = datetime.now(ist).strftime("%d %b %Y (%A)")

    db = await get_db()
    try:
        cur = await db.execute(
            "SELECT DISTINCT person_id, name, phone, role "
            "FROM agent_registered_faces "
            "WHERE person_id LIKE 'TEACHER_%' OR person_id LIKE 'PRINCIPAL_%' "
            "ORDER BY name"
        )
        teachers = [
            {"person_id": r[0], "name": r[1], "phone": r[2] or "", "role": r[3] or "Teacher"}
            for r in await cur.fetchall()
        ]

        cur = await db.execute(
            "SELECT person_id, student_name, camera_label, confidence, logged_at, "
            "notification_sent "
            "FROM attendance_records "
            "WHERE (person_id LIKE 'TEACHER_%' OR person_id LIKE 'PRINCIPAL_%') "
            "AND date(logged_at) = ? "
            "ORDER BY logged_at",
            (today,),
        )
        attendance = []
        for r in await cur.fetchall():
            logged = r[4] or ""
            time_str = ""
            if logged:
                try:
                    dt = datetime.fromisoformat(logged)
                    time_str = dt.strftime("%I:%M %p")
                except Exception:
                    time_str = logged
            attendance.append({
                "person_id": r[0],
                "name": r[1],
                "camera": r[2] or "",
                "confidence": round(r[3] * 100, 1) if r[3] else 0,
                "time": time_str,
                "notification_sent": bool(r[5]) if r[5] is not None else False,
            })
    finally:
        await db.close()

    present = len(set(a["person_id"] for a in attendance))
    total = len(teachers)
    absent = total - present

    xlsx_bytes = _generate_teacher_report_excel(teachers, attendance, today)
    filename = f"Teacher_Attendance_{today}.xlsx"

    body = (
        f"Daily Teacher/Staff Attendance Report — {today_display}\n\n"
        f"Present: {present} / {total}\n"
        f"Absent: {absent}\n\n"
        f"Please find the detailed report attached.\n\n"
        f"— PPIS Automated Attendance System"
    )

    recipients = [r.strip() for r in REPORT_RECIPIENTS.split(",") if r.strip()]
    results = []
    for email in recipients:
        ok = await send_email_async(
            email,
            f"Teacher Attendance Report — {today_display}",
            body,
            "PP International School",
            attachments=[(filename, xlsx_bytes)],
        )
        results.append({"email": email, "sent": ok})

    return {
        "status": "ok",
        "date": today,
        "present": present,
        "absent": absent,
        "total": total,
        "email_results": results,
    }


@router.get("/attendance/monitoring")
async def attendance_monitoring():
    """AI monitoring dashboard: track accuracy, false positives, camera health.

    Returns metrics for:
    - False positive detections
    - Camera-wise accuracy
    - Confidence distribution
    - Notification logs
    - Daily accuracy percentage
    """
    db = await get_db()
    try:
        today_filter = "date(logged_at) = date('now', '+5 hours', '+30 minutes')"

        # Total attendance records today
        cursor = await db.execute(
            f"SELECT COUNT(*) FROM attendance_records WHERE {today_filter}"
        )
        total_records = (await cursor.fetchone())[0]

        # Confidence distribution
        cursor = await db.execute(
            f"SELECT "
            f"  SUM(CASE WHEN confidence >= 0.60 THEN 1 ELSE 0 END) as high_conf, "
            f"  SUM(CASE WHEN confidence >= 0.45 AND confidence < 0.60 THEN 1 ELSE 0 END) as medium_conf, "
            f"  SUM(CASE WHEN confidence < 0.45 THEN 1 ELSE 0 END) as low_conf, "
            f"  AVG(confidence) as avg_conf, "
            f"  MIN(confidence) as min_conf, "
            f"  MAX(confidence) as max_conf "
            f"FROM attendance_records WHERE {today_filter}"
        )
        conf_row = await cursor.fetchone()
        confidence_stats = {
            "high_confidence_count": conf_row[0] or 0,
            "medium_confidence_count": conf_row[1] or 0,
            "low_confidence_count": conf_row[2] or 0,
            "average": round((conf_row[3] or 0) * 100, 1),
            "minimum": round((conf_row[4] or 0) * 100, 1),
            "maximum": round((conf_row[5] or 0) * 100, 1),
        }

        # Camera-wise breakdown
        cursor = await db.execute(
            f"SELECT camera_label, COUNT(*) as detections, "
            f"AVG(confidence) as avg_conf, MIN(confidence) as min_conf "
            f"FROM attendance_records WHERE {today_filter} "
            f"GROUP BY camera_label ORDER BY detections DESC"
        )
        camera_stats = [
            {
                "camera": r[0],
                "detections": r[1],
                "avg_confidence": round((r[2] or 0) * 100, 1),
                "min_confidence": round((r[3] or 0) * 100, 1),
            }
            for r in await cursor.fetchall()
        ]

        # Notification stats
        cursor = await db.execute(
            f"SELECT "
            f"  SUM(CASE WHEN notification_sent = 1 THEN 1 ELSE 0 END) as sent, "
            f"  SUM(CASE WHEN notification_sent = 0 OR notification_sent IS NULL THEN 1 ELSE 0 END) as pending "
            f"FROM attendance_records WHERE {today_filter}"
        )
        notif_row = await cursor.fetchone()
        notification_stats = {
            "sent": notif_row[0] or 0,
            "pending": notif_row[1] or 0,
        }

        # Duplicate detection check (should be 0 with proper dedup)
        cursor = await db.execute(
            f"SELECT person_id, COUNT(*) as cnt "
            f"FROM attendance_records WHERE {today_filter} "
            f"GROUP BY person_id HAVING cnt > 1"
        )
        duplicates = [{"person_id": r[0], "count": r[1]}
                      for r in await cursor.fetchall()]

        # Teacher vs student breakdown
        cursor = await db.execute(
            f"SELECT "
            f"  SUM(CASE WHEN person_id LIKE 'TEACHER_%' THEN 1 ELSE 0 END) as teachers, "
            f"  SUM(CASE WHEN person_id NOT LIKE 'TEACHER_%' THEN 1 ELSE 0 END) as students "
            f"FROM attendance_records WHERE {today_filter}"
        )
        role_row = await cursor.fetchone()

        return {
            "date": datetime.now().strftime("%Y-%m-%d"),
            "total_records": total_records,
            "teachers_present": role_row[0] or 0,
            "students_present": role_row[1] or 0,
            "confidence_stats": confidence_stats,
            "camera_stats": camera_stats,
            "notification_stats": notification_stats,
            "duplicate_detections": duplicates,
            "system_checks": {
                "multi_frame_verification": "3 sightings required",
                "confidence_threshold_student": "40%",
                "confidence_threshold_teacher": "45%",
                "entry_validation": "required",
                "anti_spoofing": "enabled",
                "quality_filtering": "sharpness + brightness",
                "time_window_student": "8:00-9:00 AM",
                "time_window_teacher": "7:00-7:45 AM",
            },
            "alerts": _generate_monitoring_alerts(
                confidence_stats, camera_stats, duplicates, total_records
            ),
        }
    finally:
        await db.close()


def _generate_monitoring_alerts(
    confidence_stats: dict, camera_stats: list,
    duplicates: list, total_records: int,
) -> list[dict]:
    """Generate automatic alerts for monitoring issues."""
    alerts = []
    if confidence_stats["average"] < 45 and total_records > 0:
        alerts.append({
            "level": "warning",
            "message": f"Average confidence is low ({confidence_stats['average']}%). "
                       f"Check camera quality and lighting.",
        })
    if confidence_stats["low_confidence_count"] > total_records * 0.2 and total_records > 5:
        alerts.append({
            "level": "warning",
            "message": f"{confidence_stats['low_confidence_count']} detections below 45% confidence. "
                       f"Review camera positions.",
        })
    if duplicates:
        alerts.append({
            "level": "error",
            "message": f"{len(duplicates)} person(s) have duplicate attendance records. "
                       f"Dedup may be failing.",
        })
    for cam in camera_stats:
        if cam["avg_confidence"] < 40 and cam["detections"] > 2:
            alerts.append({
                "level": "warning",
                "message": f"Camera '{cam['camera']}' has low avg confidence "
                           f"({cam['avg_confidence']}%). Check angle/lighting.",
            })
    if not alerts:
        alerts.append({"level": "ok", "message": "All systems operating normally."})
    return alerts


# ── Homework Docs Management ─────────────────────────────────────────────────

@router.get("/homework/docs")
async def list_homework_docs():
    """List all registered homework Google Docs."""
    from app.services.homework_delivery_service import get_registered_docs
    docs = await get_registered_docs()
    return {"docs": docs, "total": len(docs)}


@router.post("/homework/docs", dependencies=[Depends(verify_dashboard_secret)])
async def register_homework_docs(payload: dict):
    """Register homework docs in bulk.

    Expects: {"docs": [{"grade": "Grade 1A", "doc_id": "...", "doc_url": "..."}]}
    """
    from app.services.homework_delivery_service import register_homework_doc
    docs = payload.get("docs", [])
    registered = 0
    failed = 0
    for doc in docs:
        grade = doc.get("grade", "")
        doc_id = doc.get("doc_id", "")
        doc_url = doc.get("doc_url", "")
        if grade and doc_id:
            ok = await register_homework_doc(grade, doc_id, doc_url)
            if ok:
                registered += 1
            else:
                failed += 1
    return {"registered": registered, "failed": failed, "total": len(docs)}


@router.get("/homework/logs")
async def list_homework_logs(limit: int = Query(50, le=200)):
    """Get recent homework delivery logs."""
    from app.services.homework_delivery_service import get_homework_logs
    logs = await get_homework_logs(limit)
    return {"logs": logs, "total": len(logs)}


@router.post("/homework/trigger/{period}", dependencies=[Depends(verify_dashboard_secret)])
async def trigger_homework_delivery(period: int):
    """Manually trigger homework delivery for a specific period (for testing)."""
    from app.services.homework_delivery_service import run_homework_delivery
    if period < 0 or period > 6:
        raise HTTPException(status_code=400, detail="Period must be 0-6")
    result = await run_homework_delivery(period)
    return result


@router.post("/homework/clear-docs", dependencies=[Depends(verify_dashboard_secret)])
async def trigger_daily_clear():
    """Manually trigger daily doc clear (for testing)."""
    from app.services.homework_delivery_service import daily_clear_all_docs
    result = await daily_clear_all_docs()
    return result


@router.post("/homework/send-review-copies", dependencies=[Depends(verify_dashboard_secret)])
async def send_review_copies(payload: dict):
    """Send copies of today's delivered homework to a review phone number."""
    import re as _re
    import asyncio as _asyncio
    from app.services.whatsapp_service import send_cloud_template_message

    phone = payload.get("phone", "")
    if not phone:
        raise HTTPException(status_code=400, detail="phone required")

    db = await get_db()
    rows = await db.execute_fetchall(
        "SELECT DISTINCT grade, period, content FROM homework_delivery_logs "
        "WHERE date(created_at) = date('now') AND status = 'delivered' "
        "ORDER BY period, grade"
    )

    period_labels = {
        0: "Assembly (08:00-08:10)", 1: "Period 1 (08:10-08:45)",
        2: "Period 2 (09:00-09:30)", 3: "Period 3 (09:30-10:00)",
        4: "Period 4 (10:00-10:30)", 5: "Period 5 (10:30-11:00)",
        6: "Period 6 (11:00-11:30)",
    }

    seen: set[tuple[str, int]] = set()
    sent = 0
    failed = 0
    details = []
    for row in rows:
        grade, period, content = row[0], row[1], row[2]
        key = (grade, period)
        if key in seen:
            continue
        seen.add(key)

        hw = content
        hw = hw.replace("\u2013", "-").replace("\u2014", "-")
        hw = hw.replace("\r\n", " | ").replace("\n", " | ").replace("\r", " | ")
        hw = hw.replace("\t", " ")
        hw = _re.sub(r"[ ]{4,}", "   ", hw)
        hw = _re.sub(r"[ ]+", " ", hw)
        hw = hw.strip()[:900]

        label = period_labels.get(period, f"Period {period}")
        ok = await send_cloud_template_message(
            phone, "ppis_classwork_homework",
            body_params=[grade, label, hw],
        )
        if ok:
            sent += 1
        else:
            failed += 1
        details.append({"grade": grade, "period": period, "status": "ok" if ok else "failed"})
        await _asyncio.sleep(1)

    return {"phone": phone, "sent": sent, "failed": failed, "details": details}
