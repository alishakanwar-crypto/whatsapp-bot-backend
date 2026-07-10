"""
Mood & Temperament API (Multi-Person)
======================================
Receives mood observations from the campus agent, stores them,
and generates daily mood/temperament reports at 12:00 PM IST.

Supports multiple tracked persons (Chairman, Alisha, etc.).
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import os
from collections import Counter
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Request

logger = logging.getLogger("chairman_mood")

IST = timezone(timedelta(hours=5, minutes=30))
router = APIRouter()

REPORT_EMAIL = os.environ.get("MOOD_REPORT_EMAIL", "alisha.kanwar@ppischool.in")
MOOD_REPORTS_ENABLED = os.environ.get("MOOD_REPORTS_ENABLED", "false").lower() in (
    "1", "true", "yes",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_db():
    from app.database import get_db
    return await get_db()


async def _get_historical_mood(today: str, days: int = 7) -> dict[str, dict]:
    """Fetch mood summary for each person over the last N days (excluding today).

    Returns dict: person_name -> {usual_mood, avg_intensity, avg_days, deviation}
    """
    from datetime import datetime, timedelta
    try:
        today_dt = datetime.strptime(today, "%Y-%m-%d")
    except ValueError:
        return {}
    start_date = (today_dt - timedelta(days=days)).strftime("%Y-%m-%d")

    db = await _get_db()
    try:
        cur = await db.execute(
            "SELECT person, dominant_emotion, intensity, date "
            "FROM chairman_mood_log "
            "WHERE date >= ? AND date < ? ORDER BY person, date",
            (start_date, today),
        )
        rows = await cur.fetchall()
    finally:
        await db.close()

    if not rows:
        return {}

    by_person: dict[str, list] = {}
    for r in rows:
        by_person.setdefault(r[0], []).append({
            "emotion": r[1], "intensity": r[2], "date": r[3],
        })

    result: dict[str, dict] = {}
    for pname, obs in by_person.items():
        emotions = [o["emotion"] for o in obs]
        intensities = [o["intensity"] for o in obs]
        unique_dates = len(set(o["date"] for o in obs))
        usual_mood = Counter(emotions).most_common(1)[0][0]
        avg_intensity = round(sum(intensities) / len(intensities), 1)

        result[pname] = {
            "usual_mood": usual_mood,
            "avg_intensity": avg_intensity,
            "avg_days": unique_dates,
            "deviation": None,
        }

    return result


def _generate_mood_alerts(person_reports: dict,
                          historical: dict) -> list[dict]:
    """Generate mood alerts based on current report + historical comparison."""
    alerts: list[dict] = []

    for pname, pr in person_reports.items():
        mood = pr.get("dominant_mood", "neutral")
        intensity = pr.get("avg_intensity", 0)
        neg_emotions = {"angry", "sad", "fear", "disgust"}

        # High priority: repeated negative emotions
        neg_pct = sum(
            pr.get("emotion_distribution", {}).get(e, 0) for e in neg_emotions
        )
        if neg_pct > 60:
            alerts.append({
                "severity": "HIGH",
                "person": pname,
                "detail": f"Predominantly negative mood today ({neg_pct:.0f}% negative emotions). "
                          f"Dominant: {mood}, Intensity: {intensity}%",
            })
        elif neg_pct > 30:
            alerts.append({
                "severity": "MEDIUM",
                "person": pname,
                "detail": f"Elevated negative mood ({neg_pct:.0f}% negative). "
                          f"Dominant: {mood}",
            })

        # Historical deviation
        hist = historical.get(pname, {})
        if hist.get("avg_days", 0) >= 3:
            usual = hist.get("usual_mood", "neutral")
            if usual in {"happy", "neutral"} and mood in neg_emotions:
                hist["deviation"] = (
                    f"Usually {usual}, but today showing {mood}. "
                    f"May indicate stress or mood change."
                )
                alerts.append({
                    "severity": "MEDIUM",
                    "person": pname,
                    "detail": hist["deviation"],
                })
            usual_intensity = hist.get("avg_intensity", 0)
            if intensity > 0 and usual_intensity > 0:
                if intensity > usual_intensity * 1.5:
                    alerts.append({
                        "severity": "LOW",
                        "person": pname,
                        "detail": f"Expression intensity ({intensity}%) significantly "
                                  f"higher than usual ({usual_intensity}%).",
                    })

    return alerts


# ---------------------------------------------------------------------------
# POST /api/chairman/mood — receive mood observation from campus agent
# ---------------------------------------------------------------------------

@router.post("/api/chairman/mood")
async def receive_mood(request: Request):
    data = await request.json()
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    person = data.get("person", "Chairman")
    timestamp = data.get("timestamp", now.strftime("%Y-%m-%d %H:%M:%S"))
    camera = data.get("camera", "unknown")
    dominant_emotion = data.get("dominant_emotion", "unknown")
    emotions = data.get("emotions", {})
    temperament = data.get("temperament", "neutral")
    intensity = data.get("intensity", 0.0)
    face_distance = data.get("face_distance", 0.0)
    face_confidence = data.get("face_confidence", 0.0)
    face_crop = data.get("face_crop", "")
    mood_category = data.get("mood_category", "Neutral")
    mood_label = data.get("mood_label", "Normal")
    frame_count = data.get("frame_count", 1)
    agreement = data.get("agreement", 0.0)
    negative_ratio = data.get("negative_ratio", 0.0)
    description = data.get("description", "")

    # If face_crop is provided and emotion is generic, re-analyze with GPT-4o
    generic_emotions = {"detected", "unknown", "neutral", ""}
    if face_crop and len(face_crop) > 100 and dominant_emotion in generic_emotions:
        try:
            from app.routes.trueface import _analyze_face_mood
            mood = await _analyze_face_mood(face_crop)
            dominant_emotion = mood.get("dominant_emotion", dominant_emotion)
            emotions = mood.get("emotions", emotions)
            temperament = mood.get("temperament", temperament)
            intensity = float(mood.get("intensity", intensity))
            description = mood.get("description", "")
            mood_category = {
                "happy": "Positive", "confident": "Positive", "cheerful": "Positive",
                "sad": "Negative", "angry": "Negative", "anxious": "Negative",
                "fearful": "Negative", "stressed": "Negative",
            }.get(dominant_emotion, "Neutral")
            mood_label = dominant_emotion.title()
        except Exception as e:
            logger.warning("[MOOD] GPT-4o re-analysis failed: %s", e)

    db = await _get_db()
    try:
        await db.execute(
            "INSERT INTO chairman_mood_log "
            "(person, date, timestamp, camera, dominant_emotion, emotions_json, "
            "temperament, intensity, face_distance, face_confidence, face_crop, "
            "mood_category, mood_label, frame_count, agreement, negative_ratio, "
            "description) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                person, today, timestamp, camera, dominant_emotion,
                json.dumps(emotions), temperament, intensity,
                face_distance, face_confidence, face_crop,
                mood_category, mood_label, frame_count, agreement, negative_ratio,
                description,
            ),
        )
        await db.commit()
    finally:
        await db.close()

    logger.info(
        "[MOOD] Logged: %s — %s on %s — %s (%s) intensity=%.1f",
        person, timestamp, camera, dominant_emotion, temperament, intensity,
    )

    # Mood reports permanently disabled per user request
    # if person.lower() == "chairman":
    #     asyncio.ensure_future(_send_chairman_instant_report())

    return {"status": "ok", "person": person, "emotion": dominant_emotion, "temperament": temperament}


# ---------------------------------------------------------------------------
# GET /api/chairman/mood/today — today's mood summary (all persons)
# ---------------------------------------------------------------------------

@router.get("/api/chairman/mood/today")
async def today_mood(person: str | None = None):
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    db = await _get_db()
    try:
        if person:
            cur = await db.execute(
                "SELECT person, timestamp, camera, dominant_emotion, temperament, intensity "
                "FROM chairman_mood_log WHERE date = ? AND person = ? ORDER BY timestamp",
                (today, person),
            )
        else:
            cur = await db.execute(
                "SELECT person, timestamp, camera, dominant_emotion, temperament, intensity "
                "FROM chairman_mood_log WHERE date = ? ORDER BY timestamp",
                (today,),
            )
        rows = await cur.fetchall()
    finally:
        await db.close()

    observations = [
        {
            "person": r[0],
            "timestamp": r[1],
            "camera": r[2],
            "emotion": r[3],
            "temperament": r[4],
            "intensity": r[5],
        }
        for r in rows
    ]

    # Build per-person summaries
    by_person: dict[str, list] = {}
    for o in observations:
        by_person.setdefault(o["person"], []).append(o)

    person_summaries = {}
    for pname, obs_list in by_person.items():
        emotions = [o["emotion"] for o in obs_list]
        temperaments = [o["temperament"] for o in obs_list]
        emotion_counts = Counter(emotions)
        temperament_counts = Counter(temperaments)
        total = len(obs_list)

        person_summaries[pname] = {
            "total_observations": total,
            "dominant_mood": emotion_counts.most_common(1)[0][0],
            "dominant_temperament": temperament_counts.most_common(1)[0][0],
            "avg_intensity": round(sum(o["intensity"] for o in obs_list) / total, 1),
            "emotion_distribution": {k: round(v / total * 100, 1) for k, v in emotion_counts.items()},
            "temperament_distribution": {k: round(v / total * 100, 1) for k, v in temperament_counts.items()},
        }

    return {
        "date": today,
        "total_observations": len(observations),
        "person_summaries": person_summaries,
        "observations": observations,
    }


# ---------------------------------------------------------------------------
# GET /api/chairman/mood/report/{date} — full report data
# ---------------------------------------------------------------------------

@router.get("/api/chairman/mood/report/{date}")
async def mood_report(date: str, person: str | None = None):
    db = await _get_db()
    try:
        _cols = ("person, timestamp, camera, dominant_emotion, emotions_json, "
                 "temperament, intensity, face_confidence, description")
        if person:
            cur = await db.execute(
                f"SELECT {_cols} FROM chairman_mood_log "
                "WHERE date = ? AND person = ? ORDER BY timestamp",
                (date, person),
            )
        else:
            cur = await db.execute(
                f"SELECT {_cols} FROM chairman_mood_log "
                "WHERE date = ? ORDER BY timestamp",
                (date,),
            )
        rows = await cur.fetchall()
    finally:
        await db.close()

    observations = []
    for r in rows:
        emotions = json.loads(r[4]) if r[4] else {}
        observations.append({
            "person": r[0],
            "timestamp": r[1],
            "camera": r[2],
            "dominant_emotion": r[3],
            "emotions": emotions,
            "temperament": r[5],
            "intensity": r[6],
            "face_confidence": r[7],
            "description": r[8] if len(r) > 8 else "",
        })

    if not observations:
        return {"date": date, "total_observations": 0, "summary": "No data"}

    # Build per-person report
    by_person: dict[str, list] = {}
    for obs in observations:
        by_person.setdefault(obs["person"], []).append(obs)

    person_reports = {}
    for pname, obs_list in by_person.items():
        # Hourly breakdown
        hourly: dict[str, list] = {}
        for obs in obs_list:
            try:
                hour = obs["timestamp"].split(" ")[1][:2] + ":00"
            except (IndexError, AttributeError):
                hour = "unknown"
            hourly.setdefault(hour, []).append(obs)

        hourly_summary = {}
        for hour, h_obs in sorted(hourly.items()):
            emos = [o["dominant_emotion"] for o in h_obs]
            temps = [o["temperament"] for o in h_obs]
            hourly_summary[hour] = {
                "observations": len(h_obs),
                "dominant_emotion": Counter(emos).most_common(1)[0][0],
                "dominant_temperament": Counter(temps).most_common(1)[0][0],
                "avg_intensity": round(
                    sum(o["intensity"] for o in h_obs) / len(h_obs), 1
                ),
            }

        all_emotions = [o["dominant_emotion"] for o in obs_list]
        all_temperaments = [o["temperament"] for o in obs_list]
        total = len(obs_list)
        emotion_counts = Counter(all_emotions)
        temperament_counts = Counter(all_temperaments)

        best_hour = min(hourly_summary, key=lambda h: hourly_summary[h]["avg_intensity"]) if hourly_summary else "N/A"

        person_reports[pname] = {
            "total_observations": total,
            "dominant_mood": emotion_counts.most_common(1)[0][0],
            "dominant_temperament": temperament_counts.most_common(1)[0][0],
            "avg_intensity": round(sum(o["intensity"] for o in obs_list) / total, 1),
            "emotion_distribution": {
                k: round(v / total * 100, 1) for k, v in emotion_counts.items()
            },
            "temperament_distribution": {
                k: round(v / total * 100, 1) for k, v in temperament_counts.items()
            },
            "hourly_summary": hourly_summary,
            "best_time_to_approach": best_hour,
            "observations": [
                {
                    "time": o["timestamp"].split(" ")[1] if " " in o.get("timestamp", "") else o.get("timestamp", ""),
                    "camera": o["camera"],
                    "emotion": o["dominant_emotion"],
                    "temperament": o["temperament"],
                    "intensity": o["intensity"],
                    "description": o.get("description", ""),
                }
                for o in obs_list
            ],
        }

    return {
        "date": date,
        "total_observations": len(observations),
        "person_reports": person_reports,
    }


# ---------------------------------------------------------------------------
# Hourly Report (called by scheduler every hour 7 AM - 12 PM IST)
# ---------------------------------------------------------------------------

def _generate_mood_excel(date: str, person_reports: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    first_sheet = True
    for pname, report in person_reports.items():
        if first_sheet:
            ws = wb.active
            ws.title = pname
            first_sheet = False
        else:
            ws = wb.create_sheet(pname)

        ws.cell(row=1, column=1, value=f"{pname} — Mood Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Date: {date}")
        ws.cell(row=3, column=1, value=f"Total Observations: {report['total_observations']}")
        ws.cell(row=4, column=1, value=f"Dominant Mood: {report['dominant_mood']}")
        ws.cell(row=5, column=1, value=f"Dominant Temperament: {report['dominant_temperament']}")
        ws.cell(row=6, column=1, value=f"Avg Expression Intensity: {report['avg_intensity']}%")
        ws.cell(row=7, column=1, value=f"Best Time to Approach: {report.get('best_time_to_approach', 'N/A')}")

        # Emotion distribution
        row = 9
        ws.cell(row=row, column=1, value="Emotion").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Percentage").font = Font(bold=True)
        for emo, pct in sorted(report.get("emotion_distribution", {}).items(), key=lambda x: -x[1]):
            row += 1
            ws.cell(row=row, column=1, value=emo).border = border
            ws.cell(row=row, column=2, value=f"{pct}%").border = border

        # Temperament distribution
        row += 2
        ws.cell(row=row, column=1, value="Temperament").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Percentage").font = Font(bold=True)
        for temp, pct in sorted(report.get("temperament_distribution", {}).items(), key=lambda x: -x[1]):
            row += 1
            ws.cell(row=row, column=1, value=temp).border = border
            ws.cell(row=row, column=2, value=f"{pct}%").border = border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

        # Hourly breakdown
        row += 2
        ws.cell(row=row, column=1, value="Hourly Breakdown").font = Font(bold=True, size=12)
        row += 1
        headers = ["Hour", "Observations", "Dominant Mood", "Dominant Temperament", "Avg Intensity"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for hour, data in sorted(report.get("hourly_summary", {}).items()):
            row += 1
            ws.cell(row=row, column=1, value=hour).border = border
            ws.cell(row=row, column=2, value=data["observations"]).border = border
            ws.cell(row=row, column=3, value=data["dominant_emotion"]).border = border
            ws.cell(row=row, column=4, value=data["dominant_temperament"]).border = border
            ws.cell(row=row, column=5, value=f"{data['avg_intensity']}%").border = border

        # Individual observations detail
        obs_list = report.get("observations", [])
        if obs_list:
            row += 2
            ws.cell(row=row, column=1, value="Observation Details").font = Font(bold=True, size=12)
            row += 1
            detail_headers = ["#", "Time", "Camera", "Emotion", "Temperament", "Intensity", "Description"]
            for col, h in enumerate(detail_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            for idx, obs in enumerate(obs_list, 1):
                row += 1
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=obs.get("time", "")).border = border
                ws.cell(row=row, column=3, value=obs.get("camera", "")).border = border
                ws.cell(row=row, column=4, value=obs.get("emotion", "")).border = border
                ws.cell(row=row, column=5, value=obs.get("temperament", "")).border = border
                ws.cell(row=row, column=6, value=f"{obs.get('intensity', 0)}%").border = border
                ws.cell(row=row, column=7, value=obs.get("description", "")).border = border

            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 15
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Track whether we already sent an instant report today to avoid spam
_chairman_instant_sent: dict[str, bool] = {}


async def _send_chairman_instant_report():
    """Send an immediate mood report when Chairman is first detected today."""
    if not MOOD_REPORTS_ENABLED:
        logger.info("[MOOD] Chairman instant report disabled")
        return

    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    if _chairman_instant_sent.get(today):
        return
    _chairman_instant_sent[today] = True
    # Clean old dates
    for d in list(_chairman_instant_sent):
        if d != today:
            del _chairman_instant_sent[d]

    # Small delay to let the observation be fully committed to DB
    import asyncio
    await asyncio.sleep(2)

    today_display = now.strftime("%d-%m-%Y")
    time_display = now.strftime("%I:%M %p")

    db = await _get_db()
    try:
        cur = await db.execute(
            "SELECT person, timestamp, camera, dominant_emotion, temperament, intensity "
            "FROM chairman_mood_log WHERE date = ? AND person = 'Chairman' ORDER BY timestamp",
            (today,),
        )
        rows = await cur.fetchall()
    finally:
        await db.close()

    if not rows:
        return

    observations = [
        {"person": r[0], "timestamp": r[1], "camera": r[2],
         "emotion": r[3], "temperament": r[4], "intensity": r[5]}
        for r in rows
    ]

    latest = observations[-1]
    body = (
        f"⚡ CHAIRMAN DETECTED — Instant Mood Report\n"
        f"{today_display} at {time_display} IST\n\n"
        f"Chairman Rahul Gupta has been detected!\n\n"
        f"Latest Observation:\n"
        f"  Camera: {latest['camera']}\n"
        f"  Time: {latest['timestamp']}\n"
        f"  Mood: {latest['emotion']}\n"
        f"  Temperament: {latest['temperament']}\n"
        f"  Intensity: {latest['intensity']}%\n\n"
    )

    if len(observations) > 1:
        body += f"Total observations today: {len(observations)}\n"
        emotions = [o["emotion"] for o in observations]
        dominant = Counter(emotions).most_common(1)[0][0]
        body += f"Dominant mood today: {dominant}\n\n"

    body += "Full report will follow in the next hourly mood report.\n\n"
    body += "— PPIS Mood & Temperament Monitor"

    from app.services.email_service import send_email_async
    recipients = [r.strip() for r in REPORT_EMAIL.split(",") if r.strip()]
    for email in recipients:
        ok = await send_email_async(
            email,
            f"⚡ Chairman Detected — {today_display} {time_display} IST",
            body,
            "PP International School",
        )
        logger.info("[MOOD] Chairman instant report → %s: %s", email, "OK" if ok else "FAILED")

    logger.info("[MOOD] Chairman instant mood report sent at %s", time_display)


async def send_daily_mood_report():
    if not MOOD_REPORTS_ENABLED:
        logger.info("[MOOD] Daily report disabled")
        return

    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    today_display = now.strftime("%d-%m-%Y")
    time_display = now.strftime("%I:%M %p")
    current_hour = now.strftime("%H:00")

    db = await _get_db()
    try:
        cur = await db.execute(
            "SELECT COUNT(*) FROM chairman_mood_log WHERE date = ?", (today,),
        )
        row = await cur.fetchone()
        count = row[0] if row else 0
    finally:
        await db.close()

    if count == 0:
        logger.info("[MOOD] No mood observations for %s — skipping report", today)
        return

    import httpx
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                f"https://ppis-whatsapp-bot.fly.dev/api/chairman/mood/report/{today}"
            )
            report = resp.json()
    except Exception as e:
        logger.error("[MOOD] Failed to fetch report data: %s", e)
        return

    if report.get("total_observations", 0) == 0:
        logger.info("[MOOD] No observations in report for %s", today)
        return

    person_reports = report.get("person_reports", {})
    if not person_reports:
        logger.info("[MOOD] No person reports for %s", today)
        return

    xlsx_bytes = _generate_mood_excel(today, person_reports)
    filename = f"Mood_Report_{today}_{now.strftime('%H%M')}.xlsx"

    persons_tracked = list(person_reports.keys())
    body = (
        f"Hourly Mood Report — {today_display} at {time_display} IST\n"
        f"Persons tracked: {', '.join(persons_tracked)}\n\n"
    )

    # Fetch historical data for comparison (last 7 days)
    historical_summary = await _get_historical_mood(today, 7)

    for pname, pr in person_reports.items():
        hourly = pr.get("hourly_summary", {})
        # Highlight current hour
        this_hour = hourly.get(current_hour, {})
        if this_hour:
            body += (
                f"--- {pname} (This Hour: {current_hour}) ---\n"
                f"  Observations this hour: {this_hour.get('observations', 0)}\n"
                f"  Mood: {this_hour.get('dominant_emotion', 'N/A')}\n"
                f"  Temperament: {this_hour.get('dominant_temperament', 'N/A')}\n"
                f"  Intensity: {this_hour.get('avg_intensity', 0)}%\n\n"
            )
        else:
            body += f"--- {pname} (No observations this hour) ---\n\n"

        body += (
            f"Day Summary ({pr['total_observations']} total observations):\n"
            f"  Dominant Mood: {pr['dominant_mood']}\n"
            f"  Dominant Temperament: {pr['dominant_temperament']}\n"
            f"  Avg Intensity: {pr['avg_intensity']}%\n"
            f"  Best Time to Approach: {pr.get('best_time_to_approach', 'N/A')}\n\n"
        )

        # Include recent observation details (last 10)
        obs_list = pr.get("observations", [])
        if obs_list:
            body += "Recent Observations:\n"
            for obs in obs_list[-10:]:
                desc = obs.get("description", "")
                desc_str = f" — {desc}" if desc else ""
                body += (
                    f"  {obs.get('time', '?')} | "
                    f"{obs.get('emotion', '?')} | "
                    f"{obs.get('temperament', '?')} | "
                    f"{obs.get('intensity', 0)}%"
                    f"{desc_str}\n"
                )
            body += "\n"

        # Historical comparison
        hist = historical_summary.get(pname, {})
        if hist.get("avg_days", 0) > 0:
            body += (
                f"Historical Comparison (last {hist['avg_days']} days):\n"
                f"  Usual Mood: {hist.get('usual_mood', 'N/A')}\n"
                f"  Usual Intensity: {hist.get('avg_intensity', 0)}%\n"
            )
            if hist.get("deviation"):
                body += f"  ⚠ DEVIATION: {hist['deviation']}\n"
            body += "\n"

    # Alerts
    alerts = _generate_mood_alerts(person_reports, historical_summary)
    if alerts:
        body += "════════ MOOD ALERTS ════════\n"
        for a in alerts:
            body += f"  [{a['severity']}] {a['person']}: {a['detail']}\n"
        body += "\n"

    body += "— PPIS Mood & Temperament Monitor"

    from app.services.email_service import send_email_async
    recipients = [r.strip() for r in REPORT_EMAIL.split(",") if r.strip()]
    for email in recipients:
        ok = await send_email_async(
            email,
            f"Hourly Mood Report — {today_display} {time_display} IST",
            body,
            "PP International School",
            attachments=[(filename, xlsx_bytes)],
        )
        logger.info("[MOOD] Hourly report → %s: %s", email, "OK" if ok else "FAILED")

    logger.info(
        "[MOOD] Hourly report sent at %s: %d observations for %s",
        time_display, report["total_observations"], ", ".join(persons_tracked),
    )


def send_daily_mood_report_sync():
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            asyncio.ensure_future(send_daily_mood_report())
        else:
            loop.run_until_complete(send_daily_mood_report())
    except RuntimeError:
        asyncio.run(send_daily_mood_report())
