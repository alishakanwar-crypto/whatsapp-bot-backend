"""
TrueFace 3000 Attendance System
===============================
Receives face-recognition attendance events from a browser-based poller
running on the school PC (reads the TrueFace device's Search Records table).

Flow:
  1. Browser JS poller detects new OK face scan → POST /api/trueface/event
  2. First detection of the day for a teacher = "arrival" → WhatsApp sent
  3. Second detection = "departure" → WhatsApp sent
  4. 8:00 AM IST → Excel arrival report emailed to leave@ppischool.in
  5. 3:00 PM IST → Excel departure report emailed to leave@ppischool.in

Also keeps the legacy ADMS endpoints in case the device push protocol
is enabled in the future.
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import os
from datetime import datetime, timezone, timedelta
from pathlib import Path

from fastapi import APIRouter, Request
from fastapi.responses import PlainTextResponse

logger = logging.getLogger("app.trueface")

IST = timezone(timedelta(hours=5, minutes=30))
router = APIRouter()

REPORT_RECIPIENTS = os.environ.get(
    "TRUEFACE_REPORT_EMAIL", "leave@ppischool.in"
)

CHAIRMAN_PHONE = os.environ.get("TRUEFACE_CHAIRMAN_PHONE", "919971166562")
CHAIRMAN_TEMPLATE = os.environ.get(
    "TRUEFACE_CHAIRMAN_TEMPLATE", "ppis_chairman_teacher_arrival"
)

# Kill switch: set TRUEFACE_WHATSAPP_DISABLED=true to suppress all WhatsApp sends
WHATSAPP_DISABLED = os.environ.get("TRUEFACE_WHATSAPP_DISABLED", "").lower() in ("true", "1", "yes")

# Before this hour (IST), all detections count as arrival only.
# Departure is only recorded after this hour.
DEPARTURE_HOUR = int(os.environ.get("TRUEFACE_DEPARTURE_HOUR", "11"))

# Persons whose mood should be logged from TrueFace photos
MOOD_TRACKED_NAMES = {
    "RAHUL GUPTA": "Chairman",
    "ALISHA AHUJA": "Alisha",
    "ALISHA KANWAR": "Alisha",
}


# ============================================================
# Database helpers
# ============================================================

async def _get_db():
    from app.database import get_db
    return await get_db()


async def _get_teacher(db, pin: str) -> dict | None:
    cur = await db.execute(
        "SELECT pin, name, phone FROM trueface_teachers WHERE pin = ?", (pin,)
    )
    row = await cur.fetchone()
    if row:
        return {"pin": row[0], "name": row[1], "phone": row[2]}
    return None


async def _get_all_teachers(db) -> list[dict]:
    cur = await db.execute(
        "SELECT pin, name, phone FROM trueface_teachers ORDER BY name"
    )
    return [{"pin": r[0], "name": r[1], "phone": r[2]} for r in await cur.fetchall()]


async def _get_attendance_record(db, pin: str, date: str) -> dict | None:
    cur = await db.execute(
        "SELECT id, pin, name, date, arrival_time, departure_time, "
        "arrival_whatsapp, departure_whatsapp "
        "FROM trueface_attendance WHERE pin = ? AND date = ?",
        (pin, date),
    )
    row = await cur.fetchone()
    if row:
        return {
            "id": row[0], "pin": row[1], "name": row[2], "date": row[3],
            "arrival_time": row[4], "departure_time": row[5],
            "arrival_whatsapp": row[6], "departure_whatsapp": row[7],
        }
    return None


async def _get_all_attendance(db, date: str) -> list[dict]:
    cur = await db.execute(
        "SELECT pin, name, arrival_time, departure_time, "
        "arrival_whatsapp, departure_whatsapp "
        "FROM trueface_attendance WHERE date = ? ORDER BY arrival_time",
        (date,),
    )
    return [
        {"pin": r[0], "name": r[1], "arrival_time": r[2], "departure_time": r[3],
         "arrival_whatsapp": r[4], "departure_whatsapp": r[5]}
        for r in await cur.fetchall()
    ]


async def _auto_register_from_dvr(db, pin: str, evt_name: str) -> dict | None:
    """Try to match a TrueFace event name against the DVR teacher database.

    If the name matches a TEACHER_* or PRINCIPAL_* entry, auto-register
    the teacher with their TrueFace PIN and phone number.
    """
    if not evt_name:
        return None

    # Normalize name for matching: lowercase, strip extra spaces
    norm = evt_name.strip().lower()

    try:
        cur = await db.execute(
            "SELECT person_id, name, phone FROM agent_registered_faces "
            "WHERE (person_id LIKE 'TEACHER_%' OR person_id LIKE 'PRINCIPAL_%') "
            "AND phone IS NOT NULL AND phone != ''"
        )
        rows = await cur.fetchall()
    except Exception as e:
        logger.warning(f"[TRUEFACE] Auto-register DB query failed: {e}")
        return None

    for r in rows:
        db_name = (r[1] or "").strip()
        db_phone = (r[2] or "").split(",")[0].strip()

        # Match: exact (case-insensitive) or contained
        if db_name.lower() == norm or norm in db_name.lower() or db_name.lower() in norm:
            # Register this teacher with their TrueFace PIN
            try:
                await db.execute(
                    "INSERT OR REPLACE INTO trueface_teachers (pin, name, phone) "
                    "VALUES (?, ?, ?)",
                    (pin, db_name, db_phone),
                )
                await db.commit()
                logger.info(
                    f"[TRUEFACE] Auto-registered: PIN={pin} name={db_name} "
                    f"phone={db_phone} (matched from DVR: {r[0]})"
                )
                return {"pin": pin, "name": db_name, "phone": db_phone}
            except Exception as e:
                logger.error(f"[TRUEFACE] Auto-register insert failed: {e}")
                return None

    # Fallback: try matching against the contacts sheet
    result = await _auto_register_from_contacts(db, pin, evt_name)
    if result:
        return result

    logger.info(f"[TRUEFACE] No match for PIN={pin} name='{evt_name}' in DVR or contacts")
    return None


async def _auto_register_from_contacts(db, pin: str, evt_name: str) -> dict | None:
    """Match a TrueFace event against the uploaded contact sheet.

    Uses priority matching:
    0. PIN-based match (highest priority — exact employee ID match)
    1. Exact full name match
    2. Full name contained in contact name or vice versa
    3. First name match (only if exactly ONE contact matches)
    If multiple contacts match at the same priority, skip (ambiguous).
    """
    if not evt_name:
        return None

    norm = evt_name.strip().lower()
    norm_parts = norm.split()

    try:
        cur = await db.execute("SELECT pin, name, phone FROM trueface_contacts")
        rows = await cur.fetchall()
    except Exception as e:
        logger.warning(f"[TRUEFACE] Contact sheet lookup failed: {e}")
        return None

    # Priority 0: PIN-based match (exact employee ID)
    if pin:
        for r in rows:
            contact_pin = (r[0] or "").strip()
            if contact_pin and contact_pin == pin:
                contact_name = (r[1] or "").strip()
                return await _register_contact_match(
                    db, pin, evt_name, contact_name, r[2])

    # Priority 1: Exact full name match
    for r in rows:
        contact_name = (r[1] or "").strip()
        cn_lower = contact_name.lower()
        if cn_lower == norm:
            return await _register_contact_match(db, pin, evt_name, contact_name, r[2])

    # Priority 2: Full name substring match (one contains the other)
    substring_matches = []
    for r in rows:
        contact_name = (r[1] or "").strip()
        cn_lower = contact_name.lower()
        if norm in cn_lower or cn_lower in norm:
            substring_matches.append((contact_name, r[2]))

    if len(substring_matches) == 1:
        return await _register_contact_match(
            db, pin, evt_name, substring_matches[0][0], substring_matches[0][1])
    elif len(substring_matches) > 1:
        names = [m[0] for m in substring_matches]
        logger.warning(
            f"[TRUEFACE] Ambiguous substring match for PIN={pin} name='{evt_name}': "
            f"{names} — skipping auto-register")
        return None

    # Priority 3: First name match (only if unambiguous)
    if norm_parts:
        first_name = norm_parts[0]
        first_name_matches = []
        for r in rows:
            contact_name = (r[1] or "").strip()
            cn_parts = contact_name.lower().split()
            if cn_parts and cn_parts[0] == first_name:
                first_name_matches.append((contact_name, r[2]))
            elif len(cn_parts) > 1 and cn_parts[-1] == first_name:
                first_name_matches.append((contact_name, r[2]))

        if len(first_name_matches) == 1:
            return await _register_contact_match(
                db, pin, evt_name, first_name_matches[0][0], first_name_matches[0][1])
        elif len(first_name_matches) > 1:
            names = [m[0] for m in first_name_matches]
            logger.warning(
                f"[TRUEFACE] Ambiguous first-name match for PIN={pin} name='{evt_name}': "
                f"{names} — skipping auto-register")
            return None

    return None


async def _register_contact_match(db, pin: str, evt_name: str,
                                   contact_name: str, phone) -> dict | None:
    """Register a matched contact as a TrueFace teacher."""
    contact_phone = (phone or "").strip()
    try:
        await db.execute(
            "INSERT OR REPLACE INTO trueface_teachers (pin, name, phone) "
            "VALUES (?, ?, ?)",
            (pin, evt_name.strip(), contact_phone),
        )
        await db.commit()
        logger.info(
            f"[TRUEFACE] Auto-registered from contacts: PIN={pin} "
            f"name={evt_name.strip()} phone={contact_phone} "
            f"(matched: {contact_name})"
        )
        return {"pin": pin, "name": evt_name.strip(), "phone": contact_phone}
    except Exception as e:
        logger.error(f"[TRUEFACE] Contact register insert failed: {e}")
        return None


def _format_time_12h(timestamp: str) -> str:
    """Convert '2026-05-22 07:30:00' or '07:30:00' to '07:30 AM'."""
    try:
        if " " in timestamp:
            dt = datetime.strptime(timestamp.strip(), "%Y-%m-%d %H:%M:%S")
        else:
            dt = datetime.strptime(timestamp.strip(), "%H:%M:%S")
        return dt.strftime("%I:%M %p")
    except (ValueError, TypeError):
        return timestamp or ""


# ============================================================
# WhatsApp notifications
# ============================================================

async def _send_arrival_whatsapp(name: str, phone: str, time_str: str) -> bool:
    if WHATSAPP_DISABLED:
        logger.info("[TRUEFACE] WhatsApp DISABLED — skipping arrival for %s", name)
        return False
    from app.services.whatsapp_service import send_cloud_template_message
    display_name = name.title() if name == name.upper() else name
    logger.info(f"[TRUEFACE] Arrival WhatsApp → {phone} for {display_name} at {time_str}")
    try:
        ok = await send_cloud_template_message(
            to=phone,
            template_name="ppis_teacher_present_text",
            language_code="en",
            body_params=[display_name, time_str],
        )
        logger.info(f"[TRUEFACE] Arrival WhatsApp {'OK' if ok else 'FAILED'} for {phone}")
        return bool(ok)
    except Exception as e:
        logger.error(f"[TRUEFACE] Arrival WhatsApp error for {phone}: {e}")
        return False


async def _send_departure_whatsapp(name: str, phone: str, time_str: str) -> bool:
    if WHATSAPP_DISABLED:
        logger.info("[TRUEFACE] WhatsApp DISABLED — skipping departure for %s", name)
        return False
    from app.services.whatsapp_service import send_cloud_template_message
    display_name = name.title() if name == name.upper() else name
    logger.info(f"[TRUEFACE] Departure WhatsApp → {phone} for {display_name} at {time_str}")
    try:
        ok = await send_cloud_template_message(
            to=phone,
            template_name="ppis_teacher_departure",
            language_code="en",
            body_params=[display_name, time_str],
        )
        logger.info(f"[TRUEFACE] Departure WhatsApp {'OK' if ok else 'FAILED'} for {phone}")
        return bool(ok)
    except Exception as e:
        logger.error(f"[TRUEFACE] Departure WhatsApp error for {phone}: {e}")
        return False


async def _get_db_photo_b64(name: str) -> str:
    """Fetch the stored face photo from the DVR database for a teacher."""
    import base64

    try:
        db = await _get_db()
        norm = name.strip().lower()
        cur = await db.execute(
            "SELECT image_data FROM agent_registered_faces "
            "WHERE (person_id LIKE 'TEACHER_%' OR person_id LIKE 'PRINCIPAL_%') "
            "AND LOWER(name) LIKE ? "
            "ORDER BY registered_at DESC LIMIT 1",
            (f"%{norm}%",),
        )
        row = await cur.fetchone()
        await db.close()
        if row and row[0]:
            return base64.b64encode(row[0]).decode()
    except Exception as e:
        logger.debug("[TRUEFACE] DB photo lookup failed for %s: %s", name, e)
    return ""


async def _notify_chairman_arrival(
    name: str, time_str: str, photo_b64: str = "",
) -> bool:
    """Send arrival notification with face photo to the chairman.

    The template requires an IMAGE header. If no live photo was captured
    from the device, we fall back to the stored database photo.
    """
    if WHATSAPP_DISABLED:
        logger.info("[TRUEFACE] WhatsApp DISABLED — skipping chairman notify for %s", name)
        return False
    if not CHAIRMAN_PHONE:
        return False

    from app.services.whatsapp_service import (
        send_cloud_template_message,
        upload_base64_image_cloud,
    )

    display_name = name.title() if name == name.upper() else name

    # Only use live photo from the device — never fall back to database photos
    image_b64 = photo_b64
    photo_source = "live" if image_b64 else "none"

    logger.info(
        "[TRUEFACE] Chairman notify → %s: %s at %s (photo=%s)",
        CHAIRMAN_PHONE, display_name, time_str, photo_source,
    )

    if not image_b64:
        logger.warning("[TRUEFACE] No photo available for %s — skipping chairman notify", name)
        return False

    header_image_id = None
    try:
        header_image_id = await upload_base64_image_cloud(image_b64)
        if header_image_id:
            logger.info("[TRUEFACE] Uploaded %s photo, media_id=%s", photo_source, header_image_id)
        else:
            logger.warning("[TRUEFACE] Photo upload failed for %s", name)
            return False
    except Exception as e:
        logger.warning("[TRUEFACE] Photo upload error: %s", e)
        return False

    try:
        ok = await send_cloud_template_message(
            to=CHAIRMAN_PHONE,
            template_name=CHAIRMAN_TEMPLATE,
            language_code="en",
            body_params=[display_name, time_str],
            header_image_id=header_image_id,
        )
        logger.info(
            "[TRUEFACE] Chairman WhatsApp %s for %s",
            "OK" if ok else "FAILED", display_name,
        )
        return bool(ok)
    except Exception as e:
        logger.error("[TRUEFACE] Chairman WhatsApp error: %s", e)
        return False


# ============================================================
# Gate entry + teacher sighting from TrueFace events
# ============================================================

async def _log_trueface_to_gate_and_sighting(
    name: str, pin: str, timestamp: str, date: str,
):
    """Log a TrueFace detection as both a gate entry and a teacher sighting
    so the head count reconciliation report includes TrueFace data."""
    ts = timestamp or datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    db = await _get_db()
    try:
        await db.execute(
            "INSERT INTO gate_entries "
            "(date, timestamp, camera, direction, attire_color, person_crop) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (date, ts, "TrueFace 3000", "IN", "", ""),
        )
        person_id = f"TEACHER_{name.upper().replace(' ', '_')}"
        await db.execute(
            "INSERT INTO teacher_dvr_sightings "
            "(date, timestamp, person_id, name, camera, confidence, "
            "outfit_color, outfit_description, outfit_colors_json) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (date, ts, person_id, name, "TrueFace 3000", 1.0,
             "", "identified by TrueFace biometric", "[]"),
        )
        await db.commit()
        logger.info("[TRUEFACE] Logged gate entry + teacher sighting for %s", name)
    except Exception as e:
        logger.warning("[TRUEFACE] Failed to log gate/sighting for %s: %s", name, e)
    finally:
        await db.close()


# ============================================================
# Mood logging from TrueFace photos
# ============================================================

async def _analyze_face_mood(photo_b64: str) -> dict:
    """Use OpenAI Vision to analyze facial expression and return mood data."""
    if not photo_b64:
        return {
            "dominant_emotion": "neutral",
            "emotions": {"neutral": 1.0},
            "temperament": "calm",
            "intensity": 30.0,
        }
    try:
        from app.services.openai_service import get_client
        ai_client = get_client()
        if ai_client is None:
            return {
                "dominant_emotion": "neutral",
                "emotions": {"neutral": 1.0},
                "temperament": "calm",
                "intensity": 30.0,
            }

        data_uri = f"data:image/jpeg;base64,{photo_b64}"
        response = await ai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": (
                        "Analyze this person's facial expression. Respond ONLY with a JSON object, no markdown:\n"
                        '{"dominant_emotion": "<one of: happy, neutral, sad, angry, surprised, fearful, disgusted, contemplative, tired, confident, anxious>",\n'
                        ' "emotions": {"<emotion>": <0.0-1.0>, ...},\n'
                        ' "temperament": "<one of: cheerful, calm, composed, tense, irritable, melancholic, energetic, withdrawn, focused, relaxed, stressed, warm>",\n'
                        ' "intensity": <1-100 how strong the expression is>,\n'
                        ' "description": "<one sentence describing their apparent mood and demeanor>"}'
                    )},
                    {"type": "image_url", "image_url": {"url": data_uri, "detail": "low"}},
                ],
            }],
            max_tokens=200,
            temperature=0.3,
        )
        text = response.choices[0].message.content.strip()
        # Strip markdown fences if present
        if text.startswith("```"):
            text = text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        result = json.loads(text)
        # Validate required fields
        if "dominant_emotion" not in result or "temperament" not in result:
            raise ValueError("Missing required fields")
        return result
    except Exception as e:
        logger.warning("[TRUEFACE] Mood analysis failed: %s", e)
        return {
            "dominant_emotion": "neutral",
            "emotions": {"neutral": 1.0},
            "temperament": "calm",
            "intensity": 30.0,
        }


_mood_dedup_last: dict[str, str] = {}  # person_label -> last logged timestamp
_MOOD_DEDUP_WINDOW = 300  # 5 minutes — skip duplicate TrueFace mood entries


async def _log_mood_from_trueface(name: str, timestamp: str, photo_b64: str):
    """If the person is Chairman or Alisha, log a mood observation from their
    TrueFace recognition photo. Uses the same chairman_mood_log table.

    Deduplicates: only one mood entry per person per 5-minute window.
    """
    person_label = MOOD_TRACKED_NAMES.get(name.upper().strip())
    if not person_label:
        return

    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    ts = timestamp or now.strftime("%Y-%m-%d %H:%M:%S")

    # Deduplicate: skip if same person was logged within the window
    last_ts = _mood_dedup_last.get(person_label, "")
    if last_ts:
        try:
            t_prev = datetime.strptime(last_ts, "%Y-%m-%d %H:%M:%S")
            t_curr = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            if abs((t_curr - t_prev).total_seconds()) < _MOOD_DEDUP_WINDOW:
                logger.debug("[TRUEFACE] Mood dedup skip for %s (last=%s, now=%s)",
                             person_label, last_ts, ts)
                return
        except (ValueError, TypeError):
            pass

    _mood_dedup_last[person_label] = ts

    mood = await _analyze_face_mood(photo_b64)
    dominant_emotion = mood.get("dominant_emotion", "neutral")
    emotions = mood.get("emotions", {"neutral": 1.0})
    temperament = mood.get("temperament", "calm")
    intensity = float(mood.get("intensity", 30.0))
    description = mood.get("description", "")

    db = await _get_db()
    try:
        await db.execute(
            "INSERT INTO chairman_mood_log "
            "(person, date, timestamp, camera, dominant_emotion, emotions_json, "
            "temperament, intensity, face_distance, face_confidence, face_crop, "
            "description) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                person_label, today, ts,
                "TrueFace 3000", dominant_emotion, json.dumps(emotions),
                temperament, intensity, 0.0, 1.0, photo_b64[:200] if photo_b64 else "",
                description,
            ),
        )
        await db.commit()
    finally:
        await db.close()

    logger.info("[TRUEFACE] Mood logged for %s (%s) at %s: %s / %s / %.0f%% — %s",
                name, person_label, ts, dominant_emotion, temperament, intensity,
                description[:80] if description else "no description")


# ============================================================
# Core attendance event endpoint
# ============================================================

@router.post("/api/trueface/event")
async def receive_trueface_event(request: Request):
    """Receive a face-recognition event from the browser poller.

    Body: {"pin": "1", "name": "alisha ahuja", "timestamp": "2026-05-22 07:30:00"}
    Or batch: [{"pin": "1", ...}, {"pin": "2", ...}]

    Returns which events were processed as arrival/departure.
    """
    body = await request.json()
    events = body if isinstance(body, list) else [body]

    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    db = await _get_db()
    results = []
    try:
        for evt in events:
            pin = str(evt.get("pin", "")).strip()
            timestamp = evt.get("timestamp", "").strip()
            evt_name = evt.get("name", "").strip()

            if not pin:
                results.append({"pin": pin, "status": "skipped", "reason": "no pin"})
                continue

            # Use the event's own date (from the device timestamp), not today.
            # This prevents stale events from yesterday being recorded under today.
            evt_date = today
            if timestamp and " " in timestamp:
                evt_date_part = timestamp.split(" ")[0]
                if len(evt_date_part) == 10:  # YYYY-MM-DD format
                    evt_date = evt_date_part

            if evt_date != today:
                logger.info(
                    f"[TRUEFACE] Skipping stale event: PIN={pin} name={evt_name} "
                    f"event_date={evt_date} today={today}"
                )
                results.append({
                    "pin": pin, "status": "skipped",
                    "reason": f"stale event from {evt_date}",
                })
                continue

            teacher = await _get_teacher(db, pin)
            if not teacher and evt_name:
                # Auto-register: match name against DVR teacher database
                teacher = await _auto_register_from_dvr(db, pin, evt_name)
            if not teacher:
                logger.info(f"[TRUEFACE] Unknown PIN={pin} name={evt_name}")
                results.append({"pin": pin, "status": "skipped", "reason": "unknown"})
                continue

            name = teacher["name"]
            phone = teacher["phone"]
            time_str = _format_time_12h(timestamp) if timestamp else now.strftime("%I:%M %p")
            time_raw = timestamp.split(" ")[1] if " " in timestamp else now.strftime("%H:%M:%S")

            record = await _get_attendance_record(db, pin, today)

            photo_b64 = evt.get("photo", "")

            # Log mood observation for Chairman/Alisha from TrueFace photo
            asyncio.ensure_future(
                _log_mood_from_trueface(name, timestamp, photo_b64)
            )

            # Also log as gate entry + teacher sighting for reconciliation
            asyncio.ensure_future(
                _log_trueface_to_gate_and_sighting(name, pin, timestamp, today)
            )

            # Determine current IST hour from the event timestamp
            try:
                evt_hour = int(time_raw.split(":")[0])
            except (ValueError, IndexError):
                evt_hour = now.hour

            if not record:
                # First detection → arrival
                await db.execute(
                    "INSERT INTO trueface_attendance "
                    "(pin, name, date, arrival_time, arrival_whatsapp) "
                    "VALUES (?, ?, ?, ?, 0)",
                    (pin, name, today, time_raw),
                )
                await db.commit()

                wa_ok = False
                if phone:
                    wa_ok = await _send_arrival_whatsapp(name, phone, time_str)
                    if wa_ok:
                        await db.execute(
                            "UPDATE trueface_attendance SET arrival_whatsapp = 1 "
                            "WHERE pin = ? AND date = ?",
                            (pin, today),
                        )
                        await db.commit()

                # Notify chairman of each arrival
                asyncio.ensure_future(
                    _notify_chairman_arrival(name, time_str, photo_b64)
                )

                logger.info(f"[TRUEFACE] ARRIVAL: {name} at {time_str} WA={wa_ok}")
                results.append({
                    "pin": pin, "name": name, "status": "arrival",
                    "time": time_str, "whatsapp": wa_ok,
                })

            elif not record["departure_time"] and evt_hour >= DEPARTURE_HOUR:
                # After DEPARTURE_HOUR and no departure yet → mark departure
                await db.execute(
                    "UPDATE trueface_attendance SET departure_time = ?, departure_whatsapp = 0 "
                    "WHERE pin = ? AND date = ?",
                    (time_raw, pin, today),
                )
                await db.commit()

                wa_ok = False
                if phone:
                    wa_ok = await _send_departure_whatsapp(name, phone, time_str)
                    if wa_ok:
                        await db.execute(
                            "UPDATE trueface_attendance SET departure_whatsapp = 1 "
                            "WHERE pin = ? AND date = ?",
                            (pin, today),
                        )
                        await db.commit()

                logger.info(f"[TRUEFACE] DEPARTURE: {name} at {time_str} WA={wa_ok}")
                results.append({
                    "pin": pin, "name": name, "status": "departure",
                    "time": time_str, "whatsapp": wa_ok,
                })

            elif not record["departure_time"] and evt_hour < DEPARTURE_HOUR:
                # Before DEPARTURE_HOUR — ignore extra scans (arrival already recorded)
                logger.info(
                    f"[TRUEFACE] Ignored pre-{DEPARTURE_HOUR}:00 scan for {name} "
                    f"at {time_str} (arrival already at {_format_time_12h(record['arrival_time'])})"
                )
                results.append({
                    "pin": pin, "name": name, "status": "ignored_morning",
                    "time": time_str,
                })

            else:
                # Already have arrival + departure → update departure time silently
                await db.execute(
                    "UPDATE trueface_attendance SET departure_time = ? "
                    "WHERE pin = ? AND date = ?",
                    (time_raw, pin, today),
                )
                await db.commit()
                results.append({
                    "pin": pin, "name": name, "status": "updated_departure",
                    "time": time_str,
                })

    finally:
        await db.close()

    return {"status": "ok", "results": results}



# ============================================================
# Excel report generation + email
# ============================================================

def _generate_attendance_excel(
    teachers: list[dict],
    attendance: list[dict],
    report_date: str,
    report_type: str = "arrival",
) -> bytes:
    """Generate an Excel report for teacher attendance."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active

    title_label = "Arrival" if report_type == "arrival" else "Departure"
    ws.title = f"Teacher {title_label}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    present_fill = PatternFill("solid", fgColor="C6EFCE")
    absent_fill = PatternFill("solid", fgColor="FFC7CE")
    left_fill = PatternFill("solid", fgColor="BDD7EE")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = (
        f"PP International School — Teacher {title_label} Report — {report_date}"
    )
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    if report_type == "arrival":
        headers = ["S.No", "Teacher Name", "Status", "Arrival Time", "WhatsApp Sent", "Remarks"]
    else:
        headers = ["S.No", "Teacher Name", "Arrival Time", "Departure Time", "WhatsApp Sent", "Remarks"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    att_map = {a["pin"]: a for a in attendance}

    row = 4
    present_count = 0
    for i, t in enumerate(teachers, 1):
        pin = t["pin"]
        att = att_map.get(pin)

        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=t["name"]).border = border

        if report_type == "arrival":
            if att and att.get("arrival_time"):
                present_count += 1
                ws.cell(row=row, column=3, value="Present").border = border
                ws.cell(row=row, column=3).fill = present_fill
                ws.cell(row=row, column=4, value=_format_time_12h(att["arrival_time"])).border = border
                ws.cell(row=row, column=5, value="Yes" if att.get("arrival_whatsapp") else "No").border = border
            else:
                ws.cell(row=row, column=3, value="Absent").border = border
                ws.cell(row=row, column=3).fill = absent_fill
                ws.cell(row=row, column=4, value="-").border = border
                ws.cell(row=row, column=5, value="-").border = border
            ws.cell(row=row, column=6, value="").border = border
        else:
            arrival_time = _format_time_12h(att["arrival_time"]) if att and att.get("arrival_time") else "-"
            departure_time = _format_time_12h(att["departure_time"]) if att and att.get("departure_time") else "-"
            ws.cell(row=row, column=3, value=arrival_time).border = border
            if att and att.get("departure_time"):
                ws.cell(row=row, column=4, value=departure_time).border = border
                ws.cell(row=row, column=4).fill = left_fill
                ws.cell(row=row, column=5, value="Yes" if att.get("departure_whatsapp") else "No").border = border
                present_count += 1
            elif att and att.get("arrival_time"):
                ws.cell(row=row, column=4, value="Still Present").border = border
                ws.cell(row=row, column=4).fill = present_fill
                ws.cell(row=row, column=5, value="-").border = border
                present_count += 1
            else:
                ws.cell(row=row, column=4, value="Absent").border = border
                ws.cell(row=row, column=4).fill = absent_fill
                ws.cell(row=row, column=5, value="-").border = border
            ws.cell(row=row, column=6, value="").border = border

        row += 1

    # Summary row
    row += 1
    total = len(teachers)
    absent = total - present_count
    ws.cell(row=row, column=2, value=f"Total: {total} | Present: {present_count} | Absent: {absent}")
    ws.cell(row=row, column=2).font = Font(bold=True)

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20
    ws.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def send_attendance_report(report_type: str = "arrival"):
    """Generate and email attendance report."""
    from app.services.email_service import send_email_async

    today = datetime.now(IST).strftime("%Y-%m-%d")
    today_display = datetime.now(IST).strftime("%d %b %Y (%A)")

    now = datetime.now(IST)

    db = await _get_db()
    try:
        teachers = await _get_all_teachers(db)
        attendance = await _get_all_attendance(db, today)
    finally:
        await db.close()

    if not teachers:
        logger.warning("[TRUEFACE] No teachers registered — skipping report")
        return

    present = len([a for a in attendance if a.get("arrival_time")])
    total = len(teachers)

    title_label = "Arrival" if report_type == "arrival" else "Departure"
    xlsx_bytes = _generate_attendance_excel(teachers, attendance, today, report_type)
    filename = f"Teacher_{title_label}_{today}.xlsx"

    body = (
        f"Daily Teacher {title_label} Report — {today_display}\n\n"
        f"Present: {present} / {total}\n"
        f"Absent: {total - present}\n\n"
        f"Please find the detailed report attached.\n\n"
        f"— PPIS TrueFace Attendance System"
    )

    recipients = [r.strip() for r in REPORT_RECIPIENTS.split(",") if r.strip()]
    for email in recipients:
        ok = await send_email_async(
            email,
            f"Teacher {title_label} Report — {today_display}",
            body,
            "PP International School",
            attachments=[(filename, xlsx_bytes)],
        )
        logger.info(f"[TRUEFACE] {title_label} report → {email}: {'OK' if ok else 'FAILED'}")


def send_arrival_report_sync():
    """Sync wrapper for scheduler."""
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            asyncio.ensure_future(send_attendance_report("arrival"))
        else:
            loop.run_until_complete(send_attendance_report("arrival"))
    except RuntimeError:
        asyncio.run(send_attendance_report("arrival"))


def send_departure_report_sync():
    """Sync wrapper for scheduler."""
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            asyncio.ensure_future(send_attendance_report("departure"))
        else:
            loop.run_until_complete(send_attendance_report("departure"))
    except RuntimeError:
        asyncio.run(send_attendance_report("departure"))


# ============================================================
# ADMS Protocol Endpoints — TrueFace device pushes events here
# ============================================================
#
# ZKTeco ADMS (Automatic Data Master Server) protocol:
#   1. GET  /iclock/cdata?SN=xxx — device check-in / handshake
#   2. POST /iclock/cdata?SN=xxx&table=ATTLOG — attendance records push
#   3. GET  /iclock/getrequest?SN=xxx — device asks for pending commands
#   4. POST /iclock/devicecmd?SN=xxx — device reports command results
#
# The device pushes attendance records as tab-separated lines:
#   PIN\tTimestamp\tStatus\tVerify\tWorkCode\tReserved1\tReserved2
#   Example: 1395\t2026-05-25 08:30:00\t0\t15\t0\t1\t0

ADMS_SERIAL = os.environ.get("TRUEFACE_SERIAL", "TW30000001260433")


@router.get("/iclock/cdata")
async def iclock_cdata_get(request: Request):
    """ADMS handshake — device checks in and asks for options.

    The response tells the device what data tables to push and how often.
    """
    sn = request.query_params.get("SN", "unknown")
    logger.info(f"[ADMS] Handshake from SN={sn}")

    # Respond with standard ADMS options
    # TransFlag=TransData  → device should push attendance data
    # Realtime=1           → push events in near-realtime
    # TransInterval=1      → push every 1 minute
    options = (
        f"GET OPTION FROM: {sn}\r\n"
        f"Stamp=0\r\n"
        f"OpStamp=0\r\n"
        f"PhotoStamp=0\r\n"
        f"TransFlag=TransData AttLog\tOpLog\r\n"
        f"Realtime=1\r\n"
        f"TransInterval=1\r\n"
        f"TransTimes=00:00;23:59\r\n"
        f"Encrypt=0\r\n"
    )
    return PlainTextResponse(options)


@router.post("/iclock/cdata")
async def iclock_cdata_post(request: Request):
    """ADMS data push — device sends attendance or operation logs.

    For ATTLOG table, each line is a tab-separated attendance record.
    We parse these and feed them into the same event pipeline as the
    browser-based poller (receive_trueface_event).
    """
    sn = request.query_params.get("SN", "unknown")
    table = request.query_params.get("table", "").upper()
    body = await request.body()
    body_text = body.decode("utf-8", errors="replace").strip()

    logger.info(f"[ADMS] POST from SN={sn} table={table} ({len(body_text)} bytes)")

    if table == "ATTLOG" and body_text:
        events = _parse_adms_attlog(body_text)
        if events:
            logger.info(f"[ADMS] Parsed {len(events)} attendance event(s)")
            results = await _process_adms_events(events)
            logger.info(f"[ADMS] Processed: {results}")
    elif table == "OPERLOG":
        logger.info(f"[ADMS] Operation log: {body_text[:200]}")
    else:
        logger.info(f"[ADMS] Other data (table={table}): {body_text[:200]}")

    return PlainTextResponse("OK")


@router.get("/iclock/getrequest")
async def iclock_getrequest(request: Request):
    """Device asks for pending commands. Return OK (no pending commands)."""
    return PlainTextResponse("OK")


@router.post("/iclock/devicecmd")
async def iclock_devicecmd(request: Request):
    """Device reports command execution results."""
    body = await request.body()
    logger.info(f"[ADMS] Device command result: {body.decode('utf-8', errors='replace')[:200]}")
    return PlainTextResponse("OK")


def _parse_adms_attlog(body: str) -> list[dict]:
    """Parse ADMS ATTLOG push data into event dicts.

    Format (tab-separated, one record per line):
        PIN\tTimestamp\tStatus\tVerify\tWorkCode[\tReserved1\tReserved2]

    Status: 0=Check-In, 1=Check-Out, 2=Break-Out, 3=Break-In, 4=OT-In, 5=OT-Out
    Verify: 1=FP, 4=Card, 15=Face
    """
    events = []
    for line in body.split("\n"):
        line = line.strip()
        if not line:
            continue

        parts = line.split("\t")
        if len(parts) < 2:
            logger.warning(f"[ADMS] Skipping malformed line: {line[:100]}")
            continue

        pin = parts[0].strip()
        timestamp = parts[1].strip() if len(parts) > 1 else ""
        status = parts[2].strip() if len(parts) > 2 else "0"
        verify = parts[3].strip() if len(parts) > 3 else ""

        if not pin or not timestamp:
            continue

        # Map verify code to method name
        verify_method = {
            "1": "Fingerprint", "4": "Card", "15": "Face",
        }.get(verify, f"Method-{verify}")

        events.append({
            "pin": pin,
            "name": "",  # ADMS doesn't always send name; lookup happens in pipeline
            "timestamp": timestamp,
            "adms_status": status,
            "adms_verify": verify_method,
        })

    return events


async def _process_adms_events(events: list[dict]) -> list[dict]:
    """Process ADMS events through the same pipeline as browser poller events.

    Reuses the exact same arrival/departure logic, WhatsApp notifications,
    and chairman alerts as receive_trueface_event().
    """
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    db = await _get_db()
    results = []
    try:
        for evt in events:
            pin = str(evt.get("pin", "")).strip()
            timestamp = evt.get("timestamp", "").strip()
            evt_name = evt.get("name", "").strip()

            if not pin:
                results.append({"pin": pin, "status": "skipped", "reason": "no pin"})
                continue

            # Extract event date — skip stale events
            evt_date = today
            if timestamp and " " in timestamp:
                evt_date_part = timestamp.split(" ")[0]
                if len(evt_date_part) == 10:
                    evt_date = evt_date_part

            if evt_date != today:
                logger.info(
                    f"[ADMS] Skipping stale event: PIN={pin} "
                    f"event_date={evt_date} today={today}"
                )
                results.append({
                    "pin": pin, "status": "skipped",
                    "reason": f"stale event from {evt_date}",
                })
                continue

            # Look up teacher by PIN
            teacher = await _get_teacher(db, pin)
            if not teacher and evt_name:
                teacher = await _auto_register_from_dvr(db, pin, evt_name)
            if not teacher:
                # Try contact sheet lookup by PIN alone
                teacher = await _auto_register_from_contacts(db, pin, evt_name or f"PIN-{pin}")
            if not teacher:
                logger.info(f"[ADMS] Unknown PIN={pin}")
                results.append({"pin": pin, "status": "skipped", "reason": "unknown"})
                continue

            name = teacher["name"]
            phone = teacher["phone"]
            time_str = _format_time_12h(timestamp) if timestamp else now.strftime("%I:%M %p")
            time_raw = timestamp.split(" ")[1] if " " in timestamp else now.strftime("%H:%M:%S")

            record = await _get_attendance_record(db, pin, today)

            # Log mood observation for Chairman/Alisha from ADMS event
            asyncio.ensure_future(
                _log_mood_from_trueface(name, timestamp, "")
            )

            try:
                evt_hour = int(time_raw.split(":")[0])
            except (ValueError, IndexError):
                evt_hour = now.hour

            if not record:
                # First detection → arrival
                await db.execute(
                    "INSERT INTO trueface_attendance "
                    "(pin, name, date, arrival_time, arrival_whatsapp) "
                    "VALUES (?, ?, ?, ?, 0)",
                    (pin, name, today, time_raw),
                )
                await db.commit()

                wa_ok = False
                if phone:
                    wa_ok = await _send_arrival_whatsapp(name, phone, time_str)
                    if wa_ok:
                        await db.execute(
                            "UPDATE trueface_attendance SET arrival_whatsapp = 1 "
                            "WHERE pin = ? AND date = ?",
                            (pin, today),
                        )
                        await db.commit()

                asyncio.ensure_future(
                    _notify_chairman_arrival(name, time_str, "")
                )

                logger.info(f"[ADMS] ARRIVAL: {name} at {time_str} WA={wa_ok}")
                results.append({
                    "pin": pin, "name": name, "status": "arrival",
                    "time": time_str, "whatsapp": wa_ok,
                })

            elif not record["departure_time"] and evt_hour >= DEPARTURE_HOUR:
                await db.execute(
                    "UPDATE trueface_attendance SET departure_time = ?, departure_whatsapp = 0 "
                    "WHERE pin = ? AND date = ?",
                    (time_raw, pin, today),
                )
                await db.commit()

                wa_ok = False
                if phone:
                    wa_ok = await _send_departure_whatsapp(name, phone, time_str)
                    if wa_ok:
                        await db.execute(
                            "UPDATE trueface_attendance SET departure_whatsapp = 1 "
                            "WHERE pin = ? AND date = ?",
                            (pin, today),
                        )
                        await db.commit()

                logger.info(f"[ADMS] DEPARTURE: {name} at {time_str} WA={wa_ok}")
                results.append({
                    "pin": pin, "name": name, "status": "departure",
                    "time": time_str, "whatsapp": wa_ok,
                })

            elif not record["departure_time"] and evt_hour < DEPARTURE_HOUR:
                results.append({
                    "pin": pin, "name": name, "status": "ignored_morning",
                    "time": time_str,
                })

            else:
                await db.execute(
                    "UPDATE trueface_attendance SET departure_time = ? "
                    "WHERE pin = ? AND date = ?",
                    (time_raw, pin, today),
                )
                await db.commit()
                results.append({
                    "pin": pin, "name": name, "status": "updated_departure",
                    "time": time_str,
                })

    finally:
        await db.close()

    return results


# ============================================================
# Teacher Management API
# ============================================================

@router.get("/api/trueface/teachers")
async def list_teachers():
    """List all registered TrueFace teachers."""
    db = await _get_db()
    try:
        teachers = await _get_all_teachers(db)
    finally:
        await db.close()
    return {"teachers": teachers, "count": len(teachers)}


@router.post("/api/trueface/teachers")
async def register_teacher(request: Request):
    """Register a single teacher. Body: {"pin": "1", "name": "...", "phone": "..."}"""
    data = await request.json()
    pin = str(data.get("pin", "")).strip()
    name = data.get("name", "").strip()
    phone = data.get("phone", "").strip()

    if not pin or not name:
        return {"error": "pin and name are required"}

    db = await _get_db()
    try:
        await db.execute(
            "INSERT OR REPLACE INTO trueface_teachers (pin, name, phone) VALUES (?, ?, ?)",
            (pin, name, phone),
        )
        await db.commit()
    finally:
        await db.close()

    logger.info(f"[TRUEFACE] Teacher registered: PIN={pin} name={name} phone={phone}")
    return {"status": "ok", "pin": pin, "name": name}


@router.post("/api/trueface/teachers/bulk")
async def bulk_register_teachers(request: Request):
    """Bulk register teachers. Body: [{"pin": "1", "name": "...", "phone": "..."}, ...]"""
    data = await request.json()
    db = await _get_db()
    count = 0
    try:
        for entry in data:
            pin = str(entry.get("pin", "")).strip()
            name = entry.get("name", "").strip()
            phone = entry.get("phone", "").strip()
            if pin and name:
                await db.execute(
                    "INSERT OR REPLACE INTO trueface_teachers (pin, name, phone) VALUES (?, ?, ?)",
                    (pin, name, phone),
                )
                count += 1
        await db.commit()
    finally:
        await db.close()

    logger.info(f"[TRUEFACE] Bulk registered {count} teachers")
    return {"status": "ok", "registered": count}


@router.post("/api/trueface/contacts/bulk")
async def bulk_upload_contacts(request: Request):
    """Upload contact sheet data. Body: [{"pin": "...", "name": "...", "phone": "...", "category": "staff"}, ...]"""
    data = await request.json()
    db = await _get_db()
    count = 0
    try:
        # Clear existing contacts and re-insert
        await db.execute("DELETE FROM trueface_contacts")
        for entry in data:
            pin = (entry.get("pin") or "").strip()
            name = (entry.get("name") or "").strip()
            phone = (entry.get("phone") or "").strip()
            category = (entry.get("category") or "staff").strip()
            if name and phone:
                await db.execute(
                    "INSERT INTO trueface_contacts (pin, name, phone, category) VALUES (?, ?, ?, ?)",
                    (pin, name, phone, category),
                )
                count += 1
        await db.commit()
    finally:
        await db.close()

    logger.info(f"[TRUEFACE] Uploaded {count} contacts from sheet")
    return {"status": "ok", "uploaded": count}


@router.get("/api/trueface/contacts")
async def list_contacts():
    """List all uploaded contacts."""
    db = await _get_db()
    try:
        cur = await db.execute("SELECT pin, name, phone, category FROM trueface_contacts ORDER BY name")
        rows = await cur.fetchall()
    finally:
        await db.close()
    return {"contacts": [{"pin": r[0], "name": r[1], "phone": r[2], "category": r[3]} for r in rows], "count": len(rows)}


@router.get("/api/trueface/attendance")
async def get_attendance(date: str | None = None):
    """Get attendance for a given date (default: today)."""
    if not date:
        date = datetime.now(IST).strftime("%Y-%m-%d")
    db = await _get_db()
    try:
        teachers = await _get_all_teachers(db)
        attendance = await _get_all_attendance(db, date)
    finally:
        await db.close()

    att_map = {a["pin"]: a for a in attendance}
    result = []
    for t in teachers:
        att = att_map.get(t["pin"], {})
        result.append({
            "pin": t["pin"],
            "name": t["name"],
            "phone": t["phone"],
            "arrival_time": _format_time_12h(att["arrival_time"]) if att.get("arrival_time") else None,
            "departure_time": _format_time_12h(att["departure_time"]) if att.get("departure_time") else None,
            "status": "departed" if att.get("departure_time") else ("present" if att.get("arrival_time") else "absent"),
        })

    present = sum(1 for r in result if r["status"] != "absent")
    return {"date": date, "total": len(teachers), "present": present, "records": result}


@router.post("/api/trueface/report/arrival")
async def trigger_arrival_report():
    """Manually trigger the arrival report."""
    await send_attendance_report("arrival")
    return {"status": "ok", "type": "arrival"}


@router.post("/api/trueface/report/departure")
async def trigger_departure_report():
    """Manually trigger the departure report."""
    await send_attendance_report("departure")
    return {"status": "ok", "type": "departure"}


@router.get("/api/trueface/status")
async def trueface_status():
    """Get TrueFace integration status."""
    db = await _get_db()
    try:
        teachers = await _get_all_teachers(db)
        today = datetime.now(IST).strftime("%Y-%m-%d")
        attendance = await _get_all_attendance(db, today)
    finally:
        await db.close()

    return {
        "registered_teachers": len(teachers),
        "today_present": len([a for a in attendance if a.get("arrival_time")]),
        "today_departed": len([a for a in attendance if a.get("departure_time")]),
    }
