"""
Gate Head Count Reconciliation & Entry Monitoring System
========================================================
AI-powered school headcount reconciliation comparing detections across:
  - Entry Gate DVR Cameras
  - TrueFace 3000 Facial Recognition System
  - Reception Camera
  - Additional Internal Cameras

Classifies: Teachers, Visitors, Unknown Persons
Reconciliation categories:
  ✓ FULLY VERIFIED — detected on both TrueFace + DVR cameras
  ⚠ ENTRY ONLY (DVR Only) — seen on DVR but not marked on TrueFace
  ⚠ TRUEFACE ONLY — recognized by TrueFace but no DVR sighting
  — ABSENT — not detected on any system

Endpoints:
    POST /api/gate/entry             — receive gate entry events
    POST /api/gate/teacher-sighting  — receive DVR teacher face sightings
    POST /api/gate/visitor-sighting  — receive DVR visitor (unknown) sightings
    GET  /api/gate/status            — today's running totals
    GET  /api/gate/reconciliation/{date} — full reconciliation data

Scheduled:
    Every 30 min, 6 AM – 5 PM IST — reconciliation report emailed

All timestamps use IST (Asia/Kolkata, UTC+05:30).
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import os
import re
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger("app.gate")

IST = timezone(timedelta(hours=5, minutes=30))
router = APIRouter()

REPORT_RECIPIENTS = os.environ.get(
    "GATE_REPORT_EMAIL",
    "alisha.kanwar@ppischool.in",
)

CHAIRMAN_PHONE = os.environ.get("TRUEFACE_CHAIRMAN_PHONE", "919971166562")
UNKNOWN_ALERT_PHONE = os.environ.get("UNKNOWN_ALERT_PHONE", "918796105084")
UNKNOWN_ALERT_PHONES = [
    # "918796105084",   # Alisha — disabled per request
]
UNKNOWN_ALERT_TEMPLATE = "unknown_person_detected"

# Dedup: max one unknown person alert per camera per 60-second window
_unknown_alert_last: dict[str, float] = {}
_UNKNOWN_ALERT_COOLDOWN = 60  # seconds

# Cameras at actual entry/exit points (Main Gate, Basement, Dispersal)
_ENTRY_CAMERA_KEYWORDS = ("ENTRY GATE", "DISPERSAL", "BASEMENT")

# Deduplication: entries within this window from same/nearby cameras = same person
_DEDUP_WINDOW_SECONDS = 120  # 2 minutes


def _deduplicate_gate_entries(entries: list[dict]) -> list[dict]:
    """Collapse raw gate detections into estimated unique crossings.

    Same-camera re-detections within the time window are dropped.
    Cross-camera detections with matching attire within the window
    are treated as the same person.
    """
    if not entries:
        return []

    sorted_entries = sorted(entries, key=lambda e: e.get("timestamp", ""))
    unique: list[dict] = []

    for entry in sorted_entries:
        ts_str = entry.get("timestamp", "")
        attire = entry.get("attire_color", "unknown")
        direction = entry.get("direction", "IN")

        is_dup = False
        for prev in reversed(unique):
            prev_ts = prev.get("timestamp", "")
            try:
                t1 = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                t2 = datetime.strptime(prev_ts, "%Y-%m-%d %H:%M:%S")
                diff = abs((t1 - t2).total_seconds())
            except (ValueError, TypeError):
                break
            if diff > _DEDUP_WINDOW_SECONDS:
                break
            if prev.get("direction") != direction:
                continue
            # Same camera within window → definitely same person
            if prev.get("camera") == entry.get("camera"):
                is_dup = True
                break
            # Different camera but same attire within window → likely same person
            if attire and attire != "unknown" and prev.get("attire_color", "unknown") == attire:
                is_dup = True
                break

        if not is_dup:
            unique.append(entry)

    return unique


# ============================================================
# Database helpers
# ============================================================

async def _get_db():
    from app.database import get_db
    return await get_db()


async def _store_gate_entries(db, entries: list[dict]) -> int:
    """Store gate entry events in the database. Returns count stored."""
    count = 0
    for entry in entries:
        ts = entry.get("timestamp", "")
        date_part = ts.split(" ")[0] if " " in ts else datetime.now(IST).strftime("%Y-%m-%d")
        camera = entry.get("camera", "")
        direction = entry.get("direction", "IN")
        attire_color = entry.get("attire_color", "unknown")
        person_crop = entry.get("person_crop", "")

        await db.execute(
            "INSERT INTO gate_entries (date, timestamp, camera, direction, attire_color, person_crop) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (date_part, ts, camera, direction, attire_color, person_crop),
        )
        count += 1
    await db.commit()
    return count


async def _get_gate_entries(db, date: str, direction: str | None = None) -> list[dict]:
    """Get all gate entries for a date, optionally filtered by direction."""
    if direction:
        cur = await db.execute(
            "SELECT id, date, timestamp, camera, direction, attire_color, reconciled, matched_pin, notes, person_crop "
            "FROM gate_entries WHERE date = ? AND direction = ? ORDER BY timestamp",
            (date, direction),
        )
    else:
        cur = await db.execute(
            "SELECT id, date, timestamp, camera, direction, attire_color, reconciled, matched_pin, notes, person_crop "
            "FROM gate_entries WHERE date = ? ORDER BY timestamp",
            (date,),
        )
    rows = await cur.fetchall()
    return [
        {
            "id": r[0], "date": r[1], "timestamp": r[2], "camera": r[3],
            "direction": r[4], "attire_color": r[5], "reconciled": bool(r[6]),
            "matched_pin": r[7], "notes": r[8],
            "person_crop": r[9] if len(r) > 9 else "",
        }
        for r in rows
    ]


async def _get_gate_entries_with_crops(db, date: str) -> list[dict]:
    """Get gate entries including person_crop images for snapshot gallery."""
    cur = await db.execute(
        "SELECT id, timestamp, camera, direction, person_crop "
        "FROM gate_entries WHERE date = ? AND direction = 'IN' "
        "AND person_crop != '' ORDER BY timestamp DESC LIMIT 16",
        (date,),
    )
    return [
        {"id": r[0], "timestamp": r[1], "camera": r[2],
         "direction": r[3], "person_crop": r[4]}
        for r in await cur.fetchall()
    ]


async def _get_trueface_attendance(db, date: str) -> list[dict]:
    """Get all TrueFace attendance records for a date."""
    cur = await db.execute(
        "SELECT pin, name, arrival_time, departure_time "
        "FROM trueface_attendance WHERE date = ? ORDER BY arrival_time",
        (date,),
    )
    return [
        {"pin": r[0], "name": r[1], "arrival_time": r[2], "departure_time": r[3]}
        for r in await cur.fetchall()
    ]


async def _get_total_teachers(db) -> int:
    """Get total registered teachers count (all teachers, not just those with phones)."""
    cur = await db.execute("SELECT COUNT(*) FROM trueface_teachers")
    row = await cur.fetchone()
    return row[0] if row else 0


async def _get_all_teachers(db) -> list[dict]:
    """Get all registered teachers (including those without phone numbers)."""
    cur = await db.execute(
        "SELECT pin, name, phone FROM trueface_teachers ORDER BY name"
    )
    return [{"pin": r[0], "name": r[1], "phone": r[2] or ""} for r in await cur.fetchall()]


async def _get_contact_categories(db) -> dict[str, str]:
    """Get category for each contact by PIN from trueface_contacts.
    Returns dict mapping UPPERCASED name -> category."""
    cur = await db.execute(
        "SELECT name, category FROM trueface_contacts ORDER BY name"
    )
    result: dict[str, str] = {}
    for r in await cur.fetchall():
        name = (r[0] or "").upper().strip()
        category = (r[1] or "staff").lower().strip()
        if name:
            result[name] = category
    return result


async def _store_teacher_sightings(db, sightings: list[dict]) -> int:
    """Store DVR teacher sightings in the database."""
    count = 0
    for s in sightings:
        ts = s.get("timestamp", "")
        date_part = s.get("date", "")
        if not date_part and " " in ts:
            date_part = ts.split(" ")[0]
        if not date_part:
            date_part = datetime.now(IST).strftime("%Y-%m-%d")

        outfit_color = s.get("outfit_color", "")
        outfit_desc = s.get("outfit_description", "")
        outfit_colors = json.dumps(s.get("outfit_colors", []))

        direction = s.get("direction", "IN")

        await db.execute(
            "INSERT INTO teacher_dvr_sightings "
            "(date, timestamp, person_id, name, camera, confidence, "
            "outfit_color, outfit_description, outfit_colors_json, direction) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (date_part, ts, s.get("person_id", ""), s.get("name", ""),
             s.get("camera", ""), s.get("confidence", 0.0),
             outfit_color, outfit_desc, outfit_colors, direction),
        )
        count += 1
    await db.commit()
    return count


async def _get_teacher_sightings(db, date: str) -> list[dict]:
    """Get all DVR teacher sightings for a date."""
    cur = await db.execute(
        "SELECT person_id, name, camera, timestamp, confidence, "
        "outfit_color, outfit_description, outfit_colors_json, direction "
        "FROM teacher_dvr_sightings WHERE date = ? ORDER BY name, timestamp",
        (date,),
    )
    results = []
    for r in await cur.fetchall():
        try:
            outfit_colors = json.loads(r[7]) if r[7] else []
        except (json.JSONDecodeError, TypeError):
            outfit_colors = []
        results.append({
            "person_id": r[0], "name": r[1], "camera": r[2],
            "timestamp": r[3], "confidence": r[4],
            "outfit_color": r[5] or "",
            "outfit_description": r[6] or "",
            "outfit_colors": outfit_colors,
            "direction": r[8] if len(r) > 8 and r[8] else "IN",
        })
    return results


def _classify_visitor(camera: str, timestamp: str) -> str:
    """Classify an unknown visitor based on camera location and time.

    4-category system:
      - Staff (registered): handled elsewhere via TrueFace/face DB match
      - Student: handled elsewhere via name pattern / face DB match
      - Parent: unknown face on entry/reception camera during school hours (7-15)
      - Third-party/Vendor: unknown face during off-hours, on basement camera,
        or unreconciled gate entry without face match
      - Visitor: default fallback for unknown face on other cameras
    """
    cam_upper = camera.upper()
    hour = -1
    try:
        time_part = timestamp.split(" ")[1] if " " in timestamp else timestamp
        hour = int(time_part.split(":")[0])
    except (IndexError, ValueError):
        pass

    is_entry = any(k in cam_upper for k in ("ENTRY", "DISPERSAL"))
    is_reception = "RECEPTION" in cam_upper
    is_basement = "BASEMENT" in cam_upper
    is_monitored = is_entry or is_reception or is_basement

    if is_monitored:
        # Off-hours → Third-party/Vendor (any monitored camera)
        if hour < 7 or hour >= 17:
            return "Third-party/Vendor"
        # Basement → Third-party/Vendor (service providers use basement)
        if is_basement:
            return "Third-party/Vendor"
        # School hours on entry/reception → Parent
        if 7 <= hour < 15:
            return "Parent"
        # After 15:00 (dispersal time) → Parent
        return "Parent"
    return "Visitor"


async def _store_visitor_sightings(db, visitors: list[dict]) -> int:
    """Store DVR visitor (unknown face) sightings in the database."""
    count = 0
    for v in visitors:
        ts = v.get("timestamp", "")
        date_part = v.get("date", "")
        if not date_part and " " in ts:
            date_part = ts.split(" ")[0]
        if not date_part:
            date_part = datetime.now(IST).strftime("%Y-%m-%d")

        cam = v.get("camera", "")
        classification = _classify_visitor(cam, ts)
        snapshot = v.get("snapshot", "")
        direction = v.get("direction", "IN")

        await db.execute(
            "INSERT INTO visitor_dvr_sightings "
            "(date, timestamp, camera, classification, snapshot, direction) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (date_part, ts, cam, classification, snapshot, direction),
        )
        count += 1
    await db.commit()
    return count


async def _get_visitor_sightings(db, date: str) -> list[dict]:
    """Get all DVR visitor sightings for a date."""
    cur = await db.execute(
        "SELECT id, timestamp, camera, classification, snapshot, direction "
        "FROM visitor_dvr_sightings "
        "WHERE date = ? ORDER BY timestamp",
        (date,),
    )
    return [
        {
            "id": r[0], "timestamp": r[1], "camera": r[2],
            "classification": r[3] or "Visitor",
            "snapshot": r[4] or "",
            "direction": r[5] if len(r) > 5 and r[5] else "IN",
        }
        for r in await cur.fetchall()
    ]


async def _store_vehicle_entries(db, entries: list[dict]) -> int:
    """Store vehicle entry events in the database. Returns count stored."""
    count = 0
    for entry in entries:
        ts = entry.get("timestamp", "")
        date_part = ts.split(" ")[0] if " " in ts else datetime.now(IST).strftime("%Y-%m-%d")
        camera = entry.get("camera", "")
        direction = entry.get("direction", "IN")
        vehicle_type = entry.get("vehicle_type", "car")

        snapshot = entry.get("snapshot", "")

        await db.execute(
            "INSERT INTO vehicle_entries (date, timestamp, camera, direction, vehicle_type, snapshot) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (date_part, ts, camera, direction, vehicle_type, snapshot),
        )
        count += 1
    await db.commit()
    return count


async def _get_vehicle_entries(db, date: str) -> list[dict]:
    """Get all vehicle entries for a date."""
    cur = await db.execute(
        "SELECT id, timestamp, camera, direction, vehicle_type, snapshot "
        "FROM vehicle_entries WHERE date = ? ORDER BY timestamp",
        (date,),
    )
    return [
        {"id": r[0], "timestamp": r[1], "camera": r[2],
         "direction": r[3], "vehicle_type": r[4], "snapshot": r[5] if len(r) > 5 else ""}
        for r in await cur.fetchall()
    ]


# ============================================================
# Reconciliation Logic
# ============================================================

async def _reconcile(db, date: str) -> dict:
    """Perform head count reconciliation for a given date.

    Architecture (per school spec):
      Total Persons Entered  = gate counter baseline
      Recognized Students    = students identified by face recognition
      Recognized Staff       = staff identified by TrueFace / face DB
      Total Recognized       = students + staff
      Unrecognized Persons   = Total Entries - Total Recognized
      Current Occupancy      = Total Entries - Total Exits

      Unrecognized persons get temp IDs (U-001, U-002, ...) with:
        entry time, gate, clothing color, last seen camera/time.
      We do NOT classify them as parent/vendor/visitor.
    """
    from collections import Counter

    gate_in_all = await _get_gate_entries(db, date, direction="IN")
    gate_out_all = await _get_gate_entries(db, date, direction="OUT")
    trueface = await _get_trueface_attendance(db, date)
    all_teachers = await _get_all_teachers(db)
    total_registered = len(all_teachers)
    dvr_sightings = await _get_teacher_sightings(db, date)
    visitor_sightings = await _get_visitor_sightings(db, date)
    vehicle_entries = await _get_vehicle_entries(db, date)
    contact_categories = await _get_contact_categories(db)
    gate_entries_with_crops = await _get_gate_entries_with_crops(db, date)

    # Only count entry/exit cameras for gate totals (not internal cameras)
    gate_in_entry = [e for e in gate_in_all if any(k in e["camera"].upper() for k in _ENTRY_CAMERA_KEYWORDS)]
    gate_out_entry = [e for e in gate_out_all if any(k in e["camera"].upper() for k in _ENTRY_CAMERA_KEYWORDS)]
    raw_gate_in = len(gate_in_entry)
    raw_gate_out = len(gate_out_entry)

    # Deduplicate: collapse same-person detections across cameras/scans
    gate_in = _deduplicate_gate_entries(gate_in_entry)
    gate_out = _deduplicate_gate_entries(gate_out_entry)
    unique_gate_in = len(gate_in)
    unique_gate_out = len(gate_out)

    # ── CATEGORY 1: VERIFIED STAFF ──
    # Name aliases: DVR name → TrueFace canonical name (for people registered
    # under different names on different systems)
    _NAME_ALIASES: dict[str, str] = {
        "ALISHA AHUJA": "ALISHA KANWAR",
    }

    def _canonical(name: str) -> str:
        """Return canonical (TrueFace) name, resolving aliases."""
        upper = name.upper().strip()
        return _NAME_ALIASES.get(upper, upper)

    # Build TrueFace lookup by name
    tf_by_name: dict[str, dict] = {}
    for t in trueface:
        tf_by_name[_canonical(t["name"])] = t

    # Build DVR sightings grouped by person
    dvr_by_person: dict[str, list[dict]] = {}
    dvr_names: dict[str, str] = {}
    dvr_outfits: dict[str, list[dict]] = {}
    for s in dvr_sightings:
        key = _canonical(s["name"])
        dvr_by_person.setdefault(key, []).append(s)
        dvr_names[key] = s["name"]
        if s.get("outfit_color") and s["outfit_color"] != "unknown":
            dvr_outfits.setdefault(key, []).append({
                "color": s["outfit_color"],
                "description": s.get("outfit_description", ""),
                "colors": s.get("outfit_colors", []),
                "camera": s.get("camera", ""),
                "timestamp": s.get("timestamp", ""),
            })

    # All unique registered staff names (using canonical names)
    all_names: set[str] = set()
    for t in all_teachers:
        all_names.add(_canonical(t["name"]))
    for key in tf_by_name:
        all_names.add(key)
    for key in dvr_by_person:
        all_names.add(key)

    # Per-person detail with proper category separation
    staff_detail = []
    for name_upper in sorted(all_names):
        tf = tf_by_name.get(name_upper)
        dvr_list = dvr_by_person.get(name_upper, [])

        dvr_cameras = []
        dvr_times = []
        dvr_sighting_details = []
        for s in dvr_list:
            cam = s["camera"]
            ts = s["timestamp"]
            time_part = ts.split(" ")[1] if " " in ts else ts
            if cam not in dvr_cameras:
                dvr_cameras.append(cam)
            dvr_times.append(time_part)
            dvr_sighting_details.append({
                "camera": cam, "time": time_part,
                "direction": s.get("direction", "IN"),
                "confidence": s.get("confidence", 0),
                "outfit_color": s.get("outfit_color", ""),
                "outfit_description": s.get("outfit_description", ""),
            })

        display_name = tf["name"] if tf else dvr_names.get(name_upper, name_upper.title())

        outfit_observations = dvr_outfits.get(name_upper, [])
        outfit_summary = outfit_observations[-1].get("description", "unknown") if outfit_observations else "-"

        raw_category = contact_categories.get(name_upper, "staff")
        category = _normalize_category(raw_category, name=display_name)

        # Determine presence status using TrueFace + DVR direction data
        is_present = tf is not None or len(dvr_list) > 0
        has_departed_tf = tf is not None and tf.get("departure_time") is not None
        # Check if last DVR sighting was on an exit/dispersal camera
        dvr_exit_sightings = [s for s in dvr_list if s.get("direction") == "OUT"]
        dvr_entry_sightings = [s for s in dvr_list if s.get("direction") == "IN"]
        has_departed_dvr = (len(dvr_exit_sightings) > 0 and
                            (not dvr_entry_sightings or
                             dvr_exit_sightings[-1].get("timestamp", "") >
                             dvr_entry_sightings[-1].get("timestamp", "")))
        has_departed = has_departed_tf or has_departed_dvr

        if is_present and not has_departed:
            occupancy_status = "INSIDE"
        elif is_present and has_departed:
            occupancy_status = "EXITED"
        else:
            occupancy_status = "ABSENT"

        # Time trail with entry/exit events
        time_trail: list[dict] = []
        if tf and tf.get("arrival_time"):
            time_trail.append({"time": tf["arrival_time"], "source": "TrueFace 3000", "event": "Arrival"})
        for det in dvr_sighting_details:
            direction = det.get("direction", "IN")
            event_type = "DVR Exit" if direction == "OUT" else "DVR Sighting"
            time_trail.append({"time": det["time"], "source": det["camera"], "event": event_type})
        if tf and tf.get("departure_time"):
            time_trail.append({"time": tf["departure_time"], "source": "TrueFace 3000", "event": "Departure"})
        time_trail.sort(key=lambda x: x["time"])

        staff_detail.append({
            "name": display_name,
            "category": category,
            "raw_category": raw_category,
            "trueface_present": tf is not None,
            "trueface_arrival": tf["arrival_time"] if tf else None,
            "trueface_departure": tf["departure_time"] if tf else None,
            "trueface_pin": tf["pin"] if tf else None,
            "dvr_seen": len(dvr_list) > 0,
            "dvr_sighting_count": len(dvr_list),
            "dvr_cameras": dvr_cameras,
            "dvr_times": dvr_times,
            "dvr_first_seen": dvr_times[0] if dvr_times else None,
            "dvr_last_seen": dvr_times[-1] if dvr_times else None,
            "dvr_sighting_details": dvr_sighting_details,
            "outfit_summary": outfit_summary,
            "outfit_observations": outfit_observations,
            "time_trail": time_trail,
            "occupancy_status": occupancy_status,
            "status": _reconciliation_status(tf is not None, len(dvr_list) > 0),
        })

    # Staff presence counts
    staff_present = [s for s in staff_detail if s["occupancy_status"] in ("INSIDE", "EXITED")]
    staff_inside = [s for s in staff_detail if s["occupancy_status"] == "INSIDE"]
    staff_exited = [s for s in staff_detail if s["occupancy_status"] == "EXITED"]
    staff_absent = [s for s in staff_detail if s["occupancy_status"] == "ABSENT"]

    # Category breakdown for staff INSIDE
    staff_category_inside: Counter = Counter()
    for s in staff_inside:
        staff_category_inside[s["category"]] += 1

    # Category breakdown for all present staff
    staff_category_present: Counter = Counter()
    for s in staff_present:
        staff_category_present[s["category"]] += 1

    # Side-by-side (kept for compatibility)
    side_by_side = {
        "both_present": [t for t in staff_detail if t["trueface_present"] and t["dvr_seen"]],
        "trueface_only": [t for t in staff_detail if t["trueface_present"] and not t["dvr_seen"]],
        "dvr_only": [t for t in staff_detail if not t["trueface_present"] and t["dvr_seen"]],
        "neither": [t for t in staff_detail if not t["trueface_present"] and not t["dvr_seen"]],
    }

    # ── CATEGORY 2: UNKNOWN / UNREGISTERED PEOPLE ──
    # Build unknown person list from visitor sightings + gate entries,
    # deduplicating detections within 2 minutes on the same camera.
    _raw_unknowns: list[dict] = []
    for v in visitor_sightings:
        ts = v.get("timestamp", "")
        time_part = ts.split(" ")[1] if " " in ts else ts
        v_direction = v.get("direction", "IN")
        _raw_unknowns.append({
            "timestamp": ts,
            "time": time_part,
            "camera": v.get("camera", "Unknown"),
            "direction": v_direction,
            "status": "INSIDE" if v_direction == "IN" else "EXITED",
            "attire_color": v.get("attire_color", ""),
            "person_crop": v.get("snapshot", ""),
        })
    for g in gate_entries_with_crops:
        if g.get("person_crop"):
            ts = g.get("timestamp", "")
            time_part = ts.split(" ")[1] if " " in ts else ts
            _raw_unknowns.append({
                "timestamp": ts,
                "time": time_part,
                "camera": g.get("camera", "Unknown"),
                "direction": g.get("direction", "IN"),
                "status": "INSIDE" if g.get("direction") == "IN" else "EXITED",
                "person_crop": g.get("person_crop", ""),
            })

    # Deduplicate: same camera within 2 minutes → same person
    _raw_unknowns.sort(key=lambda x: (x["camera"], x["timestamp"]))
    unknown_persons: list[dict] = []
    _seen: list[tuple[str, str]] = []  # (camera, timestamp) of accepted entries
    for u in _raw_unknowns:
        cam = u["camera"]
        ts = u["timestamp"]
        is_dup = False
        for prev_cam, prev_ts in _seen:
            if prev_cam != cam:
                continue
            try:
                from datetime import datetime as _dt
                t1 = _dt.strptime(prev_ts[-8:], "%H:%M:%S")
                t2 = _dt.strptime(ts[-8:], "%H:%M:%S")
                if abs((t2 - t1).total_seconds()) < 120:
                    is_dup = True
                    break
            except (ValueError, IndexError):
                pass
        if not is_dup:
            _seen.append((cam, ts))
            u["id"] = f"U-{len(unknown_persons) + 1:03d}"
            unknown_persons.append(u)

    # Count unknown entries/exits from the actual unknown person list
    unknown_in_list = [u for u in unknown_persons if u["direction"] == "IN"]
    unknown_out_list = [u for u in unknown_persons if u["direction"] == "OUT"]
    estimated_unknown_in = len(unknown_in_list)
    estimated_unknown_out = len(unknown_out_list)
    estimated_unknown_inside = max(0, estimated_unknown_in - estimated_unknown_out)
    unique_staff_detected = len(staff_present)

    # Classification removed per spec — all unknowns are "UNRECOGNIZED PERSON"
    classification_counts: dict[str, int] = {}

    # Visitor/unknown grouped by camera
    visitor_by_camera: dict[str, list[dict]] = {}
    for v in visitor_sightings:
        cam = v.get("camera", "Unknown")
        visitor_by_camera.setdefault(cam, []).append(v)

    # ── VEHICLES ──
    vehicles_in = [v for v in vehicle_entries if v["direction"] == "IN"]
    vehicles_out = [v for v in vehicle_entries if v["direction"] == "OUT"]
    vehicle_type_counts = dict(Counter(v["vehicle_type"] for v in vehicles_in))

    # ── CATEGORY 3: CROSS-REFERENCE GATE ENTRIES WITH FACE DETECTIONS ──
    # Match each gate entry to a face detection (staff arrival or visitor
    # sighting) by timestamp proximity. Unmatched entries = people who
    # entered but could NOT be identified by any face system.
    _CROSS_REF_WINDOW = 300  # 5-minute matching window

    def _parse_ts(ts_str: str) -> datetime | None:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%H:%M:%S"):
            try:
                return datetime.strptime(ts_str, fmt)
            except (ValueError, TypeError):
                pass
        return None

    # Build face detection timeline (all identified arrivals)
    face_events: list[dict] = []
    for s in staff_present:
        arr = s.get("trueface_arrival") or s.get("dvr_first_seen")
        if arr:
            face_events.append({"time": arr, "name": s["name"], "type": "staff"})
    for v in visitor_sightings:
        ts = v.get("timestamp", "")
        t_part = ts.split(" ")[1] if " " in ts else ts
        face_events.append({"time": t_part, "name": "Unknown", "type": "visitor"})

    # Cross-reference each gate IN entry against face detections
    matched_face_indices: set[int] = set()
    reconciled_gate: list[dict] = []
    unreconciled_gate: list[dict] = []

    for g in gate_in:
        g_ts = _parse_ts(g.get("timestamp", ""))
        if g_ts is None:
            continue
        best_match = None
        best_diff = _CROSS_REF_WINDOW + 1
        best_idx = -1
        for idx, fe in enumerate(face_events):
            if idx in matched_face_indices:
                continue
            fe_ts = _parse_ts(fe["time"])
            if fe_ts is None:
                continue
            # Compare only time-of-day (handle full timestamp vs time-only)
            g_secs = g_ts.hour * 3600 + g_ts.minute * 60 + g_ts.second
            fe_secs = fe_ts.hour * 3600 + fe_ts.minute * 60 + fe_ts.second
            diff = abs(g_secs - fe_secs)
            if diff < best_diff:
                best_diff = diff
                best_match = fe
                best_idx = idx
        if best_match and best_diff <= _CROSS_REF_WINDOW:
            matched_face_indices.add(best_idx)
            reconciled_gate.append({
                "gate_timestamp": g.get("timestamp", ""),
                "camera": g.get("camera", ""),
                "attire_color": g.get("attire_color", ""),
                "matched_to": best_match["name"],
                "matched_type": best_match["type"],
                "time_diff_seconds": best_diff,
            })
        else:
            ts_str = g.get("timestamp", "")
            time_part = ts_str.split(" ")[1] if " " in ts_str else ts_str
            unreconciled_gate.append({
                "id": f"U-{len(unknown_persons) + len(unreconciled_gate) + 1:03d}",
                "timestamp": ts_str,
                "time": time_part,
                "camera": g.get("camera", "Unknown"),
                "attire_color": g.get("attire_color", "unknown"),
                "direction": g.get("direction", "IN"),
                "person_crop": g.get("person_crop", ""),
                "status": "INSIDE",
            })

    unreconciled_count = len(unreconciled_gate)

    # ── OCCUPANCY SUMMARY (per school spec) ──
    # Current Occupancy = Total Entries - Total Exits (gate counter baseline)
    total_staff_inside = len(staff_inside)
    total_staff_exited = len(staff_exited)
    total_unknown_exited = estimated_unknown_out

    # Recognized counts
    recognized_students = [s for s in staff_present if s["category"] == "Students"]
    recognized_staff = [s for s in staff_present if s["category"] != "Students"]
    total_recognized = len(staff_present)  # all recognized (staff + students)

    # Total entries = gate counter (unique IN crossings)
    total_entries = unique_gate_in
    total_exits = unique_gate_out

    # Unrecognized = Total Entries - Total Recognized
    total_unrecognized = max(0, total_entries - total_recognized)

    # Current Occupancy = Total Entries - Total Exits
    current_occupancy = max(0, total_entries - total_exits)

    # Legacy: keep for backward compat
    total_unknown_inside = estimated_unknown_inside
    total_unreconciled_inside = unreconciled_count
    total_inside = current_occupancy  # use the new formula

    # ── ALERTS ──
    alerts: list[dict] = []
    for t in side_by_side["trueface_only"]:
        alerts.append({
            "type": "ATTENDANCE_WITHOUT_ENTRY",
            "severity": "medium",
            "person": t["name"],
            "category": t.get("category", "staff"),
            "detail": f"TrueFace at {t.get('trueface_arrival', '?')} but not seen on any DVR camera.",
        })
    for t in side_by_side["dvr_only"]:
        alerts.append({
            "type": "ENTRY_WITHOUT_ATTENDANCE",
            "severity": "high",
            "person": t["name"],
            "category": t.get("category", "staff"),
            "detail": f"Seen on DVR ({', '.join(t.get('dvr_cameras', []))}) at {t.get('dvr_first_seen', '?')} but NOT marked on TrueFace.",
        })
    if total_unrecognized > 0:
        alerts.append({
            "type": "UNRECOGNIZED_PERSONS",
            "severity": "high" if total_unrecognized > 10 else "medium",
            "person": "",
            "category": "unrecognized",
            "detail": (
                f"{total_unrecognized} person(s) entered the school but could NOT "
                f"be recognized as Student or Staff. "
                f"(Total Entries: {total_entries} - Recognized: {total_recognized})"
            ),
        })
    # Flag student detections during no-school period (likely false positives)
    student_detections = [s for s in staff_detail
                         if s["category"] == "Students" and s["occupancy_status"] != "ABSENT"]
    if student_detections:
        names = ", ".join(s["name"] for s in student_detections)
        alerts.append({
            "type": "STUDENT_DETECTED",
            "severity": "low",
            "person": names,
            "category": "Students",
            "detail": (f"{len(student_detections)} student(s) detected on DVR: {names}. "
                       "May be false positive if school is not in session."),
        })

    # ── GATE vs FACE DISCREPANCY ──
    total_identified = unique_staff_detected + estimated_unknown_in
    total_accounted = total_identified  # faces matched by any system

    # Group gate entries by hourly time window for discrepancy breakdown
    hourly_gate: dict[str, int] = {}
    hourly_identified: dict[str, int] = {}
    for g in gate_in:
        ts = g.get("timestamp", "")
        hour_key = ts[11:13] + ":00" if len(ts) > 13 else "unknown"
        hourly_gate[hour_key] = hourly_gate.get(hour_key, 0) + 1
    # Staff sightings by hour
    for s in staff_detail:
        if s.get("dvr_first_seen"):
            hour_key = s["dvr_first_seen"][:2] + ":00" if len(s["dvr_first_seen"]) >= 2 else "unknown"
            hourly_identified[hour_key] = hourly_identified.get(hour_key, 0) + 1
    # Visitor sightings by hour
    for v in visitor_sightings:
        ts = v.get("timestamp", "")
        hour_key = ts[11:13] + ":00" if len(ts) > 13 else "unknown"
        hourly_identified[hour_key] = hourly_identified.get(hour_key, 0) + 1

    all_hours = sorted(set(list(hourly_gate.keys()) + list(hourly_identified.keys())))
    hourly_discrepancy = []
    for h in all_hours:
        if h == "unknown":
            continue
        gate_count = hourly_gate.get(h, 0)
        id_count = hourly_identified.get(h, 0)
        gap = max(0, gate_count - id_count)
        hourly_discrepancy.append({
            "hour": h,
            "gate_entries": gate_count,
            "identified": id_count,
            "unreconciled": gap,
        })

    discrepancy = {
        "gate_head_count": unique_gate_in,
        "faces_identified": total_identified,
        "staff_identified": unique_staff_detected,
        "visitors_identified": estimated_unknown_in,
        "unreconciled": unreconciled_count,
        "reconciled_gate_entries": len(reconciled_gate),
        "reconciliation_rate": round(
            (total_accounted / unique_gate_in * 100) if unique_gate_in > 0 else 100, 1
        ),
        "hourly_breakdown": hourly_discrepancy,
        "unreconciled_entries": unreconciled_gate[:50],
        "reconciled_entries": reconciled_gate[:50],
    }

    # Update daily summary
    await db.execute(
        "INSERT OR REPLACE INTO gate_daily_summary "
        "(date, total_in, total_out, trueface_matched, unreconciled) "
        "VALUES (?, ?, ?, ?, ?)",
        (date, raw_gate_in, raw_gate_out, len(trueface), unreconciled_count),
    )
    await db.commit()

    return {
        "date": date,
        # ── ENTRY SUMMARY ──
        "total_entries": total_entries,  # unique gate IN (baseline)
        "total_exits": total_exits,     # unique gate OUT
        "raw_gate_in": raw_gate_in,
        "raw_gate_out": raw_gate_out,
        "unique_gate_in": unique_gate_in,
        "unique_gate_out": unique_gate_out,
        "total_gate_in": unique_gate_in,   # backward compat
        "total_gate_out": unique_gate_out,  # backward compat
        # ── RECOGNITION SUMMARY ──
        "recognized_students": len(recognized_students),
        "recognized_staff_count": len(recognized_staff),
        "total_recognized": total_recognized,
        "total_registered": total_registered,
        "trueface_identified": len(trueface),
        "dvr_sighted": len([s for s in staff_detail if s["dvr_seen"]]),
        "staff_detail": staff_detail,
        "teacher_detail": staff_detail,  # backward compat
        "side_by_side": side_by_side,
        # ── UNRECOGNIZED SUMMARY ──
        "total_unrecognized": total_unrecognized,
        # ── CURRENT OCCUPANCY ──
        "current_occupancy": current_occupancy,
        "total_inside": total_inside,  # = current_occupancy
        # ── Staff breakdown ──
        "staff_inside": total_staff_inside,
        "staff_exited": total_staff_exited,
        "staff_absent": len(staff_absent),
        "staff_category_inside": dict(staff_category_inside),
        "staff_category_present": dict(staff_category_present),
        "category_counts": dict(staff_category_present),
        # ── Unknown persons (face-detected unknowns) ──
        "unknown_persons": unknown_persons,
        "estimated_unknown_in": estimated_unknown_in,
        "estimated_unknown_out": estimated_unknown_out,
        "estimated_unknown_inside": estimated_unknown_inside,
        "visitor_count": estimated_unknown_in,
        "visitor_sightings": visitor_sightings,
        "visitor_by_camera": visitor_by_camera,
        # ── Unreconciled gate entries ──
        "unreconciled_gate_entries": unreconciled_gate,
        "reconciled_gate_entries": reconciled_gate,
        "total_unreconciled_inside": total_unreconciled_inside,
        "classification_counts": classification_counts,
        # ── Exits ──
        "total_exited": total_staff_exited + total_unknown_exited,
        # ── Vehicles ──
        "vehicles_in": len(vehicles_in),
        "vehicles_out": len(vehicles_out),
        "vehicle_types": vehicle_type_counts,
        "vehicle_entries": vehicle_entries,
        # Snapshots
        "gate_entries_with_crops": gate_entries_with_crops,
        # Alerts
        "alerts": alerts,
        # Gate vs Face discrepancy
        "discrepancy": discrepancy,
        # Legacy
        "total_registered_teachers": total_registered,
        "unreconciled_count": unreconciled_count,
        "timing_trail": [],
    }


# Grade/section pattern: digits followed by optional section letter
# Matches names like "garv jain 9 - B", "student 10-A", "name 1 - C"
_STUDENT_GRADE_RE = re.compile(
    r"\b(?:nursery|prep|[kK][gG]|[1-9]|1[0-2])\s*[-–]\s*[A-Za-z]\b"
)


def _is_student_name(name: str) -> bool:
    """Detect if a name contains a grade/section pattern (e.g. '9 - B')."""
    return bool(_STUDENT_GRADE_RE.search(name))


def _normalize_category(raw: str, name: str = "") -> str:
    """Normalize raw TrueFace category into display group.

    Also detects students by name pattern (grade-section in name).
    """
    # Student detection by name pattern (overrides all other categories)
    if name and _is_student_name(name):
        return "Students"

    raw = (raw or "staff").lower().strip()
    if raw == "student":
        return "Students"
    teacher_keywords = ("teacher", "eng-teacher", "comp. teacher", "art teacher",
                        "music teacher", "sports teacher", "dance teacher",
                        "theatre teacher", "trainee teacher", "lib. teacher",
                        "taekwondo tea.", "principal")
    admin_keywords = ("admin", "admin it", "account", "est. manager",
                      "lab. incharge", "max nurse")
    if any(kw == raw for kw in teacher_keywords):
        return "Teachers"
    if any(kw == raw for kw in admin_keywords):
        return "Admin / Office Staff"
    if raw in ("advocate",):
        return "Advocates / Legal"
    if raw in ("staff",):
        return "Staff"
    if raw in ("web designers",):
        return "IT / Design"
    return raw.title()


def _reconciliation_status(trueface_present: bool, dvr_seen: bool) -> str:
    """Return a human-readable reconciliation status."""
    if trueface_present and dvr_seen:
        return "✓ FULLY VERIFIED"
    if trueface_present and not dvr_seen:
        return "⚠ TRUEFACE ONLY"
    if not trueface_present and dvr_seen:
        return "⚠ ENTRY ONLY (DVR)"
    return "— ABSENT"


# ============================================================
# Unknown Person WhatsApp Alert
# ============================================================

def _generate_placeholder_image() -> str:
    """Generate a small placeholder PNG as base64 for alerts without a snapshot."""
    import struct
    import zlib

    width, height = 200, 200
    raw = b""
    for y in range(height):
        raw += b"\x00"  # filter byte
        for x in range(width):
            raw += b"\xcc\xcc\xcc"  # light gray
    compressed = zlib.compress(raw)

    def _chunk(ctype: bytes, data: bytes) -> bytes:
        c = ctype + data
        return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)

    ihdr = struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)
    png = b"\x89PNG\r\n\x1a\n" + _chunk(b"IHDR", ihdr) + _chunk(b"IDAT", compressed) + _chunk(b"IEND", b"")
    return base64.b64encode(png).decode()


async def _send_unknown_person_alert(entry: dict):
    """Send WhatsApp template alert with 3-angle snapshots for an unrecognized person.

    Uses the 'unknown_person_detected' template (IMAGE header + body params)
    so alerts work outside the 24-hour conversation window.
    Sends up to 3 photos: face close-up (template header), body shot, context.
    Deduplicates by camera to avoid spam (one alert per camera per 60s).
    """
    import time
    camera = entry.get("camera", "Unknown")
    now = time.time()

    cache_key = camera
    last_sent = _unknown_alert_last.get(cache_key, 0)
    if now - last_sent < _UNKNOWN_ALERT_COOLDOWN:
        logger.info("[GATE] Alert skipped — cooldown active for %s (%.0fs left)",
                    camera, _UNKNOWN_ALERT_COOLDOWN - (now - last_sent))
        return

    _unknown_alert_last[cache_key] = now

    ts = entry.get("timestamp", "")
    direction = entry.get("direction", "IN")
    person_crop = entry.get("person_crop", "")
    snapshot_face = entry.get("snapshot_face", "")
    snapshot_body = entry.get("snapshot_body", "")
    snapshot_context = entry.get("snapshot_context", "")

    # Primary image: prefer face close-up, fall back to person_crop
    primary_image = snapshot_face or person_crop

    try:
        from app.services.whatsapp_service import (
            upload_base64_image_cloud,
            send_cloud_template_message,
            send_cloud_media,
        )

        image_data = primary_image if primary_image else _generate_placeholder_image()
        header_image_id = await upload_base64_image_cloud(image_data)
        if not header_image_id:
            logger.warning("[GATE] Failed to upload image for alert (camera=%s)", camera)
            return

        # Upload body and context photos for follow-up messages
        body_image_id = None
        context_image_id = None
        if snapshot_body:
            body_image_id = await upload_base64_image_cloud(snapshot_body)
        if snapshot_context:
            context_image_id = await upload_base64_image_cloud(snapshot_context)

        for phone in UNKNOWN_ALERT_PHONES:
            try:
                # 1) Template message with face close-up
                sent = await send_cloud_template_message(
                    to=phone,
                    template_name=UNKNOWN_ALERT_TEMPLATE,
                    language_code="en",
                    body_params=[camera, ts, direction],
                    header_image_id=header_image_id,
                )
                logger.info("[GATE] Unknown person alert → %s: %s (camera=%s)",
                            phone, "OK" if sent else "FAILED", camera)

                # 2) Follow-up: body shot with red highlight box
                if body_image_id:
                    await send_cloud_media(
                        to=phone,
                        media_type="image",
                        media_id=body_image_id,
                        caption=f"\U0001f534 Full body — {camera} at {ts}",
                    )

                # 3) Follow-up: context shot (full camera frame)
                if context_image_id:
                    await send_cloud_media(
                        to=phone,
                        media_type="image",
                        media_id=context_image_id,
                        caption=f"\U0001f4f7 Gate context — {camera} at {ts}",
                    )

            except Exception as inner_e:
                logger.error("[GATE] Alert to %s failed: %s", phone, inner_e)

    except Exception as e:
        logger.error("[GATE] Unknown person alert failed: %s (camera=%s)", e, camera)


# ============================================================
# Endpoints
# ============================================================

@router.post("/api/gate/entry")
async def receive_gate_entries(request: Request):
    """Receive gate entry events from the campus agent gate counter.

    Body: [{"timestamp": "...", "camera": "...", "direction": "IN", "attire_color": "blue", "person_crop": "..."}]
    """
    body = await request.json()
    entries = body if isinstance(body, list) else [body]

    db = await _get_db()
    try:
        count = await _store_gate_entries(db, entries)
    finally:
        await db.close()

    # Note: WhatsApp alerts for unknown persons are triggered from
    # /api/gate/visitor-sighting (face-recognition-verified unknowns),
    # not from gate entries (which include all persons, known and unknown).

    logger.info("[GATE] Stored %d gate entry event(s)", count)
    return {"status": "ok", "stored": count}


@router.post("/api/gate/teacher-sighting")
async def receive_teacher_sightings(request: Request):
    """Receive DVR teacher face sightings from the campus agent.

    Body: [{"person_id": "TEACHER_X", "name": "...", "camera": "...",
            "timestamp": "...", "date": "...", "confidence": 0.85}]
    """
    body = await request.json()
    sightings = body if isinstance(body, list) else [body]

    db = await _get_db()
    try:
        count = await _store_teacher_sightings(db, sightings)
    finally:
        await db.close()

    logger.info("[GATE] Stored %d DVR teacher sighting(s)", count)
    return {"status": "ok", "stored": count}


@router.post("/api/gate/visitor-sighting")
async def receive_visitor_sightings(request: Request):
    """Receive DVR visitor (unknown face) sightings from the campus agent.

    Body: [{"camera": "...", "timestamp": "...", "date": "...", "snapshot": "base64..."}]
    These are faces that did NOT match any registered person in the face DB.
    """
    body = await request.json()
    visitors = body if isinstance(body, list) else [body]

    db = await _get_db()
    try:
        count = await _store_visitor_sightings(db, visitors)
    finally:
        await db.close()

    # Send WhatsApp alert for each unknown visitor at entry cameras
    alert_count = 0
    skipped_non_entry = 0
    for v in visitors:
        cam = v.get("camera", "")
        cam_upper = cam.upper()
        is_entry_cam = any(k in cam_upper for k in _ENTRY_CAMERA_KEYWORDS)
        v_direction = v.get("direction", "IN")
        if not is_entry_cam:
            skipped_non_entry += 1
        if is_entry_cam and v_direction == "IN":
            alert_entry = {
                "camera": cam,
                "timestamp": v.get("timestamp", ""),
                "direction": v_direction,
                "person_crop": v.get("snapshot", ""),
                "snapshot_face": v.get("snapshot_face", ""),
                "snapshot_body": v.get("snapshot_body", ""),
                "snapshot_context": v.get("snapshot_context", ""),
            }
            alert_count += 1
            asyncio.create_task(_send_unknown_person_alert(alert_entry))
    logger.info("[GATE] Visitor sighting alert summary: queued=%d, skipped_non_entry=%d, total=%d",
                alert_count, skipped_non_entry, len(visitors))

    # Log classification breakdown
    cls_log = {}
    for v in visitors:
        cam = v.get("camera", "")
        ts = v.get("timestamp", "")
        cls = _classify_visitor(cam, ts)
        cls_log[cls] = cls_log.get(cls, 0) + 1
    cls_str = ", ".join(f"{k}={v}" for k, v in sorted(cls_log.items()))
    logger.info("[GATE] Stored %d DVR visitor sighting(s) [%s]", count, cls_str)
    return {"status": "ok", "stored": count}


@router.post("/api/gate/vehicle-entry")
async def receive_vehicle_entries(request: Request):
    """Receive vehicle entry events from the campus agent gate counter.

    Body: [{"timestamp": "...", "camera": "...", "direction": "IN", "vehicle_type": "car"}]
    """
    body = await request.json()
    entries = body if isinstance(body, list) else [body]

    db = await _get_db()
    try:
        count = await _store_vehicle_entries(db, entries)
    finally:
        await db.close()

    logger.info("[GATE] Stored %d vehicle entry event(s)", count)
    return {"status": "ok", "stored": count}


@router.get("/api/gate/status")
async def gate_status():
    """Get today's running head count totals."""
    today = datetime.now(IST).strftime("%Y-%m-%d")
    db = await _get_db()
    try:
        gate_in = await _get_gate_entries(db, today, direction="IN")
        gate_out = await _get_gate_entries(db, today, direction="OUT")
        trueface = await _get_trueface_attendance(db, today)
        dvr_sightings = await _get_teacher_sightings(db, today)
        visitor_sightings = await _get_visitor_sightings(db, today)
        vehicle_entries = await _get_vehicle_entries(db, today)
    finally:
        await db.close()

    # Count unique teachers seen on DVR
    dvr_unique = len({s["person_id"] for s in dvr_sightings})

    # Estimate visitors: gate entries minus identified staff
    staff_detected = len(trueface) + len({s["person_id"] for s in dvr_sightings
                                          if s["name"].upper().strip() not in
                                          {t["name"].upper().strip() for t in trueface}})
    estimated_visitors = max(0, len(gate_in) - staff_detected)
    final_visitors = max(len(visitor_sightings), estimated_visitors)

    # Vehicle counts by type
    vehicles_in = [v for v in vehicle_entries if v["direction"] == "IN"]
    vehicles_out = [v for v in vehicle_entries if v["direction"] == "OUT"]
    from collections import Counter
    vehicle_types = dict(Counter(v["vehicle_type"] for v in vehicles_in))

    return {
        "date": today,
        "gate_in": len(gate_in),
        "gate_out": len(gate_out),
        "trueface_identified": len(trueface),
        "dvr_teachers_sighted": dvr_unique,
        "dvr_total_sightings": len(dvr_sightings),
        "visitors_detected": final_visitors,
        "vehicles_in": len(vehicles_in),
        "vehicles_out": len(vehicles_out),
        "vehicle_types": vehicle_types,
    }


@router.get("/api/gate/reconciliation/{date}")
async def get_reconciliation(date: str):
    """Get full reconciliation data for a specific date."""
    db = await _get_db()
    try:
        result = await _reconcile(db, date)
    finally:
        await db.close()
    return result


# ============================================================
# Excel Report Generation
# ============================================================

def _generate_reconciliation_excel(recon: dict) -> bytes:
    """Generate the comprehensive School Headcount Reconciliation report.

    Sheets:
      1. Summary — overall counts, category breakdown, reconciliation status
      2. Verified Movements — per-person detail with category, outfit, timestamps
      3. Sighting Timeline — every DVR sighting chronologically with category
      4. Reconciliation — side-by-side comparison by status category
      5. Visitors — visitor head count, per-camera, timeline
      6. Outfit Reconciliation — outfit tracking across sightings
      7. AI Observations — automated intelligence notes
      8. Time Trail — movement reconstruction per person across cameras
      9. Mismatch Alerts — attendance/entry mismatches, unknown persons
    """
    from collections import Counter

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    light_blue_fill = PatternFill("solid", fgColor="D9E2F3")
    orange_fill = PatternFill("solid", fgColor="FFA500")
    dark_red_fill = PatternFill("solid", fgColor="C00000")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    date_str = recon["date"]
    side_by_side = recon.get("side_by_side", {})
    staff_detail = recon.get("staff_detail", recon.get("teacher_detail", []))
    visitor_sightings = recon.get("visitor_sightings", [])
    visitor_by_camera = recon.get("visitor_by_camera", {})
    unknown_persons = recon.get("unknown_persons", [])

    staff_inside = [s for s in staff_detail if s.get("occupancy_status") == "INSIDE"]
    staff_exited = [s for s in staff_detail if s.get("occupancy_status") == "EXITED"]
    staff_absent = [s for s in staff_detail if s.get("occupancy_status") == "ABSENT"]

    fully_verified = side_by_side.get("both_present", [])
    trueface_only = side_by_side.get("trueface_only", [])
    dvr_only = side_by_side.get("dvr_only", [])
    absent = side_by_side.get("neither", [])

    total_inside = recon.get("total_inside", 0)
    estimated_unknown_inside = recon.get("estimated_unknown_inside", 0)
    disc = recon.get("discrepancy", {})

    # ── Sheet 1: Summary (per school spec) ──
    ws = wb.active
    ws.title = "Reconciliation"

    ws.merge_cells("A1:D1")
    ws["A1"] = "HEAD COUNT RECONCILIATION REPORT"
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("A2:D2")
    ws["A2"] = f"PP International School - {date_str}"
    ws["A2"].font = Font(bold=True, size=11)

    total_entries = recon.get("total_entries", recon.get("unique_gate_in", 0))
    total_exits = recon.get("total_exits", recon.get("unique_gate_out", 0))
    recognized_students = recon.get("recognized_students", 0)
    recognized_staff_count = recon.get("recognized_staff_count", 0)
    total_recognized = recon.get("total_recognized", len(staff_present))
    total_unrecognized = recon.get("total_unrecognized", 0)
    current_occupancy = recon.get("current_occupancy", total_inside)

    unreconciled_entries = recon.get("unreconciled_gate_entries", [])

    r = 4
    summary_rows = [
        ("ENTRY SUMMARY", "", None),
        ("  Total Persons Entered", total_entries, green_fill),
        ("", "", None),
        ("RECOGNITION SUMMARY", "", None),
        ("  Recognized Students", recognized_students, light_blue_fill),
        ("  Recognized Staff", recognized_staff_count, light_blue_fill),
        ("  Total Recognized", total_recognized, green_fill),
        ("", "", None),
    ]
    for cat_name, cat_count in sorted(recon.get("staff_category_present", {}).items()):
        summary_rows.append((f"    {cat_name}", cat_count, None))
    summary_rows += [
        ("", "", None),
        ("UNRECOGNIZED SUMMARY", "", None),
        ("  Total Unrecognized Persons", total_unrecognized,
         red_fill if total_unrecognized > 0 else green_fill),
        ("  Formula: Entries - Recognized", f"{total_entries} - {total_recognized}", None),
        ("", "", None),
        ("EXIT SUMMARY", "", None),
        ("  Total Persons Exited", total_exits, None),
        ("", "", None),
        ("CURRENT OCCUPANCY", "", None),
        ("  Current Occupancy", current_occupancy, green_fill),
        ("  Formula: Entries - Exits", f"{total_entries} - {total_exits}", None),
        ("", "", None),
        ("RECONCILIATION CHECK", "", None),
        ("  Total Entries", total_entries, None),
        ("  Total Recognized", total_recognized, None),
        ("  Difference (Unrecognized)", total_unrecognized,
         red_fill if total_unrecognized > 0 else green_fill),
    ]
    recon_rate = round((total_recognized / total_entries * 100) if total_entries > 0 else 100, 1)
    summary_rows += [
        ("  Reconciliation Rate", f"{recon_rate}%",
         green_fill if recon_rate >= 80 else red_fill),
        ("", "", None),
        ("STAFF RECONCILIATION", "", None),
        ("  Fully Verified (TrueFace + DVR)", len(fully_verified), green_fill),
        ("  Entry Only (DVR - Not on TrueFace)", len(dvr_only), red_fill),
        ("  TrueFace Only (No DVR Sighting)", len(trueface_only), yellow_fill),
        ("  Not Detected / Absent", len(absent), gray_fill),
        ("", "", None),
        ("VEHICLE COUNT", "", None),
        ("  Vehicles IN", recon.get("vehicles_in", 0), light_blue_fill),
        ("  Vehicles OUT", recon.get("vehicles_out", 0), None),
    ]
    for vtype, vcount in sorted(recon.get("vehicle_types", {}).items()):
        summary_rows.append((f"    {vtype.title()}", vcount, None))
    for label, value, fill in summary_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=value).font = Font(bold=True, size=12)
        ws.cell(row=r, column=2).border = border
        if fill:
            ws.cell(row=r, column=1).fill = fill
            ws.cell(row=r, column=2).fill = fill
        r += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 15

    # ── Sheet 2: Staff Detail (with occupancy status) ──
    ws2 = wb.create_sheet("Staff Detail")

    ws2.merge_cells("A1:I1")
    ws2["A1"] = f"STAFF DETAIL - {date_str}"
    ws2["A1"].font = Font(bold=True, size=14)

    headers2 = [
        "#", "Name", "Category", "Status", "Occupancy",
        "Arrived", "Departed", "Cameras Seen", "Wearing",
    ]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    r = 4
    for i, t in enumerate(staff_detail, 1):
        ws2.cell(row=r, column=1, value=i).border = border
        ws2.cell(row=r, column=2, value=t["name"]).border = border
        ws2.cell(row=r, column=3, value=t.get("category", "Staff")).border = border

        status_cell = ws2.cell(row=r, column=4, value=t["status"])
        status_cell.border = border
        if "FULLY VERIFIED" in t["status"]:
            status_cell.fill = green_fill
        elif "TRUEFACE ONLY" in t["status"]:
            status_cell.fill = yellow_fill
        elif "ENTRY ONLY" in t["status"]:
            status_cell.fill = red_fill
        else:
            status_cell.fill = gray_fill

        occ = t.get("occupancy_status", "ABSENT")
        occ_cell = ws2.cell(row=r, column=5, value=occ)
        occ_cell.border = border
        if occ == "INSIDE":
            occ_cell.fill = green_fill
        elif occ == "EXITED":
            occ_cell.fill = light_blue_fill
        else:
            occ_cell.fill = gray_fill

        ws2.cell(row=r, column=6,
                 value=t.get("trueface_arrival") or t.get("dvr_first_seen") or "-").border = border
        ws2.cell(row=r, column=7,
                 value=t.get("trueface_departure") or "-").border = border
        ws2.cell(row=r, column=8,
                 value=", ".join(t.get("dvr_cameras", [])) or "-").border = border

        outfit_cell = ws2.cell(row=r, column=9, value=t.get("outfit_summary", "-"))
        outfit_cell.border = border
        outfit_cell.alignment = Alignment(wrap_text=True)
        r += 1

    for col_letter, w in [("A", 5), ("B", 25), ("C", 18), ("D", 22),
                           ("E", 12), ("F", 14), ("G", 14), ("H", 35), ("I", 30)]:
        ws2.column_dimensions[col_letter].width = w

    # ── Sheet 3: Sighting Timeline ──
    ws3 = wb.create_sheet("Sighting Timeline")

    ws3.merge_cells("A1:H1")
    ws3["A1"] = f"════════ DVR SIGHTING TIMELINE — {date_str} ════════"
    ws3["A1"].font = Font(bold=True, size=14)

    timeline_headers = ["#", "Name", "Category", "Time", "Camera Location",
                        "Confidence", "Outfit Color", "Outfit Detail"]
    for col, h in enumerate(timeline_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Collect all sightings and sort chronologically
    all_sightings = []
    for t in staff_detail:
        for det in t.get("dvr_sighting_details", []):
            all_sightings.append({"name": t["name"], "category": t.get("category", "Staff"), **det})
    all_sightings.sort(key=lambda x: x.get("time", ""))

    r = 4
    for idx, s in enumerate(all_sightings, 1):
        ws3.cell(row=r, column=1, value=idx).border = border
        ws3.cell(row=r, column=2, value=s["name"]).border = border
        ws3.cell(row=r, column=3, value=s["category"]).border = border
        ws3.cell(row=r, column=4, value=s.get("time", "")).border = border
        ws3.cell(row=r, column=5, value=s.get("camera", "")).border = border
        conf = s.get("confidence", 0)
        ws3.cell(row=r, column=6, value=f"{conf:.1%}" if conf else "—").border = border
        ws3.cell(row=r, column=7, value=s.get("outfit_color", "—")).border = border
        colors = s.get("outfit_colors", [])
        if colors:
            color_str = ", ".join(f"{c['color']} ({c['percentage']}%)" for c in colors)
        else:
            color_str = s.get("outfit_description", "—")
        detail_cell = ws3.cell(row=r, column=8, value=color_str)
        detail_cell.border = border
        detail_cell.alignment = Alignment(wrap_text=True)
        r += 1

    for col_letter, w in [("A", 5), ("B", 25), ("C", 10), ("D", 12),
                           ("E", 40), ("F", 12), ("G", 15), ("H", 40)]:
        ws3.column_dimensions[col_letter].width = w

    # ── Sheet 4: Reconciliation (Side-by-Side) ──
    ws4 = wb.create_sheet("Reconciliation")

    def _write_section(ws, start_row, title, teachers, fill):
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=8)
        title_cell = ws.cell(row=start_row, column=1,
                             value=f"{title} ({len(teachers)})")
        title_cell.font = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill = header_fill
        title_cell.border = border

        row = start_row + 1
        sub_headers = ["#", "Name", "TrueFace Arrival", "TrueFace Departure",
                       "DVR First Seen", "DVR Last Seen",
                       "Outfit / Clothing", "Sighting Count"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.border = border

        for idx, t in enumerate(teachers, 1):
            row += 1
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=t["name"]).border = border
            ws.cell(row=row, column=3,
                    value=t.get("trueface_arrival") or "—").border = border
            ws.cell(row=row, column=4,
                    value=t.get("trueface_departure") or "—").border = border
            ws.cell(row=row, column=5,
                    value=t.get("dvr_first_seen") or "—").border = border
            ws.cell(row=row, column=6,
                    value=t.get("dvr_last_seen") or "—").border = border
            ws.cell(row=row, column=7,
                    value=t.get("outfit_summary", "—")).border = border
            ws.cell(row=row, column=8,
                    value=t.get("dvr_sighting_count", 0)).border = border

        return row + 2

    ws4.merge_cells("A1:H1")
    ws4["A1"] = f"════════ RECONCILIATION — {date_str} ════════"
    ws4["A1"].font = Font(bold=True, size=14)
    row = 3

    row = _write_section(ws4, row, "✓ FULLY VERIFIED (Both TrueFace + DVR)",
                         fully_verified, green_fill)
    row = _write_section(ws4, row, "⚠ ENTRY ONLY — DVR but NOT on TrueFace",
                         dvr_only, red_fill)
    row = _write_section(ws4, row, "⚠ TRUEFACE ONLY — No DVR Sighting",
                         trueface_only, yellow_fill)
    row = _write_section(ws4, row, "— ABSENT (Neither System)",
                         absent, gray_fill)

    for col_letter, w in [("A", 5), ("B", 25), ("C", 16), ("D", 16),
                           ("E", 14), ("F", 14), ("G", 25), ("H", 14)]:
        ws4.column_dimensions[col_letter].width = w

    # ── Sheet 5: Unrecognized Persons ──
    ws5 = wb.create_sheet("Unrecognized Persons")

    ws5.merge_cells("A1:F1")
    ws5["A1"] = f"UNRECOGNIZED PERSONS - {date_str}"
    ws5["A1"].font = Font(bold=True, size=14)

    ws5["A3"] = "Total Unrecognized Persons"
    ws5["A3"].font = Font(bold=True)
    ws5["A3"].border = border
    ws5["B3"] = recon.get("total_unrecognized", 0)
    ws5["B3"].font = Font(bold=True, size=14)
    ws5["B3"].border = border
    ws5["B3"].fill = orange_fill

    ws5["A4"] = "Formula"
    ws5["A4"].font = Font(bold=True)
    ws5["A4"].border = border
    total_entries = recon.get("total_entries", 0)
    total_recognized = recon.get("total_recognized", 0)
    ws5["B4"] = f"Total Entries ({total_entries}) - Total Recognized ({total_recognized})"
    ws5["B4"].font = Font(bold=True, size=10)
    ws5["B4"].border = border

    # Unrecognized person detail table
    r = 7
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    title_cell = ws5.cell(row=r, column=1, value=f"Unrecognized Person Log ({len(unknown_persons)} entries)")
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.border = border

    r += 1
    for col, h in enumerate(["ID", "Time", "Entry Point", "Appearance", "Direction", "Status"], 1):
        cell = ws5.cell(row=r, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill
        cell.border = border

    for u in unknown_persons[:50]:
        r += 1
        ws5.cell(row=r, column=1, value=u["id"]).border = border
        ws5.cell(row=r, column=2, value=u.get("time", "-")).border = border
        ws5.cell(row=r, column=3, value=u.get("camera", "-")).border = border
        ws5.cell(row=r, column=4, value=u.get("attire_color", u.get("outfit_description", "-"))).border = border
        ws5.cell(row=r, column=5, value=u.get("direction", "IN")).border = border
        status_cell = ws5.cell(row=r, column=6, value="Unrecognized")
        status_cell.border = border
        status_cell.fill = orange_fill

    # Camera breakdown
    r += 2
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cam_title = ws5.cell(row=r, column=1, value="Detections by Camera")
    cam_title.font = Font(bold=True, size=12, color="FFFFFF")
    cam_title.fill = header_fill
    cam_title.border = border

    r += 1
    for col, h in enumerate(["Camera", "Count"], 1):
        cell = ws5.cell(row=r, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill
        cell.border = border

    for cam_name in sorted(visitor_by_camera.keys()):
        cam_visitors = visitor_by_camera[cam_name]
        r += 1
        ws5.cell(row=r, column=1, value=cam_name).border = border
        ws5.cell(row=r, column=2, value=len(cam_visitors)).border = border

    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 15
    ws5.column_dimensions["C"].width = 40
    ws5.column_dimensions["D"].width = 12
    ws5.column_dimensions["E"].width = 12

    # ── Sheet 6: Outfit Reconciliation ──
    ws6 = wb.create_sheet("Outfit Reconciliation")

    ws6.merge_cells("A1:E1")
    ws6["A1"] = f"════════ OUTFIT RECONCILIATION — {date_str} ════════"
    ws6["A1"].font = Font(bold=True, size=14)

    outfit_headers = ["#", "Name", "Outfit / Clothing", "Total Sightings",
                      "Cameras (with outfit)"]
    for col, h in enumerate(outfit_headers, 1):
        cell = ws6.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    r = 4
    idx = 0
    for t in staff_detail:
        outfit_obs = t.get("outfit_observations", [])
        if not outfit_obs:
            continue
        idx += 1
        outfit_desc = t.get("outfit_summary", "-")
        cam_set = set()
        for obs in outfit_obs:
            if obs.get("camera"):
                cam_set.add(obs["camera"])
        ws6.cell(row=r, column=1, value=idx).border = border
        ws6.cell(row=r, column=2, value=t["name"]).border = border
        outfit_cell = ws6.cell(row=r, column=3, value=outfit_desc)
        outfit_cell.border = border
        outfit_cell.alignment = Alignment(wrap_text=True)
        ws6.cell(row=r, column=4, value=len(outfit_obs)).border = border
        ws6.cell(row=r, column=5,
                 value=", ".join(sorted(cam_set)) or "—").border = border
        r += 1

    for col_letter, w in [("A", 5), ("B", 25), ("C", 35), ("D", 15), ("E", 45)]:
        ws6.column_dimensions[col_letter].width = w

    # ── Sheet 7: AI Conclusion ──
    ws7 = wb.create_sheet("AI Conclusion")

    ws7.merge_cells("A1:B1")
    ws7["A1"] = f"════════ AI CONCLUSION — {date_str} ════════"
    ws7["A1"].font = Font(bold=True, size=14)

    observations = _generate_ai_observations(recon)
    r = 3
    for obs in observations:
        ws7.cell(row=r, column=1, value="•").border = border
        obs_cell = ws7.cell(row=r, column=2, value=obs)
        obs_cell.border = border
        obs_cell.alignment = Alignment(wrap_text=True)
        r += 1

    ws7.column_dimensions["A"].width = 5
    ws7.column_dimensions["B"].width = 80

    # ── Sheet 8: Time Trail ──
    ws8 = wb.create_sheet("Time Trail")

    ws8.merge_cells("A1:E1")
    ws8["A1"] = f"════════ TIME TRAIL RECONSTRUCTION — {date_str} ════════"
    ws8["A1"].font = Font(bold=True, size=14)

    trail_headers = ["#", "Name", "Category", "Time", "Source / Camera", "Event"]
    for col, h in enumerate(trail_headers, 1):
        cell = ws8.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    r = 4
    # Show time trail for detected persons only
    detected = [t for t in staff_detail if t.get("time_trail")]
    for t in detected:
        # Person header row
        ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        person_cell = ws8.cell(row=r, column=1,
                               value=f"{t['name']} - {t.get('category', 'Staff')}")
        person_cell.font = Font(bold=True, size=10)
        person_cell.fill = light_blue_fill
        person_cell.border = border
        r += 1
        for idx, trail in enumerate(t["time_trail"], 1):
            ws8.cell(row=r, column=1, value=idx).border = border
            ws8.cell(row=r, column=2, value=t["name"]).border = border
            ws8.cell(row=r, column=3, value=t.get("category", "Staff")).border = border
            ws8.cell(row=r, column=4, value=trail.get("time", "")).border = border
            ws8.cell(row=r, column=5, value=trail.get("source", "")).border = border
            ws8.cell(row=r, column=6, value=trail.get("event", "")).border = border
            r += 1
        r += 1  # blank row between persons

    for col_letter, w in [("A", 5), ("B", 25), ("C", 20), ("D", 12),
                           ("E", 40), ("F", 30)]:
        ws8.column_dimensions[col_letter].width = w

    # ── Sheet 9: Mismatch Alerts ──
    alerts = recon.get("alerts", [])
    ws9 = wb.create_sheet("Mismatch Alerts")

    ws9.merge_cells("A1:E1")
    ws9["A1"] = f"════════ MISMATCH ALERTS — {date_str} ════════"
    ws9["A1"].font = Font(bold=True, size=14)

    alert_headers = ["#", "Alert Type", "Severity", "Person", "Detail"]
    for col, h in enumerate(alert_headers, 1):
        cell = ws9.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    high_fill = PatternFill("solid", fgColor="FF4444")
    medium_fill = PatternFill("solid", fgColor="FFA500")
    low_fill = PatternFill("solid", fgColor="FFEB9C")
    severity_fills = {"high": high_fill, "medium": medium_fill, "low": low_fill}

    r = 4
    for idx, alert in enumerate(alerts, 1):
        ws9.cell(row=r, column=1, value=idx).border = border
        ws9.cell(row=r, column=2, value=alert.get("type", "")).border = border
        sev = alert.get("severity", "medium")
        sev_cell = ws9.cell(row=r, column=3, value=sev.upper())
        sev_cell.border = border
        sev_cell.fill = severity_fills.get(sev, medium_fill)
        ws9.cell(row=r, column=4, value=alert.get("person", "—") or "—").border = border
        detail_cell = ws9.cell(row=r, column=5, value=alert.get("detail", ""))
        detail_cell.border = border
        detail_cell.alignment = Alignment(wrap_text=True)
        r += 1

    if not alerts:
        ws9.cell(row=4, column=1, value="No mismatch alerts for this period.").font = Font(italic=True)

    for col_letter, w in [("A", 5), ("B", 30), ("C", 12), ("D", 25), ("E", 60)]:
        ws9.column_dimensions[col_letter].width = w

    # ── Sheet 10: Gate vs Face Discrepancy ──
    disc = recon.get("discrepancy", {})
    ws10d = wb.create_sheet("Discrepancy")

    ws10d.merge_cells("A1:F1")
    ws10d["A1"] = f"════════ GATE vs FACE DISCREPANCY — {date_str} ════════"
    ws10d["A1"].font = Font(bold=True, size=14)

    r = 3
    disc_summary = [
        ("Gate Head Count (IN)", disc.get("gate_head_count", 0), None),
        ("Faces Identified (Total)", disc.get("faces_identified", 0), None),
        ("  Staff Identified", disc.get("staff_identified", 0), light_blue_fill),
        ("  Visitors Identified", disc.get("visitors_identified", 0), orange_fill),
        ("UNRECONCILED", disc.get("unreconciled", 0),
         red_fill if disc.get("unreconciled", 0) > 0 else green_fill),
        ("Reconciliation Rate", f"{disc.get('reconciliation_rate', 100)}%",
         green_fill if disc.get("reconciliation_rate", 100) >= 80 else red_fill),
    ]
    for label, value, fill in disc_summary:
        ws10d.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws10d.cell(row=r, column=1).border = border
        ws10d.cell(row=r, column=2, value=value).font = Font(bold=True, size=12)
        ws10d.cell(row=r, column=2).border = border
        if fill:
            ws10d.cell(row=r, column=1).fill = fill
            ws10d.cell(row=r, column=2).fill = fill
        r += 1

    # Hourly breakdown table
    hourly = disc.get("hourly_breakdown", [])
    if hourly:
        r += 1
        ws10d.cell(row=r, column=1, value="HOURLY BREAKDOWN").font = Font(bold=True, size=12)
        r += 1
        for col, h in enumerate(["Hour", "Gate Entries", "Identified", "Unreconciled"], 1):
            cell = ws10d.cell(row=r, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        r += 1
        for hd in hourly:
            ws10d.cell(row=r, column=1, value=hd["hour"]).border = border
            ws10d.cell(row=r, column=2, value=hd["gate_entries"]).border = border
            ws10d.cell(row=r, column=3, value=hd["identified"]).border = border
            unrec_cell = ws10d.cell(row=r, column=4, value=hd["unreconciled"])
            unrec_cell.border = border
            if hd["unreconciled"] > 0:
                unrec_cell.fill = red_fill
            r += 1

    # Unreconciled entries detail (CATEGORY 3)
    unrec_list = disc.get("unreconciled_entries", [])
    if unrec_list:
        r += 1
        ws10d.cell(row=r, column=1,
                    value="CATEGORY 3: UNRECONCILED GATE ENTRIES").font = Font(bold=True, size=12)
        r += 1
        ws10d.cell(row=r, column=1,
                    value="(Detected at gate but NOT matched to any face — guards, housekeeping, vendors, etc.)").font = Font(italic=True)
        r += 1
        for col, h in enumerate(["ID", "Time", "Camera", "Attire Color", "Direction"], 1):
            cell = ws10d.cell(row=r, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        r += 1
        for ue in unrec_list:
            ws10d.cell(row=r, column=1, value=ue.get("id", "")).border = border
            ws10d.cell(row=r, column=2, value=ue.get("time", "")).border = border
            ws10d.cell(row=r, column=3, value=ue.get("camera", "")).border = border
            attire_cell = ws10d.cell(row=r, column=4, value=ue.get("attire_color", ""))
            attire_cell.border = border
            attire_cell.fill = yellow_fill
            ws10d.cell(row=r, column=5, value=ue.get("direction", "IN")).border = border
            r += 1

    # Reconciled entries (cross-referenced successfully)
    rec_list = disc.get("reconciled_entries", [])
    if rec_list:
        r += 2
        ws10d.cell(row=r, column=1,
                    value="RECONCILED GATE ENTRIES (matched to face detection)").font = Font(bold=True, size=12)
        r += 1
        for col, h in enumerate(["Gate Time", "Camera", "Attire", "Matched To", "Type", "Time Diff (s)"], 1):
            cell = ws10d.cell(row=r, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        r += 1
        for re_entry in rec_list:
            ws10d.cell(row=r, column=1, value=re_entry.get("gate_timestamp", "")).border = border
            ws10d.cell(row=r, column=2, value=re_entry.get("camera", "")).border = border
            ws10d.cell(row=r, column=3, value=re_entry.get("attire_color", "")).border = border
            ws10d.cell(row=r, column=4, value=re_entry.get("matched_to", "")).border = border
            ws10d.cell(row=r, column=5, value=re_entry.get("matched_type", "")).border = border
            ws10d.cell(row=r, column=6, value=re_entry.get("time_diff_seconds", "")).border = border
            r += 1

    for col_letter, w in [("A", 40), ("B", 18), ("C", 20), ("D", 20), ("E", 15), ("F", 15)]:
        ws10d.column_dimensions[col_letter].width = w

    # ── Sheet 11: Snapshots — DISABLED per user request ──

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_ai_observations(recon: dict) -> list[str]:
    """Generate AI CONCLUSION for the reconciliation report (per school spec).

    Provides:
    1. Total people who entered.
    2. Total people who exited.
    3. Number of students recognized.
    4. Number of staff recognized.
    5. Number of unrecognized persons.
    6. Whether there are significant discrepancies requiring investigation.
    """
    observations = []
    side_by_side = recon.get("side_by_side", {})
    staff_detail = recon.get("staff_detail", recon.get("teacher_detail", []))

    dvr_only = side_by_side.get("dvr_only", [])
    trueface_only = side_by_side.get("trueface_only", [])
    fully_verified = side_by_side.get("both_present", [])

    total_entries = recon.get("total_entries", recon.get("unique_gate_in", 0))
    total_exits = recon.get("total_exits", recon.get("unique_gate_out", 0))
    recognized_students = recon.get("recognized_students", 0)
    recognized_staff_count = recon.get("recognized_staff_count", 0)
    total_recognized = recon.get("total_recognized", 0)
    total_unrecognized = recon.get("total_unrecognized", 0)
    current_occupancy = recon.get("current_occupancy", recon.get("total_inside", 0))

    # 1. Total entries and exits
    observations.append(
        f"Total persons entered: {total_entries}. "
        f"Total persons exited: {total_exits}. "
        f"Current occupancy: {current_occupancy}."
    )

    # 2-4. Recognition breakdown
    observations.append(
        f"Recognized: {total_recognized} "
        f"(Students: {recognized_students}, Staff: {recognized_staff_count}). "
        f"Unrecognized: {total_unrecognized}."
    )

    # 5. Discrepancy analysis
    if total_unrecognized > 0:
        observations.append(
            f"{total_unrecognized} person(s) entered but could not be recognized "
            f"by the Student or Staff facial recognition database. "
            f"These individuals require investigation."
        )

    if dvr_only:
        observations.append(
            f"DISCREPANCY: {len(dvr_only)} staff seen on DVR cameras but NOT marked "
            f"on TrueFace 3000. They may have bypassed biometric attendance."
        )
    if trueface_only:
        observations.append(
            f"NOTE: {len(trueface_only)} staff marked on TrueFace but NOT spotted "
            f"on any DVR camera."
        )

    # Staff detection rate
    total_reg = recon.get("total_registered", 0)
    if total_reg > 0:
        present = len(fully_verified) + len(trueface_only) + len(dvr_only)
        rate = round(present / total_reg * 100, 1)
        observations.append(
            f"Staff detection rate: {rate}% ({present}/{total_reg} registered staff detected)."
        )

    if not observations:
        observations.append("No significant observations for this period.")

    return observations


# ============================================================
# PDF Snapshot Gallery
# ============================================================

def _add_snapshot_gallery(pdf, recon: dict, section_header):
    """Add live snapshot images to the PDF report."""
    import base64
    import tempfile

    vehicle_entries = recon.get("vehicle_entries", [])
    gate_entries_raw = recon.get("gate_entries_with_crops", [])

    # Collect vehicle snapshots (latest 12)
    vehicle_snaps = [v for v in vehicle_entries if v.get("snapshot")]
    # Collect person crops from gate entries (latest 12)
    person_snaps = [g for g in gate_entries_raw if g.get("person_crop")]

    if not vehicle_snaps and not person_snaps:
        return

    def _embed_b64_image(b64_data: str, img_w: int = 45, img_h: int = 35):
        """Decode base64 image and embed in PDF."""
        try:
            img_bytes = base64.b64decode(b64_data)
            with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                tmp.write(img_bytes)
                tmp_path = tmp.name
            pdf.image(tmp_path, w=img_w, h=img_h)
            import os
            os.unlink(tmp_path)
            return True
        except Exception:
            return False

    # ── Vehicle Snapshots ──
    if vehicle_snaps:
        section_header("VEHICLE SNAPSHOTS (Live DVR Captures)")
        pdf.set_font("Helvetica", "", 8)
        cols = 3
        img_w = 58
        img_h = 40
        for i, v in enumerate(vehicle_snaps[-12:]):
            col = i % cols
            if col == 0 and i > 0:
                pdf.ln(img_h + 12)
            x = pdf.l_margin + col * (img_w + 5)
            y = pdf.get_y()
            if y + img_h + 15 > pdf.h - pdf.b_margin:
                pdf.add_page()
                y = pdf.get_y()
            try:
                img_bytes = base64.b64decode(v["snapshot"])
                with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                    tmp.write(img_bytes)
                    tmp_path = tmp.name
                pdf.image(tmp_path, x=x, y=y, w=img_w, h=img_h)
                import os
                os.unlink(tmp_path)
                # Caption below image
                pdf.set_xy(x, y + img_h)
                ts = v.get("timestamp", "")
                time_part = ts.split(" ")[1] if " " in ts else ts
                vtype = v.get("vehicle_type", "vehicle").upper()
                direction = v.get("direction", "")
                caption = f"{vtype} {direction} - {time_part}"
                pdf.cell(img_w, 4, caption, align="C")
            except Exception:
                pass
        pdf.ln(img_h + 15)

    # ── Gate Entry Person Snapshots ──
    if person_snaps:
        section_header("GATE ENTRY SNAPSHOTS (Live DVR Captures)")
        pdf.set_font("Helvetica", "", 8)
        cols = 4
        img_w = 42
        img_h = 50
        for i, g in enumerate(person_snaps[-16:]):
            col = i % cols
            if col == 0 and i > 0:
                pdf.ln(img_h + 12)
            x = pdf.l_margin + col * (img_w + 5)
            y = pdf.get_y()
            if y + img_h + 15 > pdf.h - pdf.b_margin:
                pdf.add_page()
                y = pdf.get_y()
            try:
                img_bytes = base64.b64decode(g["person_crop"])
                with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                    tmp.write(img_bytes)
                    tmp_path = tmp.name
                pdf.image(tmp_path, x=x, y=y, w=img_w, h=img_h)
                import os
                os.unlink(tmp_path)
                # Caption
                pdf.set_xy(x, y + img_h)
                ts = g.get("timestamp", "")
                time_part = ts.split(" ")[1] if " " in ts else ts
                cam = g.get("camera", "")
                direction = g.get("direction", "IN")
                caption = f"{direction} {time_part}"
                pdf.cell(img_w, 4, caption, align="C")
            except Exception:
                pass
        pdf.ln(img_h + 15)


# ============================================================
# PDF Report Generation
# ============================================================

def _generate_reconciliation_pdf(recon: dict, date_display: str,
                                  time_display: str) -> bytes:
    """Generate HEAD COUNT RECONCILIATION REPORT matching the school spec.

    Sections:
      1. ENTRY SUMMARY
      2. RECOGNITION SUMMARY
      3. UNRECOGNIZED SUMMARY
      4. EXIT SUMMARY
      5. CURRENT OCCUPANCY
      6. RECONCILIATION CHECK
      7. UNRECOGNIZED PERSON DETAILS
      8. AI CONCLUSION
    """
    from fpdf import FPDF

    staff_detail = recon.get("staff_detail", recon.get("teacher_detail", []))
    side_by_side = recon.get("side_by_side", {})
    unknown_persons = recon.get("unknown_persons", [])
    unrec_entries = recon.get("unreconciled_gate_entries", [])

    staff_inside = [s for s in staff_detail if s.get("occupancy_status") == "INSIDE"]
    staff_exited = [s for s in staff_detail if s.get("occupancy_status") == "EXITED"]
    staff_absent = [s for s in staff_detail if s.get("occupancy_status") == "ABSENT"]
    staff_present = [s for s in staff_detail if s.get("occupancy_status") in ("INSIDE", "EXITED")]

    fully_verified = side_by_side.get("both_present", [])
    trueface_only = side_by_side.get("trueface_only", [])
    dvr_only = side_by_side.get("dvr_only", [])

    total_entries = recon.get("total_entries", recon.get("unique_gate_in", 0))
    total_exits = recon.get("total_exits", recon.get("unique_gate_out", 0))
    recognized_students = recon.get("recognized_students", 0)
    recognized_staff_count = recon.get("recognized_staff_count", 0)
    total_recognized = recon.get("total_recognized", len(staff_present))
    total_unrecognized = recon.get("total_unrecognized", max(0, total_entries - total_recognized))
    current_occupancy = recon.get("current_occupancy", max(0, total_entries - total_exits))

    # Combine all unrecognized person entries (face-detected unknowns + gate-only)
    all_unrecognized: list[dict] = []
    for u in unknown_persons:
        all_unrecognized.append(u)
    for ue in unrec_entries:
        all_unrecognized.append(ue)
    # Re-number with U-001, U-002, ...
    for idx, u in enumerate(all_unrecognized, 1):
        u["id"] = f"U-{idx:03d}"

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    _orig_cell = pdf.cell

    def _safe_cell(*args, **kwargs):
        new_args = list(args)
        for i, a in enumerate(new_args):
            if isinstance(a, str):
                new_args[i] = a.replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"')
        for k, v in kwargs.items():
            if isinstance(v, str):
                kwargs[k] = v.replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"')
        return _orig_cell(*new_args, **kwargs)

    pdf.cell = _safe_cell

    # ── Title ──
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "HEAD COUNT RECONCILIATION REPORT", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, f"PP International School  |  Date: {date_display}  |  {time_display} IST",
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(6)

    # ── Helper functions ──
    def section_header(title: str):
        pdf.set_fill_color(47, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 9, f"  {title}", new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    def key_value(key: str, value, indent: int = 0, bold_val: bool = False):
        pdf.set_font("Helvetica", "", 10)
        prefix = "    " * indent
        pdf.cell(90 + indent * 10, 6, f"{prefix}{key}", new_x="RIGHT")
        pdf.set_font("Helvetica", "B" if bold_val else "", 10)
        pdf.cell(0, 6, str(value), new_x="LMARGIN", new_y="NEXT")

    def status_row(label: str, count, r: int, g: int, b: int):
        pdf.set_fill_color(r, g, b)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(120, 7, f"  {label}", border=1, fill=True, new_x="RIGHT")
        pdf.cell(30, 7, str(count), border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")

    # ══════════════════════════════════════════════
    # SECTION 1: ENTRY SUMMARY
    # ══════════════════════════════════════════════
    section_header("ENTRY SUMMARY")
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(0, 100, 0)
    pdf.cell(0, 10, f"Total Persons Entered: {total_entries}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(1)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5,
             f"Counted via: Main Gate cameras, Basement Entry, Dispersal Exit  |  "
             f"Raw detections: IN={recon.get('raw_gate_in', 0)}, OUT={recon.get('raw_gate_out', 0)}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 2: RECOGNITION SUMMARY
    # ══════════════════════════════════════════════
    section_header("RECOGNITION SUMMARY")
    key_value("Recognized Students", recognized_students, bold_val=True)
    key_value("Recognized Staff", recognized_staff_count, bold_val=True)
    status_row("Total Recognized", total_recognized, 198, 239, 206)
    pdf.ln(2)

    # Staff breakdown by category
    staff_cat_present = recon.get("staff_category_present", {})
    if staff_cat_present:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 6, "  Recognized by Category:", new_x="LMARGIN", new_y="NEXT")
        for cat_name in sorted(staff_cat_present.keys()):
            key_value(cat_name, staff_cat_present[cat_name], indent=1)
    pdf.ln(2)

    # Recognized staff inside table
    if staff_present:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(198, 239, 206)
        pdf.cell(8, 7, "#", border=1, fill=True, align="C", new_x="RIGHT")
        pdf.cell(55, 7, "Name", border=1, fill=True, new_x="RIGHT")
        pdf.cell(30, 7, "Category", border=1, fill=True, new_x="RIGHT")
        pdf.cell(25, 7, "Arrived", border=1, fill=True, align="C", new_x="RIGHT")
        pdf.cell(25, 7, "Status", border=1, fill=True, align="C", new_x="RIGHT")
        pdf.cell(0, 7, "Cameras", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        for idx, s in enumerate(staff_present, 1):
            arrival = s.get("trueface_arrival") or s.get("dvr_first_seen") or "-"
            cams = ", ".join(s.get("dvr_cameras", [])[:3]) or "-"
            occ = s.get("occupancy_status", "INSIDE")
            pdf.cell(8, 6, str(idx), border=1, align="C", new_x="RIGHT")
            pdf.cell(55, 6, s["name"], border=1, new_x="RIGHT")
            pdf.cell(30, 6, s["category"], border=1, new_x="RIGHT")
            pdf.cell(25, 6, arrival, border=1, align="C", new_x="RIGHT")
            pdf.cell(25, 6, occ, border=1, align="C", new_x="RIGHT")
            pdf.cell(0, 6, cams, border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 3: UNRECOGNIZED SUMMARY
    # ══════════════════════════════════════════════
    section_header("UNRECOGNIZED SUMMARY")
    status_row("Total Unrecognized Persons", total_unrecognized, 255, 199, 206)
    pdf.ln(1)
    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(0, 5,
             f"Formula: Total Entries ({total_entries}) - Total Recognized ({total_recognized}) "
             f"= {total_unrecognized}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5,
             "These persons were detected entering but NOT recognized as Student or Staff.",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 4: EXIT SUMMARY
    # ══════════════════════════════════════════════
    section_header("EXIT SUMMARY")
    key_value("Total Persons Exited", total_exits, bold_val=True)
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 5: CURRENT OCCUPANCY
    # ══════════════════════════════════════════════
    section_header("CURRENT OCCUPANCY")
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(0, 100, 0)
    pdf.cell(0, 12, f"Current Occupancy: {current_occupancy}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(1)
    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(0, 5,
             f"Formula: Entries ({total_entries}) - Exits ({total_exits}) = {current_occupancy}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 6: RECONCILIATION CHECK
    # ══════════════════════════════════════════════
    section_header("RECONCILIATION CHECK")
    key_value("Total Entries", total_entries, bold_val=True)
    key_value("Total Recognized", total_recognized, bold_val=True)
    key_value("Difference (Unrecognized)", total_unrecognized, bold_val=True)
    pdf.ln(2)
    pdf.set_font("Helvetica", "I", 9)
    pdf.multi_cell(0, 5,
                   "The difference represents persons detected entering but not recognized "
                   "by the Student or Staff facial recognition database.",
                   new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Reconciliation rate
    recon_rate = round((total_recognized / total_entries * 100) if total_entries > 0 else 100, 1)
    if recon_rate >= 80:
        status_row(f"Reconciliation Rate: {recon_rate}%", "OK", 198, 239, 206)
    else:
        status_row(f"Reconciliation Rate: {recon_rate}%", "LOW", 255, 199, 206)
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 7: UNRECOGNIZED PERSON DETAILS
    # ══════════════════════════════════════════════
    section_header("UNRECOGNIZED PERSON DETAILS")

    if all_unrecognized:
        for u in all_unrecognized[:50]:
            uid = u.get("id", "U-???")
            entry_time = u.get("time", u.get("timestamp", "-"))
            camera = u.get("camera", "Unknown")
            attire = u.get("attire_color", "")
            outfit = u.get("outfit_description", "")
            appearance = outfit if outfit else (attire if attire and attire != "unknown" else "-")
            direction = u.get("direction", "IN")
            status = "Unrecognized"

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_fill_color(255, 235, 156)
            pdf.cell(0, 7, f"  {uid}", fill=True, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 5, f"    Entry Time: {entry_time}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 5, f"    Entry Point: {camera}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 5, f"    Appearance: {appearance}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 5, f"    Direction: {direction}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(0, 5, f"    Status: {status}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

        if len(all_unrecognized) > 50:
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 5, f"... and {len(all_unrecognized) - 50} more (see Excel for full list)",
                     new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 6, "  No unrecognized persons detected.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ══════════════════════════════════════════════
    # SECTION 8: VEHICLE COUNT
    # ══════════════════════════════════════════════
    vehicles_in = recon.get("vehicles_in", 0)
    vehicles_out = recon.get("vehicles_out", 0)
    vehicle_types = recon.get("vehicle_types", {})
    if vehicles_in > 0 or vehicles_out > 0:
        section_header("VEHICLE COUNT (Separate from Head Count)")
        key_value("Vehicles IN", vehicles_in, bold_val=True)
        key_value("Vehicles OUT", vehicles_out)
        for vtype, vcount in sorted(vehicle_types.items()):
            key_value(vtype.title(), vcount, indent=1)
        pdf.ln(2)

    # ══════════════════════════════════════════════
    # SECTION 9: STAFF RECONCILIATION DETAIL
    # ══════════════════════════════════════════════
    section_header("STAFF RECONCILIATION DETAIL")
    status_row("Fully Verified (TrueFace + DVR)", len(fully_verified), 198, 239, 206)
    status_row("Entry Only (DVR - Not on TrueFace)", len(dvr_only), 255, 199, 206)
    status_row("TrueFace Only (No DVR Sighting)", len(trueface_only), 255, 235, 156)
    status_row("Not Detected / Absent", len(staff_absent), 217, 217, 217)
    pdf.ln(3)

    # Entry without attendance
    if dvr_only:
        for t in dvr_only:
            cams = ", ".join(t.get("dvr_cameras", []))
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(0, 6, f"{t['name']} - {t.get('category', 'Staff')}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 5, f"    DVR: {t.get('dvr_first_seen', '-')}  |  Cameras: {cams}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(200, 0, 0)
            pdf.cell(0, 5, "    NOT marked on TrueFace 3000", new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(2)

    # ── Mismatch Alerts ──
    alerts = recon.get("alerts", [])
    if alerts:
        section_header("ALERTS")
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(217, 226, 243)
        pdf.cell(8, 7, "#", border=1, fill=True, align="C", new_x="RIGHT")
        pdf.cell(45, 7, "Alert Type", border=1, fill=True, new_x="RIGHT")
        pdf.cell(18, 7, "Severity", border=1, fill=True, align="C", new_x="RIGHT")
        pdf.cell(0, 7, "Detail", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        for idx, alert in enumerate(alerts, 1):
            sev = alert.get("severity", "medium")
            if sev == "high":
                pdf.set_fill_color(255, 68, 68)
            elif sev == "medium":
                pdf.set_fill_color(255, 165, 0)
            else:
                pdf.set_fill_color(255, 235, 156)
            pdf.cell(8, 6, str(idx), border=1, align="C", new_x="RIGHT")
            pdf.cell(45, 6, alert.get("type", ""), border=1, new_x="RIGHT")
            pdf.cell(18, 6, sev.upper(), border=1, align="C", fill=True, new_x="RIGHT")
            pdf.set_fill_color(255, 255, 255)
            pdf.cell(0, 6, alert.get("detail", ""), border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

    # ══════════════════════════════════════════════
    # SECTION 10: AI CONCLUSION
    # ══════════════════════════════════════════════
    ai_obs = _generate_ai_observations(recon)
    section_header("AI CONCLUSION")
    pdf.set_font("Helvetica", "", 10)
    for obs in ai_obs:
        pdf.multi_cell(0, 6, f"  {obs}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # ── Footer ──
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, "PPIS Head Count Reconciliation & Entry Monitoring System",
             new_x="LMARGIN", new_y="NEXT", align="C")

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


# ============================================================
# Hourly Reconciliation Report (7 AM - 5 PM IST)
# ============================================================


@router.post("/api/gate/send-report")
async def trigger_reconciliation_report():
    """Manual trigger to send the reconciliation report immediately."""
    await send_reconciliation_report()
    return {"status": "sent"}


@router.post("/api/gate/test-alert")
async def test_unknown_person_alert():
    """Send a test WhatsApp template alert to verify the unknown person alert pipeline."""
    test_entry = {
        "camera": "ENTRY GATE-1",
        "timestamp": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        "direction": "IN",
        "person_crop": "",
    }
    # Clear cooldown so the test alert always sends
    _unknown_alert_last.pop("ENTRY GATE-1", None)
    await _send_unknown_person_alert(test_entry)
    return {"status": "sent", "phones": UNKNOWN_ALERT_PHONES, "template": UNKNOWN_ALERT_TEMPLATE}


# ---------------------------------------------------------------------------
# Periodic Entry Gate Snapshot → WhatsApp
# ---------------------------------------------------------------------------
GATE_SNAPSHOT_PHONES = [
    # "918796105084",   # Alisha — disabled per request
]


@router.post("/api/gate/entry-gate-snapshot")
async def receive_entry_gate_snapshot(request: Request):
    """Receive an Entry Gate snapshot from campus agent and send to WhatsApp.

    Body: {"image_b64": "<base64 JPEG>", "camera": "ENTRY GATE-1", "timestamp": "..."}
    The campus agent calls this every 10 minutes with a fresh ISAPI capture.
    """
    body = await request.json()
    image_b64 = body.get("image_b64", "")
    camera = body.get("camera", "Entry Gate")
    ts = body.get("timestamp", datetime.now(IST).strftime("%d-%m-%Y %H:%M:%S IST"))

    if not image_b64:
        return {"status": "error", "detail": "no image_b64 provided"}

    from app.services.whatsapp_service import (
        upload_base64_image_cloud,
        send_cloud_media,
    )

    media_id = await upload_base64_image_cloud(image_b64)
    if not media_id:
        logger.error("[GATE] Failed to upload entry gate snapshot to WhatsApp")
        return {"status": "error", "detail": "media upload failed"}

    # NOTE: Periodic entry gate snapshots are DISABLED on the campus agent
    # side (2026-05-30). This endpoint is kept for backward compatibility
    # but should not receive traffic. Uses regular media message.
    caption = f"\U0001f4f7 {camera} — {ts}"
    sent_count = 0
    for phone in GATE_SNAPSHOT_PHONES:
        ok = await send_cloud_media(phone, "image", media_id=media_id, caption=caption)
        if ok:
            sent_count += 1
        else:
            logger.warning("[GATE] Failed to send gate snapshot to %s", phone)

    logger.info("[GATE] Entry gate snapshot sent to %d/%d recipients", sent_count, len(GATE_SNAPSHOT_PHONES))
    return {"status": "ok", "sent": sent_count, "total": len(GATE_SNAPSHOT_PHONES)}


async def send_reconciliation_report():
    """Generate and send the head count reconciliation report (runs every 30 min, 6 AM-5 PM IST)."""
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    today_display = now.strftime("%d-%m-%Y")
    time_display = now.strftime("%I:%M %p")

    db = await _get_db()
    try:
        recon = await _reconcile(db, today)
    finally:
        await db.close()

    side_by_side = recon.get("side_by_side", {})
    if recon["trueface_identified"] == 0 and recon.get("dvr_sighted", 0) == 0 and recon.get("raw_gate_in", 0) == 0:
        logger.info("[GATE] No TrueFace, DVR, or gate records for %s — skipping report", today)
        return

    # Generate Excel
    xlsx_bytes = _generate_reconciliation_excel(recon)
    filename = f"Head_Count_Reconciliation_{today}_{now.strftime('%H%M')}.xlsx"

    # Build counts from new architecture (per school spec)
    total_entries = recon.get("total_entries", recon.get("unique_gate_in", 0))
    total_exits = recon.get("total_exits", recon.get("unique_gate_out", 0))
    recognized_students = recon.get("recognized_students", 0)
    recognized_staff_count = recon.get("recognized_staff_count", 0)
    total_recognized = recon.get("total_recognized", 0)
    total_unrecognized = recon.get("total_unrecognized", 0)
    current_occupancy = recon.get("current_occupancy", recon.get("total_inside", 0))

    fully_verified = side_by_side.get("both_present", [])
    tf_only_list = side_by_side.get("trueface_only", [])
    dvr_only_list = side_by_side.get("dvr_only", [])
    absent_list = side_by_side.get("neither", [])

    # Generate PDF report
    pdf_bytes = _generate_reconciliation_pdf(recon, today_display, time_display)
    pdf_filename = f"Head_Count_Reconciliation_{today}_{now.strftime('%H%M')}.pdf"

    # Vehicle summary for body
    v_in = recon.get("vehicles_in", 0)
    v_out = recon.get("vehicles_out", 0)
    v_types = recon.get("vehicle_types", {})
    vehicle_line = ""
    if v_in > 0 or v_out > 0:
        type_parts = ", ".join(f"{vt.title()}: {vc}" for vt, vc in sorted(v_types.items()))
        vehicle_line = f"\nVehicles IN: {v_in} | OUT: {v_out}"
        if type_parts:
            vehicle_line += f" ({type_parts})"
        vehicle_line += "\n"

    # Brief email body matching school spec
    body = (
        f"Head Count Reconciliation Report - {today_display} at {time_display} IST\n\n"
        f"ENTRY SUMMARY\n"
        f"  Total Persons Entered: {total_entries}\n\n"
        f"RECOGNITION SUMMARY\n"
        f"  Recognized Students: {recognized_students}\n"
        f"  Recognized Staff: {recognized_staff_count}\n"
        f"  Total Recognized: {total_recognized}\n\n"
        f"UNRECOGNIZED SUMMARY\n"
        f"  Total Unrecognized Persons: {total_unrecognized}\n\n"
        f"EXIT SUMMARY\n"
        f"  Total Persons Exited: {total_exits}\n\n"
        f"CURRENT OCCUPANCY: {current_occupancy}\n"
        f"  (Entries {total_entries} - Exits {total_exits})\n\n"
        f"RECONCILIATION CHECK\n"
        f"  Total Entries: {total_entries}\n"
        f"  Total Recognized: {total_recognized}\n"
        f"  Difference: {total_unrecognized}\n"
        f"{vehicle_line}\n"
        f"See attached PDF for detailed report including unrecognized person details.\n\n"
        f"-- PPIS Head Count Reconciliation & Entry Monitoring System"
    )

    from app.services.email_service import send_email_async
    recipients = [r.strip() for r in REPORT_RECIPIENTS.split(",") if r.strip()]
    for email in recipients:
        ok = await send_email_async(
            email,
            f"Hourly Head Count Reconciliation — {today_display} {time_display} IST",
            body,
            "PP International School",
            attachments=[
                (pdf_filename, pdf_bytes),
                (filename, xlsx_bytes),
            ],
        )
        logger.info("[GATE] Reconciliation report → %s: %s", email, "OK" if ok else "FAILED")

    logger.info(
        "[GATE] Reconciliation report sent at %s: Entries=%d, Recognized=%d, Unrecognized=%d, Occupancy=%d",
        time_display, total_entries, total_recognized, total_unrecognized,
        current_occupancy,
    )


def send_reconciliation_report_sync():
    """Sync wrapper for scheduler."""
    import asyncio
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            asyncio.ensure_future(send_reconciliation_report())
        else:
            loop.run_until_complete(send_reconciliation_report())
    except RuntimeError:
        asyncio.run(send_reconciliation_report())


# ============================================================
# Live Dashboard — Real-Time Unknown Persons Monitor
# ============================================================

@router.get("/api/gate/live-data")
async def live_dashboard_data():
    """Return real-time dashboard data for today."""
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    db = await _get_db()
    try:
        recon = await _reconcile(db, today)
    finally:
        await db.close()

    staff_detail = recon.get("staff_detail", [])
    staff_inside = [s for s in staff_detail if s.get("occupancy_status") == "INSIDE"]
    staff_exited = [s for s in staff_detail if s.get("occupancy_status") == "EXITED"]
    staff_absent = [s for s in staff_detail if s.get("occupancy_status") == "ABSENT"]

    unknown_persons = recon.get("unknown_persons", [])
    unknowns_inside = [u for u in unknown_persons if u.get("status") == "INSIDE"]
    unknowns_exited = [u for u in unknown_persons if u.get("status") == "EXITED"]

    ds = recon.get("data_sources", {})

    # Combine unknown + unreconciled into single "unrecognized" list
    all_unrec = list(unknown_persons) + list(recon.get("unreconciled_gate_entries", []))
    for idx, u in enumerate(all_unrec, 1):
        u["id"] = f"U-{idx:03d}"

    total_entries = recon.get("total_entries", recon.get("unique_gate_in", 0))
    total_exits = recon.get("total_exits", recon.get("unique_gate_out", 0))
    total_recognized = recon.get("total_recognized", len(staff_inside) + len(staff_exited))
    total_unrecognized = recon.get("total_unrecognized", 0)
    current_occupancy = recon.get("current_occupancy", recon.get("total_inside", 0))

    return {
        "timestamp": now.strftime("%d-%m-%Y %H:%M:%S IST"),
        "date": today,
        # Per school spec
        "entry_summary": {
            "total_persons_entered": total_entries,
        },
        "recognition_summary": {
            "recognized_students": recon.get("recognized_students", 0),
            "recognized_staff": recon.get("recognized_staff_count", 0),
            "total_recognized": total_recognized,
        },
        "unrecognized_summary": {
            "total_unrecognized": total_unrecognized,
        },
        "exit_summary": {
            "total_persons_exited": total_exits,
        },
        "current_occupancy": current_occupancy,
        # Legacy occupancy object (kept for backward compat)
        "occupancy": {
            "total_inside": current_occupancy,
            "staff_inside": len(staff_inside),
            "staff_exited": len(staff_exited),
            "staff_absent": len(staff_absent),
            "unknown_inside": len(unknowns_inside),
            "unknown_exited": len(unknowns_exited),
            "total_registered": recon.get("total_registered", 0),
        },
        "gate": {
            "unique_in": recon.get("unique_gate_in", 0),
            "unique_out": recon.get("unique_gate_out", 0),
            "raw_in": recon.get("raw_gate_in", 0),
            "raw_out": recon.get("raw_gate_out", 0),
        },
        "vehicles": {
            "in": recon.get("vehicles_in", 0),
            "out": recon.get("vehicles_out", 0),
            "types": recon.get("vehicle_types", {}),
        },
        "staff_inside": [
            {
                "name": s["name"],
                "category": s.get("category", "Staff"),
                "arrival": s.get("trueface_arrival") or s.get("dvr_first_seen") or "-",
                "cameras": s.get("dvr_cameras", [])[:3],
                "status": s.get("occupancy_status", "INSIDE"),
            }
            for s in staff_inside
        ],
        "staff_exited": [
            {
                "name": s["name"],
                "category": s.get("category", "Staff"),
                "arrival": s.get("trueface_arrival") or s.get("dvr_first_seen") or "-",
                "departure": s.get("trueface_departure") or s.get("dvr_last_seen") or "-",
            }
            for s in staff_exited
        ],
        "unrecognized_persons": [
            {
                "id": u.get("id", ""),
                "time": u.get("time", "-"),
                "camera": u.get("camera", "-"),
                "direction": u.get("direction", "IN"),
                "attire": u.get("attire_color", u.get("outfit_description", "-")),
                "status": "Unrecognized",
                "has_photo": bool(u.get("person_crop")),
            }
            for u in all_unrec[:50]
        ],
        # Legacy (kept for compat)
        "unknown_persons": [
            {
                "id": u.get("id", ""),
                "time": u.get("time", "-"),
                "camera": u.get("camera", "-"),
                "direction": u.get("direction", "IN"),
                "status": u.get("status", "INSIDE"),
                "has_photo": bool(u.get("person_crop")),
            }
            for u in unknown_persons[:50]
        ],
        "reconciliation": {
            "total_entries": total_entries,
            "total_recognized": total_recognized,
            "total_unrecognized": total_unrecognized,
            "rate": round((total_recognized / total_entries * 100) if total_entries > 0 else 100, 1),
        },
        "data_sources": ds,
        "alerts": recon.get("alerts", [])[:10],
        "staff_category_inside": recon.get("staff_category_inside", {}),
    }


_LIVE_SNAPSHOT_CAMERAS = [
    {"key": "ENTRY GATE-1", "label": "Entry Gate 1"},
    {"key": "ENTRY GATE- 2", "label": "Entry Gate 2"},
    {"key": "Basement Main Gate", "label": "Basement Main Gate"},
    {"key": "Reception C1", "label": "Reception C1"},
    {"key": "Reception C2", "label": "Reception C2"},
    {"key": "DISPERSAL EXIT", "label": "Dispersal Exit"},
]


@router.get("/api/gate/live-snapshots")
async def live_snapshot_gallery():
    """Return live DVR camera snapshots + recognized/unrecognized person photos."""
    import base64 as b64mod
    from app.routes.agent_ws import request_snapshot

    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")

    # Check if agent is connected before attempting camera snapshots
    from app.routes.agent_ws import _agent_ws
    agent_online = _agent_ws is not None

    async def _fetch_cam(cam: dict) -> dict:
        if not agent_online:
            return {"camera": cam["label"], "key": cam["key"], "success": False, "photo": ""}
        try:
            result = await request_snapshot(cam["key"], timeout=20.0)
            images = result.get("images", []) if result.get("success") else []
            return {
                "camera": cam["label"],
                "key": cam["key"],
                "success": result.get("success", False),
                "photo": images[0]["image_base64"] if images else "",
            }
        except Exception:
            return {"camera": cam["label"], "key": cam["key"], "success": False, "photo": ""}

    cam_tasks = [_fetch_cam(c) for c in _LIVE_SNAPSHOT_CAMERAS]
    camera_snapshots = await asyncio.gather(*cam_tasks)

    # Recognized staff with face photos
    db = await _get_db()
    try:
        cur = await db.execute(
            "SELECT DISTINCT arf.name, arf.image_data, ta.arrival_time, ta.departure_time "
            "FROM trueface_attendance ta "
            "JOIN agent_registered_faces arf ON arf.name = ta.name AND arf.angle = 'front' "
            "WHERE ta.date = ? ORDER BY ta.arrival_time",
            (today,),
        )
        recognized = []
        seen_names: set[str] = set()
        for r in await cur.fetchall():
            name = r[0]
            if name in seen_names:
                continue
            seen_names.add(name)
            img = r[1]
            if isinstance(img, bytes):
                img = b64mod.b64encode(img).decode()
            recognized.append({
                "name": name, "arrival": r[2] or "-",
                "departure": r[3] or "-", "photo": img or "",
            })

        # Unrecognized — visitor DVR sightings
        cur2 = await db.execute(
            "SELECT id, timestamp, camera, snapshot, direction "
            "FROM visitor_dvr_sightings "
            "WHERE date = ? AND snapshot != '' ORDER BY timestamp DESC LIMIT 30",
            (today,),
        )
        unrecognized = []
        idx = 1
        for r in await cur2.fetchall():
            unrecognized.append({
                "id": f"U-{idx:03d}", "time": r[1], "camera": r[2],
                "photo": r[3] or "", "direction": r[4] if len(r) > 4 and r[4] else "IN",
            })
            idx += 1

        # Gate entries with person_crop
        cur3 = await db.execute(
            "SELECT id, timestamp, camera, direction, person_crop, attire_color "
            "FROM gate_entries WHERE date = ? AND direction = 'IN' "
            "AND person_crop != '' ORDER BY timestamp DESC LIMIT 30",
            (today,),
        )
        for r in await cur3.fetchall():
            unrecognized.append({
                "id": f"U-{idx:03d}", "time": r[1], "camera": r[2],
                "photo": r[4] or "", "direction": r[3], "attire": r[5] or "-",
            })
            idx += 1
    finally:
        await db.close()

    return {
        "timestamp": now.strftime("%d-%m-%Y %H:%M:%S IST"),
        "agent_online": agent_online,
        "camera_snapshots": list(camera_snapshots),
        "recognized": recognized,
        "unrecognized": unrecognized,
        "total_recognized": len(recognized),
        "total_unrecognized": len(unrecognized),
    }


@router.get("/live", response_class=HTMLResponse)
async def live_dashboard_page():
    """Serve the live dashboard HTML page."""
    return _LIVE_DASHBOARD_HTML


_LIVE_DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PPIS Live Campus Monitor</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh}
.header{background:linear-gradient(135deg,#1e3a5f,#2563eb);padding:16px 24px;display:flex;justify-content:space-between;align-items:center;box-shadow:0 2px 8px rgba(0,0,0,.3)}
.header h1{font-size:20px;font-weight:700;color:#fff}
.header .meta{font-size:13px;color:#93c5fd}
.refresh-dot{display:inline-block;width:8px;height:8px;border-radius:50%;background:#22c55e;margin-right:6px;animation:pulse 2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}
.tabs{display:flex;gap:0;padding:0 24px;background:#1e293b;border-bottom:2px solid #334155}
.tab{padding:12px 24px;cursor:pointer;font-size:14px;font-weight:600;color:#94a3b8;border-bottom:3px solid transparent;transition:all .2s}
.tab:hover{color:#e2e8f0}
.tab.active{color:#3b82f6;border-bottom-color:#3b82f6}
.tab-content{display:none}
.tab-content.active{display:block}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;padding:20px 24px}
.card{background:#1e293b;border-radius:12px;padding:20px;border:1px solid #334155;transition:transform .2s}
.card:hover{transform:translateY(-2px)}
.card h2{font-size:14px;text-transform:uppercase;letter-spacing:1px;color:#94a3b8;margin-bottom:12px}
.big-num{font-size:48px;font-weight:800;line-height:1}
.big-num.green{color:#22c55e}
.big-num.amber{color:#f59e0b}
.big-num.red{color:#ef4444}
.big-num.blue{color:#3b82f6}
.sub-stat{font-size:13px;color:#94a3b8;margin-top:6px}
.sub-stat span{color:#e2e8f0;font-weight:600}
.full-width{grid-column:1/-1}
table{width:100%;border-collapse:collapse;font-size:13px}
th{text-align:left;padding:8px 12px;background:#334155;color:#94a3b8;font-weight:600;text-transform:uppercase;font-size:11px;letter-spacing:.5px}
td{padding:8px 12px;border-bottom:1px solid #334155}
tr:hover td{background:#334155}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.badge-in{background:#166534;color:#86efac}
.badge-out{background:#7c2d12;color:#fdba74}
.badge-inside{background:#164e63;color:#67e8f9}
.badge-exited{background:#3f3f46;color:#a1a1aa}
.badge-active{background:#166534;color:#86efac}
.badge-nodata{background:#7f1d1d;color:#fca5a5}
.badge-recognized{background:#166534;color:#86efac}
.badge-unrecognized{background:#7f1d1d;color:#fca5a5}
.alert-row{padding:8px 12px;border-left:3px solid;margin-bottom:4px;border-radius:0 6px 6px 0;font-size:13px}
.alert-high{border-color:#ef4444;background:rgba(239,68,68,.1)}
.alert-medium{border-color:#f59e0b;background:rgba(245,158,11,.1)}
.alert-low{border-color:#3b82f6;background:rgba(59,130,246,.1)}
.progress-bar{height:6px;background:#334155;border-radius:3px;overflow:hidden;margin-top:8px}
.progress-fill{height:100%;border-radius:3px;transition:width .5s}
.source-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px}
.source-item{background:#0f172a;border-radius:8px;padding:12px;text-align:center;border:1px solid #334155}
.source-item .count{font-size:24px;font-weight:700;color:#3b82f6}
.source-item .label{font-size:11px;color:#94a3b8;margin-top:4px}
.photo-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:16px;padding:20px 24px}
.photo-card{background:#1e293b;border-radius:12px;overflow:hidden;border:1px solid #334155;transition:transform .2s}
.photo-card:hover{transform:translateY(-2px)}
.photo-card img{width:100%;height:180px;object-fit:cover;background:#334155}
.photo-card .info{padding:12px}
.photo-card .name{font-size:14px;font-weight:700;color:#e2e8f0;margin-bottom:4px}
.photo-card .detail{font-size:12px;color:#94a3b8}
.photo-card .no-photo{width:100%;height:180px;display:flex;align-items:center;justify-content:center;background:#334155;color:#64748b;font-size:13px}
.section-header{padding:20px 24px 0;display:flex;justify-content:space-between;align-items:center}
.section-header h2{font-size:18px;font-weight:700;color:#e2e8f0}
.section-header .count-badge{font-size:14px;color:#94a3b8;background:#334155;padding:4px 12px;border-radius:20px}
.footer{text-align:center;padding:16px;color:#475569;font-size:12px}
.loading{text-align:center;padding:60px;color:#64748b;font-size:16px}
@media(max-width:768px){.grid{grid-template-columns:1fr;padding:12px}.header{padding:12px 16px}.big-num{font-size:36px}.photo-grid{grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:10px;padding:12px}.photo-card img{height:140px}}
</style>
</head>
<body>
<div class="header">
  <h1>PP International School &mdash; Live Campus Monitor</h1>
  <div class="meta"><span class="refresh-dot"></span>Auto-refresh 30s &bull; <span id="ts">Loading...</span></div>
</div>
<div class="tabs">
  <div class="tab active" onclick="switchTab('overview')">Overview</div>
  <div class="tab" onclick="switchTab('snapshots')">Additional Information</div>
</div>
<div id="tab-overview" class="tab-content active">
  <div id="app" class="loading">Loading dashboard data...</div>
</div>
<div id="tab-snapshots" class="tab-content">
  <div id="snapshots-app" class="loading">Loading snapshots...</div>
</div>
<div class="footer">PPIS Headcount Reconciliation &amp; Entry Monitoring System</div>

<script>
const API = '/api/gate/live-data';
let data = null;

async function fetchData() {
  try {
    const r = await fetch(API);
    data = await r.json();
    render();
  } catch(e) {
    document.getElementById('app').innerHTML = '<div class="loading">Error loading data. Retrying...</div>';
  }
}

function render() {
  if (!data) return;
  document.getElementById('ts').textContent = data.timestamp;
  const o = data.occupancy;
  const g = data.gate;
  const v = data.vehicles;
  const rc = data.reconciliation || {};
  const ds = data.data_sources || {};
  const es = data.entry_summary || {};
  const rs = data.recognition_summary || {};
  const us = data.unrecognized_summary || {};
  const xs = data.exit_summary || {};

  let html = '<div class="grid">';

  // Row 1: Key metrics per school spec
  html += `
    <div class="card">
      <h2>Total Persons Entered</h2>
      <div class="big-num green">${es.total_persons_entered || g.unique_in}</div>
      <div class="sub-stat">IN: <span>${g.unique_in}</span> &bull; OUT: <span>${g.unique_out}</span></div>
    </div>
    <div class="card">
      <h2>Current Occupancy</h2>
      <div class="big-num green">${data.current_occupancy || o.total_inside}</div>
      <div class="sub-stat">Entries - Exits = <span>${es.total_persons_entered || g.unique_in} - ${xs.total_persons_exited || g.unique_out}</span></div>
    </div>
    <div class="card">
      <h2>Total Recognized</h2>
      <div class="big-num blue">${rs.total_recognized || 0}</div>
      <div class="sub-stat">Students: <span>${rs.recognized_students || 0}</span> &bull; Staff: <span>${rs.recognized_staff || 0}</span></div>
    </div>
    <div class="card">
      <h2>Unrecognized Persons</h2>
      <div class="big-num ${(us.total_unrecognized||0) > 0 ? 'red' : 'green'}">${us.total_unrecognized || 0}</div>
      <div class="sub-stat">Entries - Recognized = <span>${es.total_persons_entered || 0} - ${rs.total_recognized || 0}</span></div>
    </div>`;

  // Reconciliation rate
  html += `
    <div class="card">
      <h2>Reconciliation Rate</h2>
      <div class="big-num ${rc.rate >= 90 ? 'green' : rc.rate >= 70 ? 'amber' : 'red'}">${rc.rate || 100}%</div>
      <div class="progress-bar"><div class="progress-fill" style="width:${rc.rate||100}%;background:${(rc.rate||100)>=90?'#22c55e':(rc.rate||100)>=70?'#f59e0b':'#ef4444'}"></div></div>
      <div class="sub-stat">Recognized: <span>${rc.total_recognized||0}</span> / ${rc.total_entries||0} entries</div>
    </div>`;

  // Exit summary
  html += `
    <div class="card">
      <h2>Total Persons Exited</h2>
      <div class="big-num amber">${xs.total_persons_exited || g.unique_out}</div>
      <div class="sub-stat">Staff exited: <span>${o.staff_exited}</span></div>
    </div>`;

  // Vehicles
  const vTypes = Object.entries(v.types||{}).map(([t,c])=>t[0].toUpperCase()+t.slice(1)+': '+c).join(', ');
  html += `
    <div class="card">
      <h2>Vehicles</h2>
      <div class="big-num blue">${v.in + v.out}</div>
      <div class="sub-stat">IN: <span>${v.in}</span> &bull; OUT: <span>${v.out}</span></div>
      ${vTypes ? '<div class="sub-stat">'+vTypes+'</div>' : ''}
    </div>`;

  // Data sources
  html += '<div class="card full-width"><h2>Data Sources</h2><div class="source-grid">';
  for (const key of ['entry_gate','basement','dispersal','trueface','other_dvr']) {
    const src = ds[key];
    if (!src) continue;
    const active = src.sightings > 0;
    html += `<div class="source-item">
      <div class="count">${src.sightings}</div>
      <div class="label">${src.name}</div>
      <div style="margin-top:6px"><span class="badge ${active?'badge-active':'badge-nodata'}">${active?'ACTIVE':'NO DATA'}</span></div>
      ${src.unique_persons !== undefined ? '<div class="label">'+src.unique_persons+' unique</div>' : ''}
    </div>`;
  }
  html += '</div></div>';

  // Staff category breakdown
  const cats = data.staff_category_inside || {};
  const catEntries = Object.entries(cats);
  if (catEntries.length > 0) {
    html += '<div class="card full-width"><h2>Recognized Staff by Category</h2><div class="source-grid">';
    for (const [cat, cnt] of catEntries.sort()) {
      html += `<div class="source-item"><div class="count">${cnt}</div><div class="label">${cat}</div></div>`;
    }
    html += '</div></div>';
  }

  // Unrecognized persons table (combined)
  const unrec = data.unrecognized_persons || [];
  if (unrec.length > 0) {
    html += `<div class="card full-width"><h2>Unrecognized Persons (${unrec.length})</h2>
      <div class="sub-stat" style="margin-bottom:8px">Persons detected entering but NOT recognized as Student or Staff</div>
      <table>
      <tr><th>ID</th><th>Time</th><th>Entry Point</th><th>Appearance</th><th>Direction</th><th>Status</th></tr>`;
    for (const u of unrec) {
      html += `<tr>
        <td><strong>${u.id}</strong></td><td>${u.time}</td><td>${u.camera}</td>
        <td>${u.attire||'-'}</td>
        <td><span class="badge ${u.direction==='IN'?'badge-in':'badge-out'}">${u.direction}</span></td>
        <td><span class="badge badge-inside">Unrecognized</span></td>
      </tr>`;
    }
    html += '</table></div>';
  }

  // Staff inside table
  const staffIn = data.staff_inside || [];
  if (staffIn.length > 0) {
    html += `<div class="card full-width"><h2>Recognized Staff Inside (${staffIn.length})</h2><table>
      <tr><th>#</th><th>Name</th><th>Category</th><th>Arrived</th><th>Cameras</th></tr>`;
    staffIn.forEach((s,i) => {
      html += `<tr><td>${i+1}</td><td>${s.name}</td><td>${s.category}</td><td>${s.arrival}</td><td>${(s.cameras||[]).join(', ')||'-'}</td></tr>`;
    });
    html += '</table></div>';
  }

  // Staff exited table
  const staffOut = data.staff_exited || [];
  if (staffOut.length > 0) {
    html += `<div class="card full-width"><h2>Recognized Staff Exited (${staffOut.length})</h2><table>
      <tr><th>#</th><th>Name</th><th>Category</th><th>Arrived</th><th>Departed</th></tr>`;
    staffOut.forEach((s,i) => {
      html += `<tr><td>${i+1}</td><td>${s.name}</td><td>${s.category}</td><td>${s.arrival}</td><td>${s.departure}</td></tr>`;
    });
    html += '</table></div>';
  }

  // Alerts
  const alerts = data.alerts || [];
  if (alerts.length > 0) {
    html += '<div class="card full-width"><h2>Alerts</h2>';
    for (const a of alerts) {
      const sev = a.severity || 'medium';
      html += `<div class="alert-row alert-${sev}"><strong>${a.type||''}</strong> ${a.detail||''}</div>`;
    }
    html += '</div>';
  }

  html += '</div>';
  document.getElementById('app').className = '';
  document.getElementById('app').innerHTML = html;
}

// Tab switching
let activeTab = 'overview';
let snapshotsLoaded = false;
let snapshotsLoading = false;

function switchTab(tab) {
  activeTab = tab;
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
  document.querySelector(`.tab[onclick="switchTab('${tab}')"]`).classList.add('active');
  document.getElementById('tab-' + tab).classList.add('active');
  if (tab === 'snapshots' && !snapshotsLoaded && !snapshotsLoading) {
    loadSnapshots();
  }
}

async function loadSnapshots() {
  snapshotsLoading = true;
  const el = document.getElementById('snapshots-app');
  el.innerHTML = '<div class="loading">Fetching live camera snapshots... This may take 15-20 seconds.</div>';
  try {
    const r = await fetch('/api/gate/live-snapshots');
    const d = await r.json();
    snapshotsLoaded = true;
    renderSnapshots(d);
  } catch(e) {
    el.innerHTML = '<div class="loading">Error loading snapshots. <a href="#" onclick="loadSnapshots();return false" style="color:#3b82f6">Retry</a></div>';
  }
  snapshotsLoading = false;
}

function renderSnapshots(d) {
  const el = document.getElementById('snapshots-app');
  let html = '';

  // Agent status banner
  if (!d.agent_online) {
    html += '<div style="padding:12px 24px;background:rgba(239,68,68,.15);border-left:4px solid #ef4444;margin:16px 24px 0;border-radius:0 8px 8px 0;font-size:14px;color:#fca5a5">Campus agent is offline — live camera snapshots unavailable. Showing stored person data only.</div>';
  }

  // Live Camera Feeds
  const cams = d.camera_snapshots || [];
  html += '<div class="section-header"><h2>Live Camera Feeds</h2><span class="count-badge">' + cams.length + ' cameras</span></div>';
  html += '<div class="photo-grid">';
  for (const c of cams) {
    if (c.photo) {
      html += `<div class="photo-card">
        <img src="data:image/jpeg;base64,${c.photo}" alt="${c.camera}" loading="lazy">
        <div class="info"><div class="name">${c.camera}</div>
        <div class="detail">Live snapshot &bull; ${d.timestamp}</div></div></div>`;
    } else {
      html += `<div class="photo-card">
        <div class="no-photo">${c.success === false ? 'Camera offline' : 'No image'}</div>
        <div class="info"><div class="name">${c.camera}</div>
        <div class="detail">${c.success === false ? 'Agent not connected' : 'No data'}</div></div></div>`;
    }
  }
  html += '</div>';

  // Recognized Staff
  const rec = d.recognized || [];
  html += '<div class="section-header" style="margin-top:16px"><h2>Recognized Staff Today</h2><span class="count-badge badge-recognized">' + rec.length + ' recognized</span></div>';
  if (rec.length > 0) {
    html += '<div class="photo-grid">';
    for (const p of rec) {
      if (p.photo) {
        html += `<div class="photo-card">
          <img src="data:image/jpeg;base64,${p.photo}" alt="${p.name}" loading="lazy">
          <div class="info"><div class="name">${p.name}</div>
          <div class="detail">Arrived: ${p.arrival}</div>
          ${p.departure !== '-' ? '<div class="detail">Left: ' + p.departure + '</div>' : ''}</div></div>`;
      } else {
        html += `<div class="photo-card">
          <div class="no-photo">No photo</div>
          <div class="info"><div class="name">${p.name}</div>
          <div class="detail">Arrived: ${p.arrival}</div></div></div>`;
      }
    }
    html += '</div>';
  } else {
    html += '<div style="padding:20px 24px;color:#64748b">No recognized staff data for today.</div>';
  }

  // Unrecognized Persons
  const unrec = d.unrecognized || [];
  html += '<div class="section-header" style="margin-top:16px"><h2>Unrecognized Persons</h2><span class="count-badge badge-unrecognized">' + unrec.length + ' unrecognized</span></div>';
  if (unrec.length > 0) {
    html += '<div class="photo-grid">';
    for (const u of unrec) {
      if (u.photo) {
        html += `<div class="photo-card">
          <img src="data:image/jpeg;base64,${u.photo}" alt="${u.id}" loading="lazy">
          <div class="info"><div class="name">${u.id}</div>
          <div class="detail">${u.time} &bull; ${u.camera}</div>
          <div class="detail"><span class="badge ${u.direction==='IN'?'badge-in':'badge-out'}">${u.direction}</span>
          ${u.attire ? ' &bull; ' + u.attire : ''}</div></div></div>`;
      } else {
        html += `<div class="photo-card">
          <div class="no-photo">No snapshot</div>
          <div class="info"><div class="name">${u.id}</div>
          <div class="detail">${u.time} &bull; ${u.camera}</div></div></div>`;
      }
    }
    html += '</div>';
  } else {
    html += '<div style="padding:20px 24px;color:#64748b">No unrecognized persons with snapshots for today.</div>';
  }

  // Refresh button
  html += '<div style="text-align:center;padding:24px"><button onclick="snapshotsLoaded=false;loadSnapshots()" style="background:#2563eb;color:#fff;border:none;padding:10px 24px;border-radius:8px;font-size:14px;cursor:pointer;font-weight:600">Refresh Snapshots</button></div>';

  el.className = '';
  el.innerHTML = html;
}

fetchData();
setInterval(fetchData, 30000);
</script>
</body>
</html>"""
